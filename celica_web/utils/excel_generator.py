"""
Générateur de templates Excel pour l'import de tests
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

class ExcelTemplateGenerator:
    """Générateur de templates Excel pour l'import de tests"""
    
    def __init__(self):
        self.wb = Workbook()
        self._setup_styles()
    
    def _setup_styles(self):
        """Configure les styles pour le template"""
        # Supprimer la feuille par défaut
        if self.wb.active:
            self.wb.remove(self.wb.active)
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.info_font = Font(bold=True, color="000000")
        self.info_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_test_template(self) -> BytesIO:
        """Génère un template Excel complet pour l'import de tests"""
        # Créer l'onglet Test
        self._create_test_sheet()
        
        # Créer l'onglet Questions
        self._create_questions_sheet()
        
        # Créer l'onglet Instructions
        self._create_instructions_sheet()
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_test_sheet(self):
        """Crée l'onglet Test avec les informations générales"""
        ws = self.wb.create_sheet("Test")
        
        # Titre de l'onglet
        ws['A1'] = "INFORMATIONS DU TEST"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:B1')
        
        # Données du test
        test_data = [
            ("Titre du test", "Ex: Test de culture générale"),
            ("Description", "Ex: Test pour évaluer les connaissances générales"),
            ("Durée (minutes)", "30"),
            ("Barème total", "20"),
            ("Mélanger les questions", "Oui")
        ]
        
        for i, (label, value) in enumerate(test_data, 3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = self.info_font
            ws[f'A{i}'].fill = self.info_fill
            ws[f'A{i}'].border = self.border
            
            ws[f'B{i}'] = value
            ws[f'B{i}'].border = self.border
            
            # Validation pour certains champs
            if "Durée" in label:
                dv = DataValidation(type="whole", operator="between", formula1="1", formula2="180")
                dv.add(f'B{i}')
                ws.add_data_validation(dv)
            elif "Barème" in label:
                dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
                dv.add(f'B{i}')
                ws.add_data_validation(dv)
            elif "Mélanger" in label:
                dv = DataValidation(type="list", formula1='"Oui,Non"')
                dv.add(f'B{i}')
                ws.add_data_validation(dv)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_questions_sheet(self):
        """Crée l'onglet Questions avec la structure des questions"""
        ws = self.wb.create_sheet("Questions")
        
        # En-têtes
        headers = [
            "Type", "Énoncé", "Niveau", "Points", 
            "Réponse 1", "Réponse 2", "Réponse 3", "Réponse 4", "Réponse 5", 
            "Explication"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Exemples de données
        example_data = [
            ["QCM", "Quelle est la capitale de la France?", "facile", "1", "✓Paris", "Londres", "Berlin", "Madrid", "", "Paris est la capitale de la France"],
            ["QCM", "Combien de côtés a un hexagone?", "moyen", "2", "4", "5", "✓6", "7", "", "Un hexagone a 6 côtés"],
            ["QRL", "Citez deux planètes du système solaire", "difficile", "3", "✓Mercure", "✓Vénus", "✓Terre", "✓Mars", "✓Jupiter", "Il y a 8 planètes dans le système solaire"]
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Validations
        self._add_question_validations(ws)
        
        # Ajuster les largeurs des colonnes
        column_widths = [8, 40, 10, 8, 20, 20, 20, 20, 20, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
    
    def _add_question_validations(self, ws: Worksheet):
        """Ajoute les validations de données pour l'onglet Questions"""
        # Validation pour le type de question
        type_dv = DataValidation(type="list", formula1='"QCM,QRL"')
        type_dv.add('A2:A100')
        ws.add_data_validation(type_dv)
        
        # Validation pour le niveau
        niveau_dv = DataValidation(type="list", formula1='"facile,moyen,difficile"')
        niveau_dv.add('C2:C100')
        ws.add_data_validation(niveau_dv)
        
        # Validation pour les points
        points_dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
        points_dv.add('D2:D100')
        ws.add_data_validation(points_dv)
    
    def _create_instructions_sheet(self):
        """Crée l'onglet Instructions avec les explications"""
        ws = self.wb.create_sheet("Instructions")
        
        # Titre
        ws['A1'] = "INSTRUCTIONS D'UTILISATION"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:C1')
        
        instructions = [
            ("Onglet Test", "Remplissez les informations générales du test"),
            ("Onglet Questions", "Ajoutez vos questions avec leurs réponses"),
            ("", ""),
            ("Types de questions:", ""),
            ("- QCM", "Question à choix multiples (une seule réponse correcte)"),
            ("- QRL", "Question à réponses libres (plusieurs réponses possibles)"),
            ("", ""),
            ("Niveaux de difficulté:", ""),
            ("- facile", "Question simple"),
            ("- moyen", "Question de difficulté moyenne"),
            ("- difficile", "Question complexe"),
            ("", ""),
            ("Réponses correctes:", ""),
            ("- Pour les QCM", "Préfixez la réponse correcte avec ✓"),
            ("- Pour les QRL", "Préfixez toutes les réponses correctes avec ✓"),
            ("", ""),
            ("Exemple QCM:", ""),
            ("Type: QCM", "Énoncé: Quelle est la capitale de la France?"),
            ("Réponse 1: ✓Paris", "Réponse 2: Londres"),
            ("Réponse 3: Berlin", "Réponse 4: Madrid"),
            ("", ""),
            ("Exemple QRL:", ""),
            ("Type: QRL", "Énoncé: Citez deux planètes du système solaire"),
            ("Réponse 1: ✓Mercure", "Réponse 2: ✓Vénus"),
            ("Réponse 3: ✓Terre", "Réponse 4: Mars (incorrecte)"),
        ]
        
        for i, (label, value) in enumerate(instructions, 3):
            ws[f'A{i}'] = label
            if label and not label.startswith("-"):
                ws[f'A{i}'].font = Font(bold=True)
            
            if value:
                ws[f'B{i}'] = value
                ws.merge_cells(f'B{i}:C{i}')
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
    
    def generate_example_template(self) -> BytesIO:
        """Génère un template avec des exemples de données"""
        # Réinitialiser le workbook pour ajouter des exemples
        self.wb = Workbook()
        if self.wb.active:
            self.wb.remove(self.wb.active)
        self._setup_styles()
        
        # Créer les onglets avec des exemples
        self._create_test_sheet_with_example()
        self._create_questions_sheet_with_examples()
        self._create_instructions_sheet()
        
        # Sauvegarder
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_test_sheet_with_example(self):
        """Crée l'onglet Test avec un exemple complet"""
        ws = self.wb.create_sheet("Test")
        
        # Titre
        ws['A1'] = "INFORMATIONS DU TEST"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:B1')
        
        # Exemple de données
        test_data = [
            ("Titre du test", "Test de Culture Générale"),
            ("Description", "Test pour évaluer les connaissances générales en géographie, histoire et sciences"),
            ("Durée (minutes)", "45"),
            ("Barème total", "25"),
            ("Mélanger les questions", "Oui")
        ]
        
        for i, (label, value) in enumerate(test_data, 3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = self.info_font
            ws[f'A{i}'].fill = self.info_fill
            ws[f'A{i}'].border = self.border
            
            ws[f'B{i}'] = value
            ws[f'B{i}'].border = self.border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_questions_sheet_with_examples(self):
        """Crée l'onglet Questions avec des exemples complets"""
        ws = self.wb.create_sheet("Questions")
        
        # En-têtes
        headers = [
            "Type", "Énoncé", "Niveau", "Points", 
            "Réponse 1", "Réponse 2", "Réponse 3", "Réponse 4", "Réponse 5", 
            "Explication"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Exemples de questions
        examples = [
            ["QCM", "Quelle est la capitale de la France?", "facile", "1", "✓Paris", "Londres", "Berlin", "Madrid", "", "Paris est la capitale de la France depuis le 12ème siècle"],
            ["QCM", "Combien de côtés a un hexagone?", "moyen", "2", "4", "5", "✓6", "7", "", "Un hexagone est un polygone à 6 côtés"],
            ["QCM", "Quel est le plus grand océan du monde?", "facile", "1", "Atlantique", "✓Pacifique", "Indien", "Arctique", "", "Le Pacifique couvre environ 1/3 de la surface terrestre"],
            ["QRL", "Citez deux planètes du système solaire", "difficile", "3", "✓Mercure", "✓Vénus", "✓Terre", "✓Mars", "✓Jupiter", "Il y a 8 planètes dans notre système solaire"],
            ["QRL", "Nommez trois pays d'Europe", "moyen", "2", "✓France", "✓Allemagne", "✓Italie", "✓Espagne", "✓Royaume-Uni", "L'Europe compte 44 pays"],
            ["QCM", "Quel est le symbole chimique de l'or?", "moyen", "2", "Ag", "✓Au", "Fe", "Cu", "", "Au vient du latin 'aurum' qui signifie or"]
        ]
        
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Ajouter les validations
        self._add_question_validations(ws)
        
        # Ajuster les largeurs
        column_widths = [8, 40, 10, 8, 20, 20, 20, 20, 20, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width 