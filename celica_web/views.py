from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.urls import reverse
from django.contrib.auth.models import Group
from django.db import models
from django.db.models import Count, Avg, Sum, Q
from django.core.serializers.json import DjangoJSONEncoder
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import pandas as pd
import zipfile
import random
import json
import os
from .models import (
    Cours, Utilisateur, Test, Question, Reponse, Module, Notification, Planning, Groupe,
    Aide, APropos, Resultat, Statistiques, SecurityViolation, TestEventLog, generer_ou_maj_statistiques
)
from .forms import (
    CoursForm, TestForm, QuestionForm, ReponseFormSet, ImportForm, ImportCoursForm,
    UtilisateurForm, ChangerMotDePasseForm, PlanningForm, GroupeForm, NotificationForm, ResultatForm,
    ModuleForm, LoginForm, ManualQuestionForm, SelectQuestionForm, CustomPasswordResetForm
)
from django.db import OperationalError, IntegrityError, transaction
from .utils.tirage_aleatoire import effectuer_tirage_aleatoire_questions, ajouter_questions_tirage_au_test
from .utils.deduplication import creer_question_sans_doublon, ajouter_reponses_sans_doublon, verifier_question_existante
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt
import logging
from django.template.loader import render_to_string
import uuid

def send_test_violation_notification(security_violation):
    """
    Envoie une notification ciblée à l'instructeur du test et à l'administrateur
    quand une violation de sécurité est détectée
    """
    try:
        apprenant = security_violation.utilisateur
        violation_type = security_violation.violation_type
        violation_text = security_violation.violation
        
        # Identifier le test en cours depuis l'URL
        test_id = None
        test = None
        instructeur_test = None
        
        try:
            if '/test/passer/' in security_violation.url:
                # Extraire l'ID du test de l'URL
                url_parts = security_violation.url.split('/test/passer/')
                if len(url_parts) > 1:
                    test_id_part = url_parts[1].split('/')[0]
                    if test_id_part.isdigit():
                        test_id = int(test_id_part)
                        test = Test.objects.get(id=test_id)
                        instructeur_test = test.instructeur
        except Exception as e:
            print(f"⚠️ Impossible d'identifier le test: {str(e)}")
        
        # Récupérer l'administrateur principal
        admin = Utilisateur.objects.filter(role='admin', statut='actif').first()
        
        # Créer le message de notification
        if test and instructeur_test:
            message = f"""
🚨 VIOLATION DE SÉCURITÉ - TEST INTERROMPU

Apprenant: {apprenant.get_full_name()} ({apprenant.email})
Test: {test.titre} (ID: {test.id})
Type de violation: {violation_type}
Action détectée: {violation_text}
Date/Heure: {security_violation.timestamp.strftime('%d/%m/%Y %H:%M:%S')}

Le test a été automatiquement interrompu en raison de violations répétées.
Vous pouvez reprogrammer le test pour cet apprenant si nécessaire.

Actions possibles:
• Reprogrammer le test
• Contacter l'apprenant
• Consulter les détails complets
        """.strip()
        else:
            message = f"""
🚨 VIOLATION DE SÉCURITÉ DÉTECTÉE

Apprenant: {apprenant.get_full_name()} ({apprenant.email})
Type de violation: {violation_type}
Action détectée: {violation_text}
Date/Heure: {security_violation.timestamp.strftime('%d/%m/%Y %H:%M:%S')}
URL: {security_violation.url}

Test non identifié - vérifiez manuellement.
        """.strip()
        
        # Envoyer à l'instructeur du test (si identifié)
        if instructeur_test:
            Notification.objects.create(
                titre="🚨 Test interrompu - Violation de sécurité",
                message=message,
                type_notice='urgence',
                priorite='haute',
                utilisateur=instructeur_test,
                date_expiration=timezone.now() + timezone.timedelta(days=7)
            )
            print(f"📧 Notification envoyée à l'instructeur: {instructeur_test.email}")
        
        # Envoyer à l'administrateur
        if admin:
            Notification.objects.create(
                titre="🚨 Violation de sécurité - Test interrompu",
                message=message,
                type_notice='urgence',
                priorite='haute',
                utilisateur=admin,
                date_expiration=timezone.now() + timezone.timedelta(days=7)
            )
            print(f"📧 Notification envoyée à l'administrateur: {admin.email}")
        
        print(f"✅ Notifications envoyées (Instructeur: {instructeur_test is not None}, Admin: {admin is not None})")
        
    except Exception as e:
        print(f"❌ Erreur lors de l'envoi des notifications: {str(e)}")

@login_required
def home(request):
    options = []
    user = request.user

    if user.has_perm('celica_web.gerer_cours'):
        options.append({'name': 'Gérer les cours', 'url': 'course_list'})
    if user.has_perm('celica_web.gerer_tests'):
        options.append({'name': 'Gérer les tests', 'url': 'test_list'})
    if user.has_perm('celica_web.gerer_notifications'):
        options.append({'name': 'Gérer les notifications', 'url': 'notification_list'})
    if user.has_perm('celica_web.gerer_modules'):
        options.append({'name': 'Gérer les modules', 'url': 'module_list'})
    if user.has_perm('celica_web.gerer_plannings'):
        options.append({'name': 'Gérer les plannings', 'url': 'planning_list'})

    return render(request, 'celica_web/home.html', {'options': options})

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from .forms import LoginForm
from .models import Groupe

@csrf_protect
@never_cache
def login_view(request):
    if request.user.is_authenticated:
        # Si l'utilisateur est déjà connecté, rediriger selon son rôle
        return redirect_by_role(request.user)
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            mot_de_passe = form.cleaned_data['mot_de_passe']
            
            # Authentification
            user = authenticate(request, username=email, password=mot_de_passe)
            
            if user is not None:
                if user.statut != 'actif':
                    messages.error(request, "Votre compte n'est pas encore activé. Veuillez attendre la validation par l'administrateur.")
                    return render(request, 'celicaweb/login.html', {'form': form})
                login(request, user)
                if user.doit_changer_mot_de_passe:
                    messages.info(request, "Vous devez changer votre mot de passe avant d'accéder à l'application.")
                    return redirect('celica_web:changer_mot_de_passe', user_id=user.id)
                messages.success(request, f"Bienvenue {user.first_name or user.last_name}!")
                
                # Gestion des groupes pour les instructeurs
                if user.role == 'instructeur':
                    ensure_instructor_group(user)
                
                # Redirection selon le rôle
                return redirect_by_role(user)
            else:
                messages.error(request, "Échec de l'authentification. Veuillez réessayer.")
        else:
            # Les erreurs du formulaire seront affichées automatiquement
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = LoginForm()
    
    return render(request, 'celicaweb/login.html', {'form': form})

def redirect_by_role(user):
    """Redirige l'utilisateur selon son rôle"""
    role_redirects = {
        'apprenant': 'celica_web:apprenant_dashboard',
        'instructeur': 'celica_web:instructeur_dashboard', 
        'admin': 'celica_web:admin_dashboard',
    }
    redirect_url = role_redirects.get(user.role, 'celica_web:admin_dashboard')
    return redirect(redirect_url)

def ensure_instructor_group(user):
    """S'assure que l'instructeur est dans le bon groupe"""
    try:
        instructeur_group, created = Groupe.objects.get_or_create(
            nom='Instructeurs',
            defaults={
                'description': 'Groupe des instructeurs CELICA',
                'capacite_max': 50
            }
        )
        
        if created:
            messages.info(user.request if hasattr(user, 'request') else None, 
                         "Groupe 'Instructeurs' créé automatiquement.")
        
        # Ajouter l'utilisateur au groupe s'il n'y est pas déjà
        if not user.groupes.filter(nom='Instructeurs').exists():
            user.groupes.add(instructeur_group)
            
    except Exception as e:
        pass  # Ignorer les erreurs lors de l'ajout au groupe

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def cours_form_new(request):
    if request.method == 'POST':
        form = CoursForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Le cours a été créé avec succès.")
            return redirect('celica_web:cours_list')
    else:
        form = CoursForm()
    return render(request, 'celicaweb/importer_cours.html', {'form': form})

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def cours_form_edit(request, cours_id):
    cours = get_object_or_404(Cours, id=cours_id)
    if request.method == 'POST':
        form = CoursForm(request.POST, request.FILES, instance=cours)
        if form.is_valid():
            form.save()
            messages.success(request, "Le cours a été modifié avec succès.")
            return redirect('celica_web:cours_list')
    else:
        form = CoursForm(instance=cours)
    return render(request, 'celicaweb/importer_cours.html', {'form': form, 'cours': cours})

def visitor_index(request):
    if request.user.is_authenticated:
        return redirect_by_role(request.user)
    return render(request, 'celicaweb/visitor_index.html')

@login_required
def index(request):
    user = request.user
    options = []
    if user.is_superuser:
        options = [
            {"text": "Gérer les utilisateurs", "url": "celica_web:gerer_utilisateurs"},
            {"text": "Gérer les modules", "url": "celica_web:module_list"},
            {"text": "Gérer les cours", "url": "celica_web:cours_list"},
            {"text": "Gérer les tests", "url": "celica_web:test_list"},
            {"text": "Gérer les notifications", "url": "celica_web:notification_list"},
            {"text": "Voir les statistiques", "url": "celica_web:consulter_statistiques"},
        ]
    elif hasattr(user, 'role') and user.role == 'instructeur':
        options = [
            {"text": "Gérer les cours", "url": "celica_web:cours_list"},
            {"text": "Gérer les tests", "url": "celica_web:test_list"},
            {"text": "Gérer les notifications", "url": "celica_web:notification_list"},
        ]
    else:
        options = [
            {"text": "Voir les cours", "url": "celica_web:mes_cours"},
            {"text": "Voir mes résultats", "url": "celica_web:mes_resultats"},
            {"text": "Passer un test", "url": "celica_web:test_list"},
        ]
    return render(request, 'celicaweb/index.html', {'options': options})

def logout_view(request):
    logout(request)
    return redirect('celica_web:visitor_index')

def custom_password_reset(request):
    """Vue personnalisée pour la réinitialisation de mot de passe"""
    if request.user.is_authenticated:
        return redirect('celica_web:index')
    
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            nouveau_mot_de_passe = form.cleaned_data.get('nouveau_mot_de_passe')
            confirmer_mot_de_passe = form.cleaned_data.get('confirmer_mot_de_passe')
            
            # Vérifier si l'email existe dans la base
            try:
                utilisateur = Utilisateur.objects.get(email=email)
                
                # Si l'utilisateur existe et que les mots de passe sont fournis
                if nouveau_mot_de_passe and confirmer_mot_de_passe:
                    if nouveau_mot_de_passe == confirmer_mot_de_passe:
                        # Changer le mot de passe
                        utilisateur.set_password(nouveau_mot_de_passe)
                        utilisateur.save()
                        messages.success(request, "Votre mot de passe a été réinitialisé avec succès. Vous pouvez maintenant vous connecter.")
                        return redirect('celica_web:login')
                    else:
                        messages.error(request, "Les mots de passe ne correspondent pas.")
                        # Réafficher le formulaire avec l'email pré-rempli
                        form.initial['email'] = email
                        return render(request, 'celicaweb/custom_password_reset.html', {
                            'form': form,
                            'email_verified': True
                        })
                else:
                    # Email vérifié, afficher les champs de mot de passe
                    return render(request, 'celicaweb/custom_password_reset.html', {
                        'form': form,
                        'email_verified': True
                    })
                    
            except Utilisateur.DoesNotExist:
                messages.error(request, "Cette adresse email n'est pas enregistrée dans notre base de données.")
                return render(request, 'celicaweb/custom_password_reset.html', {
                    'form': form,
                    'email_verified': False
                })
    else:
        form = CustomPasswordResetForm()
    
    return render(request, 'celicaweb/custom_password_reset.html', {
        'form': form,
        'email_verified': False
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.db import connection  # AJOUT MANQUANT
from django.http import JsonResponse  # AJOUT POUR AJAX
from .models import Test, Module, Question, Planning, Reponse
from .forms import TestForm, ManualQuestionForm, ReponseFormSet, ImportForm
import pandas as pd
import logging
import json  # AJOUT MANQUANT
from django.utils import timezone as django_timezone

# Configurer le logging
logger = logging.getLogger(__name__)

def process_import_file(uploaded_file, import_format, test_instance, user):
    """
    FONCTION MANQUANTE - Traite l'importation de questions depuis un fichier
    """
    questions_imported = 0
    
    try:
        if import_format == 'csv':
            df = pd.read_csv(uploaded_file)
        elif import_format == 'excel':
            df = pd.read_excel(uploaded_file)
        else:
            raise ValueError("Format de fichier non supporté")
        
        # Colonnes attendues: enonce, reponse1, reponse2, reponse3, reponse4, correct
        required_columns = ['enonce', 'reponse1', 'reponse2', 'correct']
        
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Colonnes manquantes. Colonnes requises: {', '.join(required_columns)}")
        
        for index, row in df.iterrows():
            if pd.isna(row['enonce']) or not row['enonce'].strip():
                continue
                
            # Créer la question
            question = Question.objects.create(
                enonce=row['enonce'].strip(),
                test=test_instance,
                module=test_instance.module,
                instructeur=user,
                type_question='qcm'
            )
            
            # Créer les réponses
            reponses_data = []
            for i in range(1, 5):  # reponse1 à reponse4
                col_name = f'reponse{i}'
                if col_name in df.columns and not pd.isna(row[col_name]):
                    reponse_text = str(row[col_name]).strip()
                    if reponse_text:
                        reponses_data.append({
                            'texte': reponse_text,
                            'est_correcte': str(row['correct']).strip() == str(i)
                        })
            
            # Sauvegarder les réponses
            for reponse_data in reponses_data:
                Reponse.objects.create(
                    question=question,
                    texte=reponse_data['texte'],
                    est_correcte=reponse_data['est_correcte']
                )
            
            questions_imported += 1
            
    except Exception as e:
        logger.error(f"Erreur lors de l'importation: {str(e)}")
        raise e
    
    return questions_imported
import json
import json
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from .models import Test, Module, Question, Reponse  # Ajout des imports manquants

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def test_form(request, test_id=None):
    """Vue complète pour créer/modifier un test avec questions dynamiques"""
    test = get_object_or_404(Test, id=test_id) if test_id else None
    modules = Module.objects.all()
    questions_existantes = []
    
    if test:
        # Recharger le test depuis la base pour s'assurer que les données sont à jour
        test.refresh_from_db()
        
        # Charger les questions du test avec leurs réponses
        questions_existantes = Question.objects.filter(test=test).prefetch_related('reponses')
        for question in questions_existantes:
            question.reponses_count = question.reponses.count()
        
        # Calculer le statut de la pondération
        est_valide, somme_totale, barème, message = valider_ponderation_test(test)
        test.ponderation_status = {
            'est_valide': est_valide,
            'somme_totale': somme_totale,
            'barème': barème,
            'message': message,
            'points_disponibles': barème - somme_totale
        }
    
    if request.method == 'POST':
        action = request.POST.get('action', 'save_test')
        
        # LIGNES DE DEBUG - Ajoutez ceci temporairement
        
        # === SAUVEGARDE DU TEST ===
        if action == 'save_test':
            try:
                titre = request.POST.get('titre', '').strip()
                description = request.POST.get('description', '').strip()
                module_id = request.POST.get('module')
                duree = request.POST.get('duree', '').strip()
                bareme = request.POST.get('bareme', '').strip()
                
                # Validation
                if not titre:
                    messages.error(request, "Le titre est obligatoire")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                if not module_id:
                    messages.error(request, "Le module est obligatoire")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                # Validation de la durée
                if not duree:
                    messages.error(request, "La durée est obligatoire et doit être un nombre positif.")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                try:
                    duree_int = int(duree)
                    if duree_int <= 0:
                        raise ValueError("La durée doit être positive")
                except ValueError:
                    messages.error(request, "La durée doit être un nombre entier positif.")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                # Validation du barème
                if not bareme:
                    messages.error(request, "Le barème est obligatoire et doit être un nombre positif.")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                try:
                    bareme_float = float(bareme)
                    if bareme_float <= 0:
                        raise ValueError("Le barème doit être positif")
                except ValueError:
                    messages.error(request, "Le barème doit être un nombre positif valide.")
                    return render(request, 'celicaweb/test_form.html', {
                        'test': test, 'modules': modules, 'questions_existantes': questions_existantes
                    })
                
                module = get_object_or_404(Module, id=module_id)
                
                # Créer ou modifier le test
                if test:
                    test.titre = titre
                    test.description = description
                    test.module = module
                    test.duree = duree_int
                    test.bareme = bareme_float
                    test.save()
                    messages.success(request, "Test modifié avec succès! Vous pouvez maintenant ajouter des questions.")
                else:
                    test = Test.objects.create(
                        titre=titre,
                        description=description,
                        module=module,
                        duree=duree_int,
                        bareme=bareme_float,
                        instructeur=request.user,
                        date_creation=timezone.now()
                    )
                    messages.success(request, "Test créé avec succès! Vous pouvez maintenant ajouter des questions.")
                
                # Recharger le test depuis la base pour s'assurer que les données sont à jour
                test.refresh_from_db()
                
                return redirect('celica_web:test_form_edit', test_id=test.id)
                
            except Exception as e:
                messages.error(request, f"Erreur lors de la sauvegarde: {str(e)}")
        
        # === AJOUT D'UNE QUESTION DYNAMIQUE ===
        elif action == 'add_question' and test:
            return ajouter_question_dynamique(request, test)
        
        # === IMPORT DE QUESTIONS DEPUIS FICHIER ===
        elif action == 'import_questions' and test:
            try:
                # print("DEBUG: Traitement de l'import en cours...")  # Debug
                
                # Vérifier si un fichier a été uploadé
                if 'fichier_questions' not in request.FILES:
                    messages.error(request, "Aucun fichier sélectionné pour l'import.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                fichier = request.FILES['fichier_questions']
                format_fichier = request.POST.get('format_fichier', 'auto')
                valider_import = request.POST.get('valider_import') == 'on'
                ignorer_erreurs = request.POST.get('ignorer_erreurs') == 'on'
                
                # print(f"DEBUG: Fichier reçu = {fichier.name}, taille = {fichier.size}")  # Debug
                
                # Validation du fichier
                if fichier.size > 10 * 1024 * 1024:  # 10 MB
                    messages.error(request, "Le fichier est trop volumineux (maximum 10 MB).")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Extensions autorisées
                nom_fichier = fichier.name.lower()
                extensions_autorisees = ['.csv', '.txt', '.json']
                if not any(nom_fichier.endswith(ext) for ext in extensions_autorisees):
                    messages.error(request, "Format de fichier non supporté. Utilisez CSV, TXT ou JSON.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Lire le contenu du fichier
                try:
                    contenu = fichier.read().decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        contenu = fichier.read().decode('latin-1')
                    except:
                        messages.error(request, "Impossible de lire le fichier. Vérifiez l'encodage (UTF-8 recommandé).")
                        return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # print(f"DEBUG: Contenu lu, longueur = {len(contenu)} caractères")  # Debug
                
                # Parser le contenu selon le format
                try:
                    if nom_fichier.endswith('.json'):
                        questions_data = json.loads(contenu)
                    else:
                        questions_data = parser_fichier_texte(contenu)
                    
                    if not questions_data:
                        messages.error(request, "Aucune question valide trouvée dans le fichier.")
                        return redirect('celica_web:test_form_edit', test_id=test.id)
                    
                    # print(f"DEBUG: {len(questions_data)} questions parsées")  # Debug
                    
                except json.JSONDecodeError as e:
                    messages.error(request, f"Erreur de format JSON: {str(e)}")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                except Exception as e:
                    messages.error(request, f"Erreur de format du fichier: {str(e)}")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Traiter les questions
                questions_ajoutees = 0
                questions_ignorees = 0
                erreurs = []
                
                for i, question_data in enumerate(questions_data):
                    try:
                        # Validation des données de la question
                        if not question_data.get('enonce', '').strip():
                            raise ValueError("Énoncé manquant")
                        
                        if question_data.get('type', '') not in ['QCM', 'QRL']:
                            raise ValueError(f"Type de question invalide: {question_data.get('type', '')}")
                        
                        points = question_data.get('points', 1.0)
                        if not isinstance(points, (int, float)) or points <= 0:
                            raise ValueError(f"Points invalides: {points}")
                        
                        niveau = question_data.get('niveau', 'moyen').lower()
                        if niveau not in ['facile', 'moyen', 'difficile']:
                            niveau = 'moyen'
                        
                        # Validation de la somme des pondérations
                        est_valide, somme_totale, barème, message = valider_ponderation_test(test, float(points))
                        
                        if not est_valide:
                            if somme_totale > barème:
                                raise ValueError(f"Impossible d'ajouter cette question. {message}")
                            else:
                                # Afficher un avertissement mais continuer
                                print(f"AVERTISSEMENT: {message}")
                        
                        # Créer la question
                        nouvelle_question = Question.objects.create(
                            enonce=question_data['enonce'].strip(),
                            type_question=question_data['type'],
                            niveau_difficulte=niveau,
                            ponderation=float(points),
                            explication=question_data.get('explication', '').strip(),
                            module=test.module,
                            test=test
                        )
                        
                        # Ajouter les réponses pour QCM
                        if question_data['type'] == 'QCM':
                            reponses_data = question_data.get('reponses', [])
                            
                            if len(reponses_data) < 2:
                                raise ValueError("Un QCM doit avoir au moins 2 réponses")
                            
                            reponses_correctes = 0
                            for reponse_data in reponses_data:
                                if not reponse_data.get('texte', '').strip():
                                    continue
                                    
                                Reponse.objects.create(
                                    question=nouvelle_question,
                                    texte=reponse_data['texte'].strip(),
                                    est_correcte=bool(reponse_data.get('correcte', False))
                                )
                                
                                if reponse_data.get('correcte', False):
                                    reponses_correctes += 1
                            
                            if reponses_correctes == 0:
                                raise ValueError("Un QCM doit avoir au moins une réponse correcte")
                        
                        questions_ajoutees += 1
                        # print(f"DEBUG: Question {i+1} créée avec succès")  # Debug
                        
                    except Exception as e:
                        erreur_msg = f"Question {i+1}: {str(e)}"
                        erreurs.append(erreur_msg)
                        questions_ignorees += 1
                        # print(f"DEBUG: Erreur question {i+1}: {str(e)}")  # Debug
                        
                        # Si on n'ignore pas les erreurs, arrêter le processus
                        if not ignorer_erreurs:
                            # Supprimer la question si elle a été créée partiellement
                            try:
                                if 'nouvelle_question' in locals():
                                    nouvelle_question.delete()
                            except:
                                pass
                            
                            messages.error(request, f"Import arrêté à cause d'une erreur: {erreur_msg}")
                            return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Messages de résultat
                if questions_ajoutees > 0:
                    messages.success(request, f"✅ {questions_ajoutees} question(s) importée(s) avec succès!")
                
                if questions_ignorees > 0:
                    messages.warning(request, f"⚠️ {questions_ignorees} question(s) ignorée(s) à cause d'erreurs.")
                
                if erreurs and len(erreurs) <= 5:  # Afficher jusqu'à 5 erreurs détaillées
                    for erreur in erreurs[:5]:
                        messages.error(request, erreur)
                elif len(erreurs) > 5:
                    messages.error(request, f"Et {len(erreurs)-5} autres erreurs...")
                
                return redirect('celica_web:test_form_edit', test_id=test.id)
                
            except Exception as e:
                # print(f"DEBUG: Erreur critique: {str(e)}")  # Debug
                messages.error(request, f"Erreur critique lors de l'import: {str(e)}")
                return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # === AJOUT DE QUESTIONS EXISTANTES ===
        elif action == 'add_existing_questions' and test:
            try:
                questions_ids = request.POST.getlist('questions_selectionnees')
                
                if not questions_ids:
                    messages.error(request, "Aucune question sélectionnée.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                questions_ajoutees = 0
                
                for question_id in questions_ids:
                    try:
                        question_originale = Question.objects.get(id=question_id)
                        
                        # Créer une copie de la question
                        nouvelle_question = Question.objects.create(
                            enonce=question_originale.enonce,
                            type_question=question_originale.type_question,
                            niveau_difficulte=question_originale.niveau_difficulte,
                            ponderation=question_originale.ponderation,
                            explication=question_originale.explication,
                            module=test.module,  # Associer au module du test
                            test=test  # Associer directement au test
                        )
                        
                        # Copier les réponses pour les QCM
                        if question_originale.type_question == 'QCM':
                            for reponse_originale in question_originale.reponses.all():
                                Reponse.objects.create(
                                    question=nouvelle_question,
                                    texte=reponse_originale.texte,
                                    est_correcte=reponse_originale.est_correcte
                                )
                        
                        questions_ajoutees += 1
                        
                    except Question.DoesNotExist:
                        messages.warning(request, f"Question avec ID {question_id} non trouvée.")
                        continue
                    except Exception as e:
                        messages.error(request, f"Erreur lors de l'ajout de la question {question_id}: {str(e)}")
                        continue
                
                if questions_ajoutees > 0:
                    messages.success(request, f"{questions_ajoutees} question(s) ajoutée(s) avec succès au test.")
                
                return redirect('celica_web:test_form_edit', test_id=test.id)
                
            except Exception as e:
                messages.error(request, f"Erreur lors de l'ajout des questions: {str(e)}")
                return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # === TIRAGE ALÉATOIRE DE QUESTIONS ===
        elif action == 'add_random_questions' and test:
            try:
                # Récupérer les paramètres du formulaire
                nombre_questions = request.POST.get('nombre_questions_aleatoires')
                
                # Validation du nombre de questions
                if not nombre_questions:
                    messages.error(request, "Le nombre de questions à tirer est obligatoire.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                try:
                    nombre_questions = int(nombre_questions)
                    if nombre_questions <= 0:
                        raise ValueError("Le nombre doit être positif")
                    if nombre_questions > 50:
                        raise ValueError("Le nombre ne peut pas dépasser 50")
                except ValueError:
                    messages.error(request, "Le nombre de questions doit être un entier positif (maximum 50).")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Préparer les filtres
                filtres = {
                    'module_filter_aleatoire': request.POST.get('module_filter_aleatoire', 'tous'),
                    'niveau_filter_aleatoire': request.POST.get('niveau_filter_aleatoire', 'tous'),
                    'type_question_filter_aleatoire': request.POST.get('type_question_filter_aleatoire', 'tous'),
                    'ponderation_max_aleatoire': request.POST.get('ponderation_max_aleatoire', '')
                }
                
                # Préparer les options
                options = {
                    'equilibrer_types': request.POST.get('equilibrer_types') == 'on',
                    'equilibrer_difficultes': request.POST.get('equilibrer_difficultes') == 'on',
                    'eviter_doublons': request.POST.get('eviter_doublons') == 'on',
                    'optimiser_bareme': request.POST.get('optimiser_bareme') == 'on'
                }
                
                # Effectuer le tirage aléatoire
                success, questions_selectionnees, message_tirage, statistiques = effectuer_tirage_aleatoire_questions(
                    test, nombre_questions, filtres, options
                )
                
                if not success:
                    messages.error(request, f"Échec du tirage aléatoire : {message_tirage}")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                if not questions_selectionnees:
                    messages.warning(request, "Aucune question n'a pu être sélectionnée avec les critères choisis.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                
                # Ajouter les questions sélectionnées au test
                success_ajout, nb_ajoutees, message_ajout = ajouter_questions_tirage_au_test(
                    test, questions_selectionnees, request.user
                )
                
                if success_ajout:
                    # Construire un message détaillé avec les statistiques
                    stats_selection = statistiques.get('selection', {})
                    message_detaille = f"{message_tirage} {message_ajout}"
                    
                    if stats_selection:
                        details = []
                        if stats_selection.get('repartition_types'):
                            types_str = ', '.join([f"{k}: {v}" for k, v in stats_selection['repartition_types'].items()])
                            details.append(f"Types: {types_str}")
                        
                        if stats_selection.get('repartition_niveaux'):
                            niveaux_str = ', '.join([f"{k}: {v}" for k, v in stats_selection['repartition_niveaux'].items()])
                            details.append(f"Niveaux: {niveaux_str}")
                        
                        if details:
                            message_detaille += f" Répartition: {' | '.join(details)}"
                    
                    messages.success(request, message_detaille)
                else:
                    messages.warning(request, message_ajout)
                
                return redirect('celica_web:test_form_edit', test_id=test.id)
                
            except Exception as e:
                messages.error(request, f"Erreur lors du tirage aléatoire: {str(e)}")
                return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # === SUPPRESSION D'UNE QUESTION ===
        elif action == 'delete_question' and test:
            question_id = request.POST.get('question_id')
            if question_id:
                try:
                    question = get_object_or_404(Question, id=question_id, test=test)
                    question.delete()
                    messages.success(request, "Question supprimée avec succès!")
                except Exception as e:
                    messages.error(request, f"Erreur lors de la suppression: {str(e)}")
            return redirect('celica_web:test_form_edit', test_id=test.id)
    
    # Recharger les questions après modification
    if test:
        questions_existantes = Question.objects.filter(test=test).prefetch_related('reponses')
        for question in questions_existantes:
            question.reponses_count = question.reponses.count()
    
    context = {
        'test': test,
        'modules': modules,
        'questions_existantes': questions_existantes,
    }
    
    return render(request, 'celicaweb/test_form.html', context)

def parser_fichier_texte(contenu):
    """Parse un fichier texte et retourne une liste de questions"""
    lines = [line.strip() for line in contenu.split('\n') if line.strip()]
    questions = []
    i = 0
    
    while i < len(lines):
        try:
            # Type de question (obligatoire)
            if i >= len(lines):
                break
                
            type_question = lines[i].strip().upper()
            if type_question not in ['QCM', 'QRL']:
                raise ValueError(f"Type de question invalide à la ligne {i+1}: '{lines[i]}'. Utilisez 'QCM' ou 'QRL'")
            
            # Énoncé (obligatoire)
            if i+1 >= len(lines):
                raise ValueError(f"Énoncé manquant après le type '{type_question}' à la ligne {i+2}")
            
            enonce = lines[i+1].strip()
            if not enonce:
                raise ValueError(f"Énoncé vide à la ligne {i+2}")
            
            # Niveau de difficulté (obligatoire)
            if i+2 >= len(lines):
                raise ValueError(f"Niveau de difficulté manquant à la ligne {i+3}")
            
            niveau = lines[i+2].strip().lower()
            if niveau not in ['facile', 'moyen', 'difficile']:
                raise ValueError(f"Niveau invalide à la ligne {i+3}: '{lines[i+2]}'. Utilisez: facile, moyen, ou difficile")
            
            # Points (obligatoire)
            if i+3 >= len(lines):
                raise ValueError(f"Points manquants à la ligne {i+4}")
            
            try:
                points = float(lines[i+3])
                if points <= 0:
                    raise ValueError("Les points doivent être positifs")
            except ValueError:
                raise ValueError(f"Points invalides à la ligne {i+4}: '{lines[i+3]}'. Utilisez un nombre décimal positif (ex: 2.0)")
            
            question_data = {
                'type': type_question,
                'enonce': enonce,
                'niveau': niveau,
                'points': points,
                'explication': '',
                'reponses': []
            }
            
            i += 4
            
            # Explication (optionnelle) - si la ligne suivante ne commence pas par ✓ ou ○
            if i < len(lines) and not (lines[i].startswith('✓') or lines[i].startswith('○')):
                question_data['explication'] = lines[i].strip()
                i += 1
            
            # Réponses pour QCM
            if type_question == 'QCM':
                reponses_trouvees = 0
                reponses_correctes = 0
                
                while i < len(lines) and (lines[i].startswith('✓') or lines[i].startswith('○')):
                    ligne_reponse = lines[i].strip()
                    
                    if ligne_reponse.startswith('✓'):
                        est_correcte = True
                        reponses_correctes += 1
                    elif ligne_reponse.startswith('○'):
                        est_correcte = False
                    else:
                        break
                    
                    texte_reponse = ligne_reponse[1:].strip()
                    if not texte_reponse:
                        raise ValueError(f"Texte de réponse vide à la ligne {i+1}")
                    
                    question_data['reponses'].append({
                        'texte': texte_reponse,
                        'correcte': est_correcte
                    })
                    
                    reponses_trouvees += 1
                    i += 1
                
                # Validation des réponses QCM
                if reponses_trouvees < 2:
                    raise ValueError(f"Un QCM doit avoir au moins 2 réponses (trouvées: {reponses_trouvees})")
                
                if reponses_correctes == 0:
                    raise ValueError("Un QCM doit avoir au moins une réponse correcte (marquée avec ✓)")
            
            questions.append(question_data)
            
        except Exception as e:
            raise ValueError(f"Erreur de format: {str(e)}")
    
    if not questions:
        raise ValueError("Aucune question valide trouvée dans le fichier")
    
    return questions

@login_required
@user_passes_test(lambda u: u.is_staff or hasattr(u, 'instructeur'))
def test_duplicate(request, test_id):
    """Vue temporaire pour la duplication de test"""
    try:
        test_original = get_object_or_404(Test, id=test_id)
        messages.info(request, f"Fonctionnalité de duplication en cours de développement pour le test: {test_original.titre}")
        return redirect('celica_web:test_form', test_id=test_id)
    except Exception as e:
        messages.error(request, f"Erreur lors de la duplication: {str(e)}")
        return redirect('celica_web:test_list')

def parser_fichier_texte(contenu):
    """Parse un fichier texte et retourne une liste de questions"""
    lines = [line.strip() for line in contenu.split('\n') if line.strip()]
    questions = []
    i = 0
    
    while i < len(lines):
        try:
            # Type de question (obligatoire)
            if i >= len(lines):
                break
                
            type_question = lines[i].strip().upper()
            if type_question not in ['QCM', 'QRL']:
                raise ValueError(f"Type de question invalide à la ligne {i+1}: '{lines[i]}'. Utilisez 'QCM' ou 'QRL'")
            
            # Énoncé (obligatoire)
            if i+1 >= len(lines):
                raise ValueError(f"Énoncé manquant après le type '{type_question}' à la ligne {i+2}")
            
            enonce = lines[i+1].strip()
            if not enonce:
                raise ValueError(f"Énoncé vide à la ligne {i+2}")
            
            # Niveau de difficulté (obligatoire)
            if i+2 >= len(lines):
                raise ValueError(f"Niveau de difficulté manquant à la ligne {i+3}")
            
            niveau = lines[i+2].strip().lower()
            if niveau not in ['facile', 'moyen', 'difficile']:
                raise ValueError(f"Niveau invalide à la ligne {i+3}: '{lines[i+2]}'. Utilisez: facile, moyen, ou difficile")
            
            # Points (obligatoire)
            if i+3 >= len(lines):
                raise ValueError(f"Points manquants à la ligne {i+4}")
            
            try:
                points = float(lines[i+3])
                if points <= 0:
                    raise ValueError("Les points doivent être positifs")
            except ValueError:
                raise ValueError(f"Points invalides à la ligne {i+4}: '{lines[i+3]}'. Utilisez un nombre décimal positif (ex: 2.0)")
            
            question_data = {
                'type': type_question,
                'enonce': enonce,
                'niveau': niveau,
                'points': points,
                'explication': '',
                'reponses': []
            }
            
            i += 4
            
            # Explication (optionnelle) - si la ligne suivante ne commence pas par ✓ ou ○
            if i < len(lines) and not (lines[i].startswith('✓') or lines[i].startswith('○')):
                question_data['explication'] = lines[i].strip()
                i += 1
            
            # Réponses pour QCM
            if type_question == 'QCM':
                reponses_trouvees = 0
                reponses_correctes = 0
                
                while i < len(lines) and (lines[i].startswith('✓') or lines[i].startswith('○')):
                    ligne_reponse = lines[i].strip()
                    
                    if ligne_reponse.startswith('✓'):
                        est_correcte = True
                        reponses_correctes += 1
                    elif ligne_reponse.startswith('○'):
                        est_correcte = False
                    else:
                        break
                    
                    texte_reponse = ligne_reponse[1:].strip()
                    if not texte_reponse:
                        raise ValueError(f"Texte de réponse vide à la ligne {i+1}")
                    
                    question_data['reponses'].append({
                        'texte': texte_reponse,
                        'correcte': est_correcte
                    })
                    
                    reponses_trouvees += 1
                    i += 1
                
                # Validation des réponses QCM
                if reponses_trouvees < 2:
                    raise ValueError(f"Un QCM doit avoir au moins 2 réponses (trouvées: {reponses_trouvees})")
                
                if reponses_correctes == 0:
                    raise ValueError("Un QCM doit avoir au moins une réponse correcte (marquée avec ✓)")
            
            questions.append(question_data)
            
        except Exception as e:
            raise ValueError(f"Erreur de format: {str(e)}")
    
    if not questions:
        raise ValueError("Aucune question valide trouvée dans le fichier")
    
    return questions

# === FONCTIONS AUXILIAIRES ===

def valider_ponderation_test(test, nouvelle_ponderation=0):
    """
    Valide que la somme des pondérations des questions du test soit égale au barème.
    
    Args:
        test: Instance du test
        nouvelle_ponderation: Pondération de la nouvelle question à ajouter (optionnel)
    
    Returns:
        tuple: (est_valide, somme_actuelle, barème, message)
    """
    # Calculer la somme des pondérations actuelles
    somme_actuelle = test.questions.aggregate(
        total=Sum('ponderation')
    )['total'] or 0
    
    # Ajouter la nouvelle pondération si fournie
    somme_totale = somme_actuelle + nouvelle_ponderation
    
    # Comparer avec le barème
    if abs(somme_totale - test.bareme) < 0.01:  # Tolérance pour les erreurs de précision float
        return True, somme_totale, test.bareme, f"✅ Somme des pondérations ({somme_totale}) = Barème ({test.bareme})"
    elif somme_totale > test.bareme:
        return False, somme_totale, test.bareme, f"❌ Somme des pondérations ({somme_totale}) > Barème ({test.bareme}) - Excédent: {somme_totale - test.bareme}"
    else:
        return False, somme_totale, test.bareme, f"⚠️ Somme des pondérations ({somme_totale}) < Barème ({test.bareme}) - Manque: {test.bareme - somme_totale}"

def valider_ponderation_stricte(test, nouvelle_ponderation):
    """
    Validation stricte de la pondération - refuse catégoriquement les dépassements
    
    Args:
        test: Instance du test
        nouvelle_ponderation: Pondération de la nouvelle question à ajouter
    
    Returns:
        tuple: (autorise_ajout, message_erreur)
    """
    # Calculer la somme actuelle des pondérations
    somme_actuelle = test.questions.aggregate(
        total=Sum('ponderation')
    )['total'] or 0
    
    # Calculer la somme totale après ajout
    somme_totale = somme_actuelle + nouvelle_ponderation
    barème = test.bareme
    
    # Refuser catégoriquement si ça dépasse le barème
    if somme_totale > barème:
        excédent = somme_totale - barème
        message = f"Impossible d'ajouter cette question. La somme des pondérations ({somme_totale}) dépasserait le barème du test ({barème}) de {excédent} points."
        return False, message
    
    # Autoriser l'ajout si dans les limites
    if abs(somme_totale - barème) < 0.01:  # Tolérance pour float
        message = f"Question ajoutée. Barème complet ({somme_totale}/{barème})."
    else:
        restant = barème - somme_totale
        message = f"Question ajoutée. Barème restant: {restant} points ({somme_totale}/{barème})."
    
    return True, message

def ajouter_question_dynamique(request, test):
    """Ajouter une question créée dynamiquement au test"""
    try:
        enonce = request.POST.get('enonce', '').strip()
        type_question = request.POST.get('type_question')
        niveau_difficulte = request.POST.get('niveau_difficulte', 'moyen')
        ponderation = request.POST.get('ponderation', 1.0)
        explication = request.POST.get('explication', '').strip()
        
        # Validation
        if not enonce:
            messages.error(request, "L'énoncé est obligatoire")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        if not type_question:
            messages.error(request, "Le type de question est obligatoire")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Valider la pondération
        try:
            ponderation_float = float(ponderation)
            if ponderation_float <= 0:
                raise ValueError("La pondération doit être positive")
        except ValueError:
            messages.error(request, "La pondération doit être un nombre positif.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Validation stricte de la somme des pondérations
        autorise_ajout, message = valider_ponderation_stricte(test, ponderation_float)
        
        if not autorise_ajout:
            messages.error(request, message)
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Créer la question
        question = Question.objects.create(
            enonce=enonce,
            type_question=type_question,
            niveau_difficulte=niveau_difficulte,
            ponderation=ponderation_float,
            explication=explication,
            module=test.module,
            test=test
        )
        
        # Pour les QCM, créer les réponses
        if type_question == 'QCM':
            reponses_ajoutees = 0
            correcte_trouvee = False
            
            # Parcourir tous les champs de réponse possibles avec le format correct
            for i in range(10):  # Jusqu'à 10 réponses maximum (indices 0-9)
                reponse_text = request.POST.get(f'new_reponse-{i}-texte', '').strip()
                est_correcte = request.POST.get(f'new_reponse-{i}-est_correcte') == 'on'
                
                if reponse_text:
                    Reponse.objects.create(
                        question=question,
                        texte=reponse_text,
                        est_correcte=est_correcte
                    )
                    reponses_ajoutees += 1
                    if est_correcte:
                        correcte_trouvee = True
            
            # Validation des réponses QCM
            if reponses_ajoutees < 2:
                question.delete()
                messages.error(request, "Un QCM doit avoir au moins 2 réponses")
                return redirect('celica_web:test_form_edit', test_id=test.id)
            
            if not correcte_trouvee:
                question.delete()
                messages.error(request, "Un QCM doit avoir au moins une réponse correcte")
                return redirect('celica_web:test_form_edit', test_id=test.id)
        
        messages.success(request, "Question ajoutée avec succès!")
        return redirect('celica_web:test_form_edit', test_id=test.id)
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'ajout de la question: {str(e)}")
        return redirect('celica_web:test_form_edit', test_id=test.id)

def importer_questions_fichier(request, test):
    """Importer des questions depuis un fichier CSV/Excel"""
    try:
        fichier = request.FILES.get('fichier_import')
        
        if not fichier:
            messages.error(request, "Aucun fichier sélectionné")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Traitement du fichier (à implémenter selon vos besoins)
        messages.info(request, "Fonctionnalité d'import en cours de développement")
        return redirect('celica_web:test_form_edit', test_id=test.id)
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'import: {str(e)}")
        return redirect('celica_web:test_form_edit', test_id=test.id)
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'import: {str(e)}")
        return redirect('celica_web:test_form_edit', test_id=test.id)

def ajouter_question_dynamique(request, test):
    """Ajoute une question avec réponses dynamiques"""
    try:
        # Récupérer les données de la question
        enonce = request.POST.get('enonce', '').strip()
        type_question = request.POST.get('type_question', '').strip()
        niveau_difficulte = request.POST.get('niveau_difficulte', 'moyen')
        ponderation = request.POST.get('ponderation', '1.0')
        explication = request.POST.get('explication', '').strip()
        
        # Validation de base
        if not enonce:
            messages.error(request, "L'énoncé de la question est obligatoire.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        if not type_question:
            messages.error(request, "Le type de question est obligatoire.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        try:
            ponderation_float = float(ponderation)
            if ponderation_float <= 0:
                raise ValueError("La pondération doit être positive")
        except ValueError:
            messages.error(request, "La pondération doit être un nombre positif.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Validation stricte de la somme des pondérations
        autorise_ajout, message = valider_ponderation_stricte(test, ponderation_float)
        
        if not autorise_ajout:
            messages.error(request, message)
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Créer la question avec vérification de doublon
        question, est_doublon, message = creer_question_sans_doublon(
            enonce=enonce,
            type_question=type_question,
            niveau_difficulte=niveau_difficulte,
            ponderation=ponderation_float,
            module=test.module,
            instructeur=request.user,
            explication=explication,
            test=test
        )
        
        if est_doublon:
            messages.warning(request, message)
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Traitement des réponses pour les QCM
        if type_question == 'QCM':
            reponses_ajoutees = 0
            au_moins_une_correcte = False
            reponses_textes = []
            
            # Parcourir tous les champs POST pour trouver les réponses (new_reponse-0-texte, new_reponse-1-texte, etc.)
            for i in range(10):  # Maximum 10 réponses (indices 0-9)
                texte_key = f'new_reponse-{i}-texte'
                correcte_key = f'new_reponse-{i}-est_correcte'
                
                texte_reponse = request.POST.get(texte_key, '').strip()
                
                if texte_reponse:
                    # Vérifier si cette réponse est marquée comme correcte
                    est_correcte = request.POST.get(correcte_key) == 'on'
                    
                    # Vérifier les doublons
                    if texte_reponse.lower() in [r.lower() for r in reponses_textes]:
                        messages.warning(request, f"Réponse dupliquée ignorée: '{texte_reponse}'")
                        continue
                    
                    reponses_textes.append(texte_reponse)
                    
                    Reponse.objects.create(
                        question=question,
                        texte=texte_reponse,
                        est_correcte=est_correcte
                    )
                    
                    reponses_ajoutees += 1
                    if est_correcte:
                        au_moins_une_correcte = True
            
            # Validation des réponses QCM
            if reponses_ajoutees < 2:
                question.delete()
                messages.error(request, "Un QCM doit avoir au minimum 2 réponses.")
                return redirect('celica_web:test_form_edit', test_id=test.id)
            
            if not au_moins_une_correcte:
                question.delete()
                messages.error(request, "Un QCM doit avoir au moins une réponse correcte.")
                return redirect('celica_web:test_form_edit', test_id=test.id)
            
            success_message = f"Question QCM ajoutée avec succès! ({reponses_ajoutees} réponses)"
        else:
            success_message = "Question à réponse libre ajoutée avec succès!"
        
        # Sauvegarder dans la banque de questions si demandé
        sauvegarder_banque = request.POST.get('sauvegarder_banque') == 'on'
        if sauvegarder_banque:
            try:
                # Créer une copie dans la banque (sans test assigné)
                question_banque = Question.objects.create(
                    enonce=enonce,
                    type_question=type_question,
                    niveau_difficulte=niveau_difficulte,
                    ponderation=ponderation_float,
                    explication=explication,
                    module=test.module,
                    instructeur=request.user
                    # Pas de test assigné pour la banque
                )
                
                # Copier les réponses si QCM
                if type_question == 'QCM':
                    # CORRECTION : Utiliser 'reponses' au lieu de 'reponses'
                    for reponse in question.reponses.all():
                        Reponse.objects.create(
                            question=question_banque,
                            texte=reponse.texte,
                            est_correcte=reponse.est_correcte
                        )
                
                success_message += " Question sauvegardée dans la banque."
                
            except Exception as e:
                messages.warning(request, f"Question ajoutée au test mais erreur lors de la sauvegarde en banque: {str(e)}")
        
        messages.success(request, success_message)
        
        # Vérifier si l'utilisateur veut continuer à ajouter des questions
        continuer_ajout = request.POST.get('continuer_ajout') == 'on'
        if continuer_ajout:
            messages.info(request, "Vous pouvez ajouter une autre question.")
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'ajout de la question: {str(e)}")
        logger.error(f"Erreur dans ajouter_question_dynamique: {str(e)}")
    
    return redirect('celica_web:test_form_edit', test_id=test.id)

def importer_questions_fichier(request, test):
    """Importe des questions depuis un fichier CSV/Excel"""
    try:
        fichier = request.FILES.get('fichier_import')
        if not fichier:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Déterminer le format du fichier
        nom_fichier = fichier.name.lower()
        if nom_fichier.endswith('.csv'):
            format_fichier = 'csv'
        elif nom_fichier.endswith(('.xlsx', '.xls')):
            format_fichier = 'excel'
        else:
            messages.error(request, "Format de fichier non supporté. Utilisez CSV ou Excel.")
            return redirect('celica_web:test_form_edit', test_id=test.id)
        
        # Traiter l'import
        questions_importees = process_import_file(fichier, format_fichier, test, request.user)
        
        if questions_importees > 0:
            messages.success(request, f"{questions_importees} question(s) importée(s) avec succès!")
        else:
            messages.warning(request, "Aucune question n'a été importée. Vérifiez le format du fichier.")
            
    except Exception as e:
        messages.error(request, f"Erreur lors de l'importation: {str(e)}")
        logger.error(f"Erreur dans importer_questions_fichier: {str(e)}")
    
    return redirect('celica_web:test_form_edit', test_id=test.id)

def ajouter_questions_existantes_au_test(request, test_id):
    """Ajoute des questions existantes de la banque au test"""
    try:
        questions_ids = request.POST.getlist('questions_existantes')
        if not questions_ids:
            messages.warning(request, "Aucune question sélectionnée.")
            return redirect('celica_web:test_form_edit', test_id=test_id)
        
        questions_ajoutees = 0
        for question_id in questions_ids:
            try:
                question_originale = get_object_or_404(Question, id=question_id)
                
                # Dupliquer la question pour ce test
                nouvelle_question = Question.objects.create(
                    enonce=question_originale.enonce,
                    type_question=question_originale.type_question,
                    niveau_difficulte=question_originale.niveau_difficulte,
                    ponderation=question_originale.ponderation,
                    explication=question_originale.explication,
                    test=test,
                    module=test.module,
                    instructeur=request.user
                )
                
                # Dupliquer les réponses si QCM
                if question_originale.type_question == 'QCM':
                    # CORRECTION : Utiliser 'reponses' au lieu de 'reponses'
                    for reponse in question_originale.reponses.all():
                        Reponse.objects.create(
                            question=nouvelle_question,
                            texte=reponse.texte,
                            est_correcte=reponse.est_correcte,
                            explication=reponse.explication
                        )
                
                questions_ajoutees += 1
            except Exception as e:
                messages.warning(request, f"Erreur lors de l'ajout de la question {question_id}: {str(e)}")
        
        if questions_ajoutees > 0:
            messages.success(request, f"{questions_ajoutees} question(s) ajoutée(s) au test avec succès!")
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'ajout des questions: {str(e)}")
        logger.error(f"Erreur dans ajouter_questions_existantes_au_test: {str(e)}")
    
    return redirect('celica_web:test_form_edit', test_id=test_id)
# === FONCTION AJAX POUR SUPPRESSION ===
from django.http import JsonResponse

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def delete_question_ajax(request, test_id, question_id):
    """Suppression AJAX d'une question"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        test = get_object_or_404(Test, id=test_id)
        question = get_object_or_404(Question, id=question_id, test=test)
        
        # Vérifier les permissions
        if test.instructeur != request.user and not request.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Permission refusée'})
        
        question.delete()
        return JsonResponse({'success': True, 'message': 'Question supprimée avec succès'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def nettoyer_doublons_questions(request):
    """
    Vue pour nettoyer les doublons dans la banque de questions ou toutes les questions selon le rôle
    """
    from .utils.deduplication import nettoyer_doublons_questions as nettoyer_doublons
    from django.db.models import Count
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Administrateurs').exists()

    # Déterminer le queryset selon le rôle
    if is_admin:
        questions_queryset = Question.objects.all()
    else:
        questions_queryset = Question.objects.filter(instructeur=user)

    if request.method == 'POST':
        try:
            stats = nettoyer_doublons(questions_queryset)
            if stats['questions_supprimées'] > 0:
                messages.success(
                    request, 
                    f"Nettoyage terminé ! {stats['questions_supprimées']} doublons supprimés "
                    f"parmi {stats['questions_analysées']} questions analysées."
                )
            else:
                messages.info(request, "Aucun doublon trouvé dans la banque de questions.")
            if stats['erreurs']:
                for erreur in stats['erreurs']:
                    messages.warning(request, erreur)
        except Exception as e:
            messages.error(request, f"Erreur lors du nettoyage : {str(e)}")

    total_questions = questions_queryset.count()
    questions_par_module = questions_queryset.values('module__intitule').annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'total_questions': total_questions,
        'questions_par_module': questions_par_module,
    }
    return render(request, 'celicaweb/nettoyer_doublons.html', context)


# Reste des fonctions...
def traiter_question_manuelle(request, test_instance):
    """Traite l'ajout d'une question manuelle"""
    
    try:
        # Créer les formulaires avec les données POST
        question_form = ManualQuestionForm(request.POST, request.FILES)
        
        print(f"Question form valid: {question_form.is_valid()}")
        
        if question_form.is_valid():
            # **CORRECTION: Créer la question**
            question = question_form.save(commit=False)
            question.test = test_instance
            question.module = test_instance.module
            question.instructeur = request.user
            
            # Assigner des valeurs par défaut si nécessaire
            if not hasattr(question, 'ordre') or question.ordre is None:
                question.ordre = 1
            if not hasattr(question, 'actif') or question.actif is None:
                question.actif = True
                
            question.save()
            
            # **CORRECTION: Traiter les réponses manuellement**
            reponses_count = 0
            correctes_count = 0
            
            # Parcourir toutes les réponses possibles (max 10)
            for i in range(10):
                texte_key = f'new_reponse-{i}-texte'
                correcte_key = f'new_reponse-{i}-est_correcte'
                
                texte = request.POST.get(texte_key, '').strip()
                est_correcte = request.POST.get(correcte_key) == 'on'
                
                if texte:  # Si le texte n'est pas vide
                    Reponse.objects.create(
                        question=question,
                        texte=texte,
                        est_correcte=est_correcte,
                        ordre=reponses_count + 1
                    )
                    reponses_count += 1
                    if est_correcte:
                        correctes_count += 1
                    print(f"  📋 Réponse {i+1}: '{texte[:30]}...' (correcte: {est_correcte})")

            # **CORRECTION: Validation des réponses**
            if reponses_count < 2:
                question.delete()
                messages.error(request, "Au moins deux réponses sont requises.")
                return redirect('celica_web:test_form', test_id=test_instance.id)
            
            if question.type_question == 'QCM' and correctes_count == 0:
                question.delete()
                messages.error(request, "Un QCM doit avoir au moins une réponse correcte.")
                return redirect('celica_web:test_form', test_id=test_instance.id)
            
            messages.success(request, f"Question ajoutée avec succès avec {reponses_count} réponses!")
            return redirect('celica_web:test_form', test_id=test_instance.id)
        
        else:
            # **CORRECTION: Gestion des erreurs**
            errors = []
            if not question_form.is_valid():
                errors.extend([f"Question: {error}" for error in question_form.errors.values()])
            
            messages.error(request, f"Erreur lors de l'ajout de la question: {'; '.join(map(str, errors))}")
            return redirect('celica_web:test_form', test_id=test_instance.id)
    
    except Exception as e:
        logger.error(f"Erreur traitement question manuelle: {e}")
        messages.error(request, f"Erreur lors de l'ajout de la question: {str(e)}")
        return redirect('celica_web:test_form', test_id=test_instance.id)

def traiter_import_questions(request, test_instance):
    """Traite l'importation de questions depuis un fichier"""
    messages.info(request, "Fonctionnalité d'importation en cours de développement.")
    return redirect('celica_web:test_form', test_id=test_instance.id)

def traiter_selection_questions(request, test_instance):
    """Traite la sélection de questions existantes"""
    messages.info(request, "Fonctionnalité de sélection en cours de développement.")
    return redirect('celica_web:test_form', test_id=test_instance.id)

def traiter_question_manuelle(request, test):
    """Traiter l'ajout d'une question manuelle avec gestion d'erreurs améliorée"""
    
    try:
        # Récupérer les données de la question
        question_data = {
            'enonce': request.POST.get('enonce', '').strip(),
            'type_question': request.POST.get('type_question', ''),
            'ponderation': request.POST.get('ponderation', '1'),
        }

        # Validation manuelle des données de base
        if not question_data['enonce']:
            messages.error(request, "L'énoncé de la question est obligatoire.")
            return redirect('celica_web:test_form', test_id=test.id)
        
        if not question_data['type_question']:
            messages.error(request, "Le type de question est obligatoire.")
            return redirect('celica_web:test_form', test_id=test.id)
        
        # Créer la question directement
        try:
            ponderation = float(question_data['ponderation'])
        except (ValueError, TypeError):
            ponderation = 1.0
        
        question = Question.objects.create(
            test=test,
            module=test.module,
            instructeur=request.user,
            enonce=question_data['enonce'],
            type_question=question_data['type_question'],
            ponderation=ponderation
        )

        # Traiter les réponses
        reponses_creees = 0
        reponses_correctes = 0
        
        # Parcourir toutes les réponses possibles
        for i in range(10):  # Maximum 10 réponses
            texte_key = f'new_reponse-{i}-texte'
            correcte_key = f'new_reponse-{i}-est_correcte'
            
            texte = request.POST.get(texte_key, '').strip()
            est_correcte = request.POST.get(correcte_key) == 'on'
            
            if texte:  # Si le texte n'est pas vide
                Reponse.objects.create(
                    question=question,
                    texte=texte,
                    est_correcte=est_correcte
                )
                reponses_creees += 1
                if est_correcte:
                    reponses_correctes += 1
                print(f"  📋 Réponse {i+1}: '{texte[:30]}...' (correcte: {est_correcte})")
        
        # Validation des réponses
        if reponses_creees < 2:
            question.delete()  # Supprimer la question créée
            messages.error(request, "Au moins deux réponses sont requises pour chaque question.")
            return redirect('celica_web:test_form', test_id=test.id)
        
        if question.type_question == 'QCM' and reponses_correctes == 0:
            question.delete()  # Supprimer la question créée
            messages.error(request, "Un QCM doit avoir au moins une réponse correcte.")
            return redirect('celica_web:test_form', test_id=test.id)
        
        messages.success(request, f"Question ajoutée avec succès! ({reponses_creees} réponses créées, {reponses_correctes} correctes)")
        return redirect('celica_web:test_form', test_id=test.id)
        
    except Exception as e:
        logger.error(f"Erreur traitement question manuelle: {e}")
        messages.error(request, f"Erreur lors de l'ajout de la question: {str(e)}")
        return redirect('celica_web:test_form', test_id=test.id)

def traiter_import_questions(request, test):
    """Traiter l'importation de questions depuis un fichier"""
    messages.info(request, "Fonctionnalité d'importation en cours de développement.")
    return redirect('celica_web:test_form', test_id=test.id)

def traiter_selection_questions(request, test):
    """Traiter la sélection de questions existantes"""
    messages.info(request, "Fonctionnalité de sélection en cours de développement.")
    return redirect('celica_web:test_form', test_id=test.id)

def test_preview(request, test_id):
    """Vue pour afficher l'aperçu d'un test"""
    test = get_object_or_404(Test, id=test_id)
    questions = test.questions.all().prefetch_related('reponses')
    
    context = {
        'test': test,
        'questions': questions,
        'is_preview': True,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'celicaweb/test_preview_partial.html', context)
    
    return render(request, 'celicaweb/test_preview.html', context)

# **NOUVELLES FONCTIONS AUXILIAIRES**

def traiter_question_manuelle(request, test_instance):
    """Traite l'ajout d'une question manuelle"""
    
    try:
        # Créer les formulaires avec les données POST
        question_form = ManualQuestionForm(request.POST, request.FILES)
        
        print(f"Question form valid: {question_form.is_valid()}")
        
        if question_form.is_valid():
            # **CORRECTION: Créer la question**
            question = question_form.save(commit=False)
            question.test = test_instance
            question.module = test_instance.module
            question.instructeur = request.user
            
            # Assigner des valeurs par défaut si nécessaire
            if not hasattr(question, 'ordre') or question.ordre is None:
                question.ordre = 1
            if not hasattr(question, 'actif') or question.actif is None:
                question.actif = True
                
            question.save()
            
            # **CORRECTION: Traiter les réponses manuellement**
            reponses_count = 0
            correctes_count = 0
            
            # Parcourir toutes les réponses possibles (max 10)
            for i in range(10):
                texte_key = f'new_reponse-{i}-texte'
                correcte_key = f'new_reponse-{i}-est_correcte'
                
                texte = request.POST.get(texte_key, '').strip()
                est_correcte = request.POST.get(correcte_key) == 'on'
                
                if texte:  # Si le texte n'est pas vide
                    Reponse.objects.create(
                        question=question,
                        texte=texte,
                        est_correcte=est_correcte,
                        ordre=reponses_count + 1
                    )
                    reponses_count += 1
                    if est_correcte:
                        correctes_count += 1
                    print(f"  📋 Réponse {i+1}: '{texte[:30]}...' (correcte: {est_correcte})")

            # **CORRECTION: Validation des réponses**
            if reponses_count < 2:
                question.delete()
                messages.error(request, "Au moins deux réponses sont requises.")
                return redirect('celica_web:test_form', test_id=test_instance.id)
            
            if question.type_question == 'QCM' and correctes_count == 0:
                question.delete()
                messages.error(request, "Un QCM doit avoir au moins une réponse correcte.")
                return redirect('celica_web:test_form', test_id=test_instance.id)
            
            messages.success(request, f"Question ajoutée avec succès avec {reponses_count} réponses!")
            return redirect('celica_web:test_form', test_id=test_instance.id)
        
        else:
            # **CORRECTION: Gestion des erreurs**
            errors = []
            if not question_form.is_valid():
                errors.extend([f"Question: {error}" for error in question_form.errors.values()])
            
            messages.error(request, f"Erreur lors de l'ajout de la question: {'; '.join(map(str, errors))}")
            return redirect('celica_web:test_form', test_id=test_instance.id)
    
    except Exception as e:
        logger.error(f"Erreur traitement question manuelle: {e}")
        messages.error(request, f"Erreur lors de l'ajout de la question: {str(e)}")
        return redirect('celica_web:test_form', test_id=test_instance.id)

def traiter_import_questions(request, test_instance):
    """Traite l'importation de questions depuis un fichier"""
    
    try:
        import_form = ImportForm(request.POST, request.FILES)
        
        if import_form.is_valid():
            uploaded_file = import_form.cleaned_data['fichier']
            
            # **CORRECTION: Traitement basique du fichier**
            questions_imported = 0
            # Ici vous pouvez ajouter votre logique d'importation
            # Pour l'instant, on simule
            
            messages.success(request, f"Importation en cours de développement. Fichier reçu: {uploaded_file.name}")
            return redirect('celica_web:test_edit', test_id=test_instance.id)
        else:
            messages.error(request, "Erreur dans le fichier d'importation.")
            return redirect('celica_web:test_edit', test_id=test_instance.id)
            
    except Exception as e:
        logger.error(f"Erreur importation: {e}")
        messages.error(request, f"Erreur lors de l'importation: {str(e)}")
        return redirect('celica_web:test_edit', test_id=test_instance.id)

def traiter_selection_questions(request, test_instance):
    """Traite la sélection de questions existantes"""
    
    try:
        selected_questions = request.POST.getlist('selected_questions')
        
        if selected_questions:
            questions_added = 0
            # Ici vous pouvez ajouter votre logique de sélection
            # Pour l'instant, on simule
            
            messages.success(request, f"Sélection en cours de développement. {len(selected_questions)} question(s) sélectionnée(s)")
            return redirect('celica_web:test_edit', test_id=test_instance.id)
        else:
            messages.warning(request, "Aucune question sélectionnée.")
            return redirect('celica_web:test_edit', test_id=test_instance.id)
            
    except Exception as e:
        logger.error(f"Erreur sélection: {e}")
        messages.error(request, f"Erreur lors de la sélection: {str(e)}")
        return redirect('celica_web:test_edit', test_id=test_instance.id)

def test_preview(request, test_id):
    """Vue pour afficher l'aperçu d'un test"""
    test = get_object_or_404(Test, id=test_id)
    questions = test.questions.all().prefetch_related('reponses')
    
    context = {
        'test': test,
        'questions': questions,
        'is_preview': True,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'celicaweb/test_preview_partial.html', context)
    
    return render(request, 'celicaweb/test_preview.html', context)

def process_import_file(uploaded_file, import_format, test_instance, user):
    """
    Fonction utilitaire pour traiter l'importation de questions depuis un fichier
    """
    import csv
    import io
    
    questions_imported = 0
    
    try:
        # Lire le contenu du fichier
        if import_format == 'csv':
            # Décoder le fichier CSV
            file_content = uploaded_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(file_content))
            
            for row in csv_reader:
                try:
                    # Récupérer la pondération
                    ponderation_float = float(row.get('ponderation', 1.0))
                    
                    # Validation stricte de la pondération
                    autorise_ajout, message = valider_ponderation_stricte(test_instance, ponderation_float)
                    
                    if not autorise_ajout:
                        logger.warning(f"Question ignorée lors de l'import: {message}")
                        continue
                    
                    # Créer la question
                    question = Question.objects.create(
                        enonce=row.get('enonce', '').strip(),
                        type_question=row.get('type_question', 'QCM'),
                        niveau_difficulte=row.get('niveau_difficulte', 'moyen'),
                        ponderation=ponderation_float,
                        module=test_instance.module,
                        test=test_instance,
                        instructeur=user
                    )
                    
                    # Créer les réponses
                    reponse_count = 1
                    while f'reponse_{reponse_count}' in row and row[f'reponse_{reponse_count}'].strip():
                        Reponse.objects.create(
                            question=question,
                            texte=row[f'reponse_{reponse_count}'].strip(),
                            est_correcte=row.get(f'est_correcte_{reponse_count}', 'False').lower() == 'true'
                        )
                        reponse_count += 1
                    
                    questions_imported += 1
                    
                except Exception as e:
                    logger.warning(f"Erreur lors de l'importation d'une ligne: {str(e)}")
                    continue
        
        elif import_format == 'excel':
            # Traitement pour Excel (à implémenter selon vos besoins)
            pass
    
    except Exception as e:
        logger.error(f"Erreur lors du traitement du fichier d'importation: {str(e)}")
        raise
    
    return questions_imported

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def test_list(request):
    utilisateur = request.user

    # Vérifier si l'utilisateur est un instructeur ou admin
    if utilisateur.role not in ['instructeur', 'admin']:
        messages.error(request, _("Seul un instructeur ou un administrateur peut accéder à la liste des tests."))
        return redirect_by_role(request.user)

    # Filtrage des tests selon le rôle
    if utilisateur.role == 'admin':
        # Les admins voient TOUS les tests
        tests = Test.objects.select_related('module', 'instructeur').all()
    else:
        # Les instructeurs voient seulement leurs tests
        tests = Test.objects.filter(instructeur=utilisateur).select_related('module')

    # Filtrage par module
    module_id = request.GET.get('module')
    if module_id and module_id != '':
        tests = tests.filter(module_id=module_id)

    # Filtrage par recherche
    search_query = request.GET.get('search')
    if search_query and search_query.strip():
        tests = tests.filter(
            titre__icontains=search_query
        ).distinct()

    # Pagination des tests
    tests = tests.order_by('-date_creation')  # Plus récents en premier
    paginator = Paginator(tests, 15)  # 15 tests par page
    page_number = request.GET.get('page')
    tests_page = paginator.get_page(page_number)

    # Modules pour le filtrage (afficher seulement les modules de l'instructeur)
    if utilisateur.role == 'admin':
        modules = Module.objects.all().order_by('intitule')
    else:
        modules = Module.objects.filter(
            models.Q(instructeur_principal=utilisateur) | models.Q(groupes__instructeurs=utilisateur)
        ).distinct().order_by('intitule')

    # Calcul du nombre total de questions (pour les tests visibles)
    from .models import Question
    total_questions = Question.objects.filter(test__in=tests).count()

    context = {
        'tests': tests_page,
        'modules': modules,
        'selected_module': module_id,
        'search_query': search_query,
        'is_admin': utilisateur.role == 'admin',
        'total_tests': tests.count(),
        'total_questions': total_questions,
    }
    return render(request, 'celicaweb/test_list.html', context)

def ajouter_questions_existantes_au_test(request, test_id):
    """
    Ajoute les questions sélectionnées au test en créant des copies
    """
    if request.method == 'POST' and request.POST.get('action') == 'add_existing_questions':
        try:
            test = Test.objects.get(id=test_id)
            questions_ids = request.POST.getlist('questions_selectionnees')
            
            if not questions_ids:
                messages.error(request, "Aucune question sélectionnée.")
                return redirect('celica_web:test_form_edit', test_id=test_id)
            
            questions_ajoutees = 0
            
            for question_id in questions_ids:
                try:
                    question_originale = Question.objects.get(id=question_id)
                    
                    # Validation stricte de la somme des pondérations
                    autorise_ajout, message = valider_ponderation_stricte(test, question_originale.ponderation)
                    
                    if not autorise_ajout:
                        messages.error(request, f"Question '{question_originale.enonce[:50]}...' ignorée: {message}")
                        continue
                    
                    # Créer une copie de la question
                    nouvelle_question = Question.objects.create(
                        enonce=question_originale.enonce,
                        type_question=question_originale.type_question,
                        niveau_difficulte=question_originale.niveau_difficulte,
                        ponderation=question_originale.ponderation,
                        explication=question_originale.explication,
                        module=test.module,  # Associer au module du test
                        test=test  # Associer directement au test
                    )
                    
                    # Copier les réponses pour les QCM
                    if question_originale.type_question == 'QCM':
                        for reponse_originale in question_originale.reponses.all():
                            Reponse.objects.create(
                                question=nouvelle_question,
                                texte=reponse_originale.texte,
                                est_correcte=reponse_originale.est_correcte
                            )
                    
                    questions_ajoutees += 1
                    
                except Question.DoesNotExist:
                    messages.warning(request, f"Question avec ID {question_id} non trouvée.")
                    continue
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'ajout de la question {question_id}: {str(e)}")
                    continue
            
            if questions_ajoutees > 0:
                messages.success(request, f"{questions_ajoutees} question(s) ajoutée(s) avec succès au test.")
            
            return redirect('celica_web:test_form_edit', test_id=test_id)
            
        except Test.DoesNotExist:
            messages.error(request, "Test non trouvé.")
            return redirect('celica_web:test_list')
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout des questions: {str(e)}")
            return redirect('celica_web:test_form_edit', test_id=test_id)
    
    return redirect('celica_web:test_form_edit', test_id=test_id)
@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def test_delete(request, test_id):
    """
    Supprime un test donné son test_id.
    """
    test = get_object_or_404(Test, id=test_id)

    # Vérifier si l'utilisateur est un instructeur ou admin
    if request.user.role not in ['instructeur', 'admin']:
        messages.error(request, _("Seul un instructeur ou un administrateur peut supprimer un test."))
        return redirect('celica_web:test_list')

    # Vérifier si le test appartient à l'instructeur
    if test.instructeur != request.user:
        messages.error(request, _("Vous n'êtes pas autorisé à supprimer ce test."))
        return redirect('celica_web:test_list')

    # Vérifier si le test est lié à une session active
    if Planning.objects.filter(test=test, date_debut__lte=timezone.now(), date_fin__gte=timezone.now()).exists():
        messages.error(request, _("Impossible de supprimer un test lié à une session active."))
        return redirect('celica_web:test_list')

    if request.method == "POST":
        try:
            test.delete()
            messages.success(request, _("Test supprimé avec succès."))
        except Exception as e:
            logger.error(f"Erreur lors de la suppression du test {test_id} : {str(e)}")
            messages.error(request, _("Une erreur s'est produite lors de la suppression du test : {}").format(str(e)))
        return redirect('celica_web:test_list')

    # Afficher une page de confirmation pour la suppression
    context = {
        'test': test,
    }
    return render(request, 'celicaweb/test_confirm_delete.html', context)
@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def generer_test_automatique(request):
    class GenererTestForm(forms.Form):
        titre = forms.CharField(max_length=200, label="Titre du test")
        module = forms.ModelChoiceField(queryset=Module.objects.all(), label="Module")
        nombre_questions = forms.IntegerField(min_value=1, label="Nombre de questions")
        niveau_difficulte = forms.ChoiceField(
            choices=[('', 'Tous')] + Question._meta.get_field('niveau_difficulte').choices,
            required=False,
            label="Niveau de difficulté"
        )
        type_question = forms.ChoiceField(
            choices=[('', 'Tous'), ('QCM', 'QCM'), ('QRL', 'QRL')],
            required=False,
            label="Type de question"
        )
    if request.method == 'POST':
        form = GenererTestForm(request.POST)
        if form.is_valid():
            titre = form.cleaned_data['titre']
            module = form.cleaned_data['module']
            nombre_questions = form.cleaned_data['nombre_questions']
            niveau_difficulte = form.cleaned_data['niveau_difficulte']
            type_question = form.cleaned_data['type_question']
            # Inclure toutes les questions du module : banque + questions d'autres tests
            questions = Question.objects.filter(
                Q(test__isnull=True, module=module) |  # Questions de la banque du module
                Q(test__isnull=False, module=module)   # Questions d'autres tests du même module
            )
            if niveau_difficulte:
                questions = questions.filter(niveau_difficulte=niveau_difficulte)
            if type_question:
                questions = questions.filter(type_question=type_question)
            if questions.count() < nombre_questions:
                messages.error(request, f"Pas assez de questions disponibles. Trouvé : {questions.count()}, requis : {nombre_questions}.")
                return render(request, 'celicaweb/generer_test_form.html', {'form': form})
            selected_questions = random.sample(list(questions), nombre_questions)
            test = Test.objects.create(titre=titre, module=module, createur=request.user)
            for question in selected_questions:
                question.test = test
                question.save()
            messages.success(request, f"Test '{titre}' généré avec succès avec {nombre_questions} questions.")
            return redirect('celica_web:test_list')
    else:
        form = GenererTestForm()
    return render(request, 'celicaweb/generer_test_form.html', {'form': form})

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def question_form_new(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES)
        reponse_formset = ReponseFormSet(request.POST)
        if form.is_valid() and reponse_formset.is_valid():
            question = form.save(commit=False)
            type_question = form.cleaned_data['type_question']
            
            # Compter seulement les réponses valides (non vides et non supprimées)
            response_count = 0
            correct_count = 0
            for form in reponse_formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    if form.cleaned_data.get('texte', '').strip():  # Réponse non vide
                        response_count += 1
                        if form.cleaned_data.get('est_correcte'):
                            correct_count += 1
            
            if request.POST.get('save_to_db'):
                if type_question == 'QCM' and response_count < 2:
                    messages.error(request, "Un QCM doit avoir au moins 2 réponses.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset})
                if type_question == 'QCM' and correct_count == 0:
                    messages.error(request, "Un QCM doit avoir au moins une bonne réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset})
                if type_question == 'QRL' and response_count != 1:
                    messages.error(request, "Un QRL doit avoir exactement une réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset})
                if type_question == 'QRL' and correct_count != 1:
                    messages.error(request, "Un QRL doit avoir exactement une bonne réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset})
                question.instructeur = request.user
                question.save()
                reponse_formset.instance = question
                reponse_formset.save()
                messages.success(request, 'Question créée avec succès.')
            return redirect('celica_web:question_list')
    else:
        form = QuestionForm()
        reponse_formset = ReponseFormSet()
    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset})

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def question_form_edit(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    # Vérifier si la question est utilisée dans un test actif
    now = timezone.now()
    active_tests = Test.objects.filter(questions=question)
    for test in active_tests:
        active_plannings = Planning.objects.filter(test=test, date_debut__lte=now, date_fin__gte=now, statut__in=['planifie', 'en_cours'])
        if active_plannings.exists():
            messages.error(request, "Cette question est utilisée dans un test actif et ne peut pas être modifiée.")
            return redirect('celica_web:question_list')
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES, instance=question)
        reponse_formset = ReponseFormSet(request.POST, instance=question)
        if form.is_valid() and reponse_formset.is_valid():
            question = form.save(commit=False)
            type_question = form.cleaned_data['type_question']
            
            # Compter seulement les réponses valides (non vides et non supprimées)
            response_count = 0
            correct_count = 0
            for form in reponse_formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    if form.cleaned_data.get('texte', '').strip():  # Réponse non vide
                        response_count += 1
                        if form.cleaned_data.get('est_correcte'):
                            correct_count += 1
            
            if request.POST.get('save_to_db'):
                if type_question == 'QCM' and response_count < 2:
                    messages.error(request, "Un QCM doit avoir au moins 2 réponses.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset, 'question': question})
                if type_question == 'QCM' and correct_count == 0:
                    messages.error(request, "Un QCM doit avoir au moins une bonne réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset, 'question': question})
                if type_question == 'QRL' and response_count != 1:
                    messages.error(request, "Un QRL doit avoir exactement une réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset, 'question': question})
                if type_question == 'QRL' and correct_count != 1:
                    messages.error(request, "Un QRL doit avoir exactement une bonne réponse.")
                    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset, 'question': question})
                question.save()
                reponse_formset.instance = question
                reponse_formset.save()
                messages.success(request, 'Question modifiée avec succès.')
            return redirect('celica_web:question_list')
    else:
        form = QuestionForm(instance=question)
        reponse_formset = ReponseFormSet(instance=question)
    return render(request, 'celicaweb/question_form.html', {'form': form, 'reponse_formset': reponse_formset, 'question': question})
@login_required
def question_list(request):
    # Déterminer les questions selon le rôle
    is_admin = request.user.is_superuser or request.user.groups.filter(name='Administrateurs').exists()
    
    if is_admin:
        # Les administrateurs voient toutes les questions
        all_questions = Question.objects.select_related('module', 'instructeur').all()
        total_questions = all_questions.count()
    else:
        # Les instructeurs voient seulement leurs questions
        all_questions = Question.objects.filter(instructeur=request.user)
        total_questions = all_questions.count()
    
    # Questions affichées (après filtrage)
    questions = all_questions
    
    # Filtrage par recherche
    search = request.GET.get('search', '')
    if search:
        questions = questions.filter(enonce__icontains=search)
    
    # Filtrage par type de question
    question_type = request.GET.get('type', '')
    if question_type:
        questions = questions.filter(type_question=question_type)
    
    # Filtrage par module
    module_id = request.GET.get('module', '')
    if module_id:
        questions = questions.filter(module_id=module_id)
    
    # Calculer les statistiques
    qcm_count = all_questions.filter(type_question='QCM').count()
    qrl_count = all_questions.filter(type_question='QRL').count()
    modules = Module.objects.all().order_by('intitule')
    
    # Pagination
    questions = questions.order_by('-date_creation')  # Plus récentes en premier
    paginator = Paginator(questions, 12)  # 12 questions par page
    page_number = request.GET.get('page')
    questions_page = paginator.get_page(page_number)
    
    # Gestion de l'ajout de questions à un test
    if request.method == 'POST':
        selected_questions = request.POST.getlist('selected_questions')
        if selected_questions:
            test_id = request.session.get('current_test_id')
            if test_id:
                test = Test.objects.get(id=test_id)
                for question_id in selected_questions:
                    question = Question.objects.get(id=question_id)
                    test.questions.add(question)
                messages.success(request, "Questions ajoutées au test avec succès.")
                return redirect('celica_web:instructeur_dashboard')
            else:
                messages.error(request, "Aucun test sélectionné pour ajouter les questions.")
    
    context = {
        'questions': questions_page,
        'total_questions': total_questions,
        'qcm_count': qcm_count,
        'qrl_count': qrl_count,
        'modules': modules,
        'is_admin': is_admin,
        'search_query': search,
        'selected_type': question_type,
        'selected_module': module_id,
    }
    
    return render(request, 'celicaweb/question_list.html', context)

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def importer_questions(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    form = ImportQuestionsForm()
    if request.method == 'POST':
        form = ImportQuestionsForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = form.cleaned_data['fichier']
            try:
                if fichier.name.endswith('.csv'):
                    df = pd.read_csv(fichier)
                else:
                    df = pd.read_excel(fichier)
                questions = []
                for idx, row in df.iterrows():
                    type_question = row['type_question']
                    if type_question not in ['QCM', 'QRL']:
                        messages.error(request, f"Type de question invalide : {type_question}")
                        continue
                    question = Question(
                        enonce=row['enonce'],
                        type_question=type_question,
                        niveau_difficulte=row.get('niveau_difficulte', 'facile'),
                        module_id=row.get('module_id', test.module.id),
                        test=test,
                        instructeur=request.user
                    )
                    if request.POST.get('save_to_db'):
                        question.save()
                        questions.append((idx + 1, question))  # Stocker la question avec son index
                    else:
                        continue
                    if type_question == 'QRL':
                        if pd.notna(row.get('reponse_1')):
                            Reponse.objects.create(question=question, texte=row['reponse_1'], est_correcte=True)
                        else:
                            messages.error(request, f"QRL sans réponse : {row['enonce']}")
                            continue
                    else:
                        response_count = 0
                        correct_count = 0
                        for i in range(1, 6):
                            response_key = f'reponse_{i}'
                            correct_key = f'est_correcte_{i}'
                            if response_key in row and pd.notna(row[response_key]):
                                response_count += 1
                                est_correcte = row.get(correct_key, False)
                                Reponse.objects.create(question=question, texte=row[response_key], est_correcte=est_correcte)
                                if est_correcte:
                                    correct_count += 1
                        if response_count < 2:
                            messages.error(request, f"QCM avec moins de 2 réponses : {row['enonce']}")
                            continue
                        if correct_count == 0:
                            messages.error(request, f"QCM sans bonne réponse : {row['enonce']}")
                            continue
                if form.cleaned_data['images']:
                    with zipfile.ZipFile(form.cleaned_data['images'], 'r') as zip_ref:
                        zip_ref.extractall('media/questions/images/')
                        for idx, question in questions:
                            # Associer l'image si elle existe (par exemple, question_1.jpg)
                            image_path = os.path.join('media/questions/images/', f'question_{idx}.jpg')
                            if os.path.exists(image_path):
                                question.image = os.path.join('questions/images/', f'question_{idx}.jpg')
                                question.save()
                messages.success(request, 'Questions importées avec succès.')
                return redirect('celica_web:test_form', test_id=test.id)
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {str(e)}")
    return render(request, 'celicaweb/importer_questions.html', {'form': form, 'test': test})

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def importer_question_depuis_fichier(request):
    form = ImportQuestionsForm()
    if request.method == 'POST':
        form = ImportQuestionsForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = form.cleaned_data['fichier']
            try:
                if fichier.name.endswith('.csv'):
                    df = pd.read_csv(fichier)
                else:
                    df = pd.read_excel(fichier)
                for _, row in df.iterrows():
                    type_question = row['type_question']
                    if type_question not in ['QCM', 'QRL']:
                        messages.error(request, f"Type de question invalide : {type_question}")
                        continue
                    module_id = row.get('module_id')
                    module = Module.objects.get(id=module_id) if module_id else None
                    if not module:
                        messages.error(request, f"Module non spécifié ou introuvable pour la question : {row['enonce']}")
                        continue
                    question = Question(
                        enonce=row['enonce'],
                        type_question=type_question,
                        niveau_difficulte=row.get('niveau_difficulte', 'facile'),
                        module=module
                    )
                    if request.POST.get('save_to_db'):
                        question.save()
                    if type_question == 'QRL':
                        if pd.notna(row.get('reponse_1')):
                            Reponse.objects.create(question=question, texte=row['reponse_1'], est_correcte=True)
                        else:
                            messages.error(request, f"QRL sans réponse : {row['enonce']}")
                            continue
                    else:
                        response_count = 0
                        correct_count = 0
                        for i in range(1, 6):
                            response_key = f'reponse_{i}'
                            correct_key = f'est_correcte_{i}'
                            if response_key in row and pd.notna(row[response_key]):
                                response_count += 1
                                est_correcte = row.get(correct_key, False)
                                Reponse.objects.create(question=question, texte=row[response_key], est_correcte=est_correcte)
                                if est_correcte:
                                    correct_count += 1
                        if response_count < 2:
                            messages.error(request, f"QCM avec moins de 2 réponses : {row['enonce']}")
                            continue
                        if correct_count == 0:
                            messages.error(request, f"QCM sans bonne réponse : {row['enonce']}")
                            continue
                if form.cleaned_data['images']:
                    with zipfile.ZipFile(form.cleaned_data['images'], 'r') as zip_ref:
                        zip_ref.extractall('media/questions/images/')
                messages.success(request, 'Questions importées avec succès.')
                return redirect('celica_web:question_list')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {str(e)}")
    return render(request, 'celicaweb/importer_question_depuis_fichier.html', {'form': form})

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def exporter_test(request, pk, format_fichier):
    test = get_object_or_404(Test, pk=pk)
    try:
        data = test.exporter(format_fichier)
        response = HttpResponse(data, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="test_{pk}.{format_fichier}"'
        return response
    except ValueError as e:
        messages.error(request, f"Erreur lors de l'exportation : {str(e)}")
        return redirect('celica_web:test_list')

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def importer_test_depuis_fichier(request, format_fichier):
    modules = Module.objects.all()
    if request.method == 'POST':
        module_id = request.POST.get('module_id')
        module = get_object_or_404(Module, id=module_id) if module_id else None
        fichier = request.FILES.get('fichier')
        if fichier:
            test = Test.importer_depuis_fichier(fichier, module, format_fichier)
            messages.success(request, "Test importé avec succès.")
            return redirect('celica_web:test_list')
        messages.error(request, "Aucun fichier fourni.")
    return render(request, 'celicaweb/importer_test.html', {'modules': modules})

# Fonction utilitaire pour vérifier le rôle apprenant
def check_apprenant_role(user):
    return hasattr(user, 'role') and user.role == 'apprenant'

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponseForbidden, JsonResponse
from django.utils import timezone
import random
import logging

logger = logging.getLogger(__name__)

@login_required
@permission_required('celica_web.passer_tests', raise_exception=True)
def passer_test(request, test_id):
    # S'assurer que le token CSRF est généré
    from django.middleware.csrf import get_token
    get_token(request)
    test = get_object_or_404(Test, id=test_id)
    
    # Vérifier si l'utilisateur est un apprenant
    if not check_apprenant_role(request.user):
        return HttpResponseForbidden("Accès interdit.")

    # Vérifier si l'apprenant a déjà passé ce test
    existing_result = Resultat.objects.filter(test=test, apprenant=request.user).first()
    if existing_result:
        messages.warning(request, "Vous avez déjà passé ce test.")
        return redirect('celica_web:mes_resultats')

    # Vérifier si le test est accessible selon le planning
    now = timezone.now()
    groupes_apprenant = request.user.groupes.all()
    
    if groupes_apprenant.exists():
        # Vérifier si le test a un planning accessible
        planning = test.plannings.filter(
            groupe__in=groupes_apprenant,
            statut__in=['planifie', 'en_cours']
        ).order_by('date_debut').first()
        
        if planning:
            # Vérifier si le test est accessible maintenant
            is_accessible = planning.date_debut <= now <= planning.date_fin
            if not is_accessible:
                if now < planning.date_debut:
                    messages.error(request, f"Ce test n'est pas encore accessible. Il sera disponible à partir du {planning.date_debut.strftime('%d/%m/%Y à %H:%M')}.")
                else:
                    messages.error(request, f"Ce test n'est plus accessible. Il était disponible jusqu'au {planning.date_fin.strftime('%d/%m/%Y à %H:%M')}.")
                return redirect('celica_web:apprenant_tests')
        else:
            messages.error(request, "Aucun planning valide trouvé pour ce test.")
            return redirect('celica_web:apprenant_tests')
    else:
        messages.error(request, "Vous n'êtes assigné à aucun groupe pour ce test.")
        return redirect('celica_web:apprenant_tests')

    # Récupérer les questions du test
    questions = list(test.questions.all())
    if not questions:
        messages.error(request, f"Le test '{test.titre}' ne contient aucune question. Veuillez contacter l'administrateur.")
        return redirect('celica_web:apprenant_dashboard')

    # Gérer l'ordre des questions dans la session
    session_key = f'test_questions_order_{test_id}'
    if session_key not in request.session:
        if test.randomize_questions:
            random.shuffle(questions)
        else:
            questions = sorted(questions, key=lambda q: (q.ordre, q.id))
        request.session[session_key] = [q.id for q in questions]
        request.session.modified = True
    else:
        # Restaurer l'ordre des questions depuis la session
        question_ids = request.session.get(session_key, [])
        questions = [get_object_or_404(Question, id=qid) for qid in question_ids if qid in [q.id for q in questions]]

    # Stocker l'ID du test actuel dans la session
    request.session['current_test_id'] = test_id
    request.session.modified = True
    
    # Gérer le chronomètre
    start_time_key = f'test_start_time_{test_id}'
    if start_time_key not in request.session:
        request.session[start_time_key] = timezone.now().timestamp()
        request.session.modified = True
        
        # RÉINITIALISER LE COMPTEUR DE VIOLATIONS AU DÉBUT DU TEST
        violation_count_key = f'violation_count_{test_id}_{request.user.id}'
        if violation_count_key in request.session:
            del request.session[violation_count_key]
        request.session[violation_count_key] = 0
        request.session.modified = True
        print(f"🔄 Compteur de violations réinitialisé pour le test {test_id} et l'utilisateur {request.user.id}")

    # VÉRIFIER LES VIOLATIONS DE SÉCURITÉ
    violation_count_key = f'violation_count_{test_id}_{request.user.id}'
    violation_count = request.session.get(violation_count_key, 0)
    if violation_count >= 3:
        print(f"🚨 VIOLATIONS DÉPASSÉES ({violation_count}/3) - REDIRECTION VERS TEST TERMINÉ")
        # Nettoyer les données de session
        test_keys_to_clean = [key for key in request.session.keys() if key.startswith('test_') or key.startswith('violation_')]
        for key in test_keys_to_clean:
            del request.session[key]
        request.session.modified = True
        return redirect('celica_web:test_terminated')

    elapsed_seconds = int(timezone.now().timestamp() - request.session[start_time_key])
    total_seconds = test.duree * 60
    temps_restant_seconds = max(0, total_seconds - elapsed_seconds)

    # Si le temps est écoulé, finaliser le test
    if temps_restant_seconds <= 0:
        return finaliser_test(request, test, questions)

    # Déterminer la question actuelle
    question_id = request.GET.get('question')
    if question_id:
        try:
            question = get_object_or_404(Question, id=question_id, test=test)
            question_index = next(i for i, q in enumerate(questions) if q.id == question.id)
        except (Question.DoesNotExist, StopIteration):
            messages.error(request, "Question introuvable.")
            question = questions[0]
            question_index = 0
    else:
        question = questions[0]
        question_index = 0

    # Gérer les soumissions POST
    if request.method == 'POST':
        action = request.POST.get('action')
        question_id = request.POST.get('question_id')

        # Initialiser la structure de session pour les réponses
        if 'test_responses' not in request.session:
            request.session['test_responses'] = {}
        if str(test_id) not in request.session['test_responses']:
            request.session['test_responses'][str(test_id)] = {}

        # Sauvegarder la réponse
        if question_id:
            try:
                question = get_object_or_404(Question, id=question_id, test=test)
                if question.type_question == 'QCM':
                    reponse_selectionnee = request.POST.get('reponse_qcm')
                    if reponse_selectionnee:
                        request.session['test_responses'][str(test_id)][question_id] = [reponse_selectionnee]
                    else:
                        request.session['test_responses'][str(test_id)][question_id] = []
                elif question.type_question == 'QRL':
                    reponse_libre = request.POST.get('reponse_libre', '').strip()
                    request.session['test_responses'][str(test_id)][question_id] = reponse_libre
                request.session.modified = True
            except Question.DoesNotExist:
                messages.error(request, "Question introuvable.")
                return redirect(f"/test/passer/{test_id}/?question={questions[0].id}")

        # Gérer les actions
        if action == 'save_only':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Réponse sauvegardée'})
            messages.success(request, "Réponse sauvegardée.")
            return redirect(f"/test/passer/{test_id}/?question={question_id}")

        elif action == 'save_and_next':
            try:
                current_index = next(i for i, q in enumerate(questions) if q.id == int(question_id))
                if current_index < len(questions) - 1:
                    next_question = questions[current_index + 1]
                    return redirect(f"/test/passer/{test_id}/?question={next_question.id}")
                else:
                    return finaliser_test(request, test, questions)
            except StopIteration:
                messages.error(request, "Erreur de navigation.")
                return redirect(f"/test/passer/{test_id}/?question={questions[0].id}")

        elif action == 'finish_test':
            return finaliser_test(request, test, questions)

    # Préparer le contexte pour le template
    responses_data = request.session.get('test_responses', {}).get(str(test_id), {})
    score_actuel = round((len(responses_data) / len(questions)) * 100, 1) if len(questions) > 0 else 0
    reponses_utilisateur = responses_data.get(str(question.id), []) if question.type_question == 'QCM' else []
    reponse_libre = responses_data.get(str(question.id), '') if question.type_question == 'QRL' else ''

    # --- RANDOMISATION DES RÉPONSES QCM ---
    reponses_ordonnees = []
    if question.type_question == 'QCM':
        reponses = list(question.reponses.all())
        session_reponses_key = f'test_{test_id}_question_{question.id}_reponses_order'
        if session_reponses_key not in request.session:
            random.shuffle(reponses)
            request.session[session_reponses_key] = [r.id for r in reponses]
            request.session.modified = True
        else:
            # Restaurer l'ordre depuis la session
            reponse_ids = request.session[session_reponses_key]
            reponses_dict = {r.id: r for r in reponses}
            reponses = [reponses_dict[rid] for rid in reponse_ids if rid in reponses_dict]
        reponses_ordonnees = reponses

    context = {
        'test': test,
        'question': question,
        'question_actuelle_numero': question_index + 1,
        'total_questions': len(questions),
        'progression_pourcentage': round(((question_index + 1) / len(questions)) * 100, 1),
        'question_precedente': questions[question_index - 1] if question_index > 0 else None,
        'question_suivante': questions[question_index + 1] if question_index < len(questions) - 1 else None,
        'reponses_utilisateur': reponses_utilisateur,
        'reponse_libre': reponse_libre,
        'temps_restant_seconds': temps_restant_seconds,
        'alphabet': 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',
        'score_actuel': score_actuel,
        'reponses_ordonnees': reponses_ordonnees,  # Ajouté pour le template
    }

    return render(request, 'celicaweb/passer_test_single.html', context)


def finaliser_test(request, test, questions):
    """Finalise le test et calcule le score"""
    print(f"🎯 FINALISATION DU TEST {test.id} POUR {request.user.email}")
    
    # Nettoyer les données de session liées aux violations
    violation_count_key = f'violation_count_{test.id}_{request.user.id}'
    if violation_count_key in request.session:
        del request.session[violation_count_key]
        request.session.modified = True
        print(f"🧹 Compteur de violations nettoyé pour le test {test.id}")
    
    # Récupérer les réponses depuis la session
    responses_data = request.session.get('test_responses', {}).get(str(test.id), {})
    print(f"📊 Réponses récupérées: {len(responses_data)} questions répondues sur {len(questions)}")
    
    # Calculer le score
    score = 0
    total_points = 0
    details_reponses = []
    
    for question in questions:
        question_points = question.ponderation or 1
        total_points += question_points
        
        # Récupérer la réponse de l'utilisateur
        user_response = responses_data.get(str(question.id))
        
        if question.type_question == 'QCM':
            if user_response and isinstance(user_response, list) and len(user_response) > 0:
                # Vérifier si la réponse sélectionnée est correcte
                reponse_selectionnee = user_response[0]
                try:
                    reponse_obj = Reponse.objects.get(id=reponse_selectionnee, question=question)
                    if reponse_obj.est_correcte:
                        score += question_points
                        details_reponses.append({
                            'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                            'reponse_utilisateur': reponse_obj.texte[:30] + '...' if len(reponse_obj.texte) > 30 else reponse_obj.texte,
                            'correcte': True,
                            'points': question_points
                        })
                    else:
                        details_reponses.append({
                            'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                            'reponse_utilisateur': reponse_obj.texte[:30] + '...' if len(reponse_obj.texte) > 30 else reponse_obj.texte,
                            'correcte': False,
                            'points': 0
                        })
                except Reponse.DoesNotExist:
                    details_reponses.append({
                        'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                        'reponse_utilisateur': 'Réponse non trouvée',
                        'correcte': False,
                        'points': 0
                    })
            else:
                details_reponses.append({
                    'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                    'reponse_utilisateur': 'Aucune réponse',
                    'correcte': False,
                    'points': 0
                })
        
        elif question.type_question == 'QRL':
            if user_response and user_response.strip():
                # Pour les QRL, on peut implémenter une logique de scoring plus complexe
                # Pour l'instant, on donne les points si une réponse a été fournie
                score += question_points
                details_reponses.append({
                    'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                    'reponse_utilisateur': user_response[:50] + '...' if len(user_response) > 50 else user_response,
                    'correcte': True,
                    'points': question_points
                })
            else:
                details_reponses.append({
                    'question': question.enonce[:50] + '...' if len(question.enonce) > 50 else question.enonce,
                    'reponse_utilisateur': 'Aucune réponse',
                    'correcte': False,
                    'points': 0
                })
    
    # Calculer le pourcentage et la note sur 20
    pourcentage = (score / total_points * 100) if total_points > 0 else 0
    note_sur_20 = round((score / total_points * 20), 2) if total_points > 0 else 0
    
    # Déterminer l'appréciation
    if pourcentage >= 80:
        appreciation = "Excellent"
    elif pourcentage >= 70:
        appreciation = "Très bien"
    elif pourcentage >= 60:
        appreciation = "Bien"
    elif pourcentage >= 50:
        appreciation = "Assez bien"
    elif pourcentage >= 40:
        appreciation = "Passable"
    else:
        appreciation = "Insuffisant"
    
    print(f"📈 SCORE FINAL: {score}/{total_points} ({pourcentage:.1f}%) - Note: {note_sur_20}/20 - {appreciation}")
    
    # Calculer le temps passé
    start_time_key = f'test_start_time_{test.id}'
    if start_time_key in request.session:
        elapsed_seconds = int(timezone.now().timestamp() - request.session[start_time_key])
        temps_passe_minutes = max(1, round(elapsed_seconds / 60))
    else:
        temps_passe_minutes = test.duree or 1

    # Créer le résultat
    resultat = Resultat.objects.create(
        test=test,
        apprenant=request.user,
        score=score,
        note_sur_20=note_sur_20,
        appreciation=appreciation,
        temps_passe=temps_passe_minutes,
        details_reponses=details_reponses
    )

    # Nettoyer la session
    if 'test_responses' in request.session and str(test.id) in request.session['test_responses']:
        del request.session['test_responses'][str(test.id)]
    if f'test_questions_order_{test.id}' in request.session:
        del request.session[f'test_questions_order_{test.id}']
    if start_time_key in request.session:
        del request.session[start_time_key]
    request.session.modified = True

    # Stocker les informations du résultat dans la session pour l'affichage
    request.session['dernier_resultat'] = {
        'test_titre': test.titre,
        'note_sur_20': note_sur_20,
        'appreciation': appreciation,
        'score': score,
        'total_points': total_points,
        'details_reponses': details_reponses
    }
    request.session.modified = True
    
    return redirect('celica_web:resultat_test_termine', resultat_id=resultat.id)

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def resultat_test_termine(request, resultat_id):
    """Affiche les résultats immédiatement après la fin du test"""
    resultat = get_object_or_404(Resultat, id=resultat_id, apprenant=request.user)
    
    # Récupérer les détails depuis la session si disponibles
    dernier_resultat_session = request.session.get('dernier_resultat', {})
    
    # Calculer le pourcentage correct basé sur la note sur 20
    # Pourcentage = (note_sur_20 / 20) * 100
    pourcentage = (resultat.note_sur_20 / 20) * 100 if resultat.note_sur_20 else 0
    
    # Déterminer la couleur de l'appréciation
    couleur_appreciation = {
        'excellent': 'success',
        'tres_bien': 'success', 
        'bien': 'info',
        'assez_bien': 'warning',
        'passable': 'warning',
        'insuffisant': 'danger'
    }.get(resultat.appreciation, 'secondary')
    
    context = {
        'resultat': resultat,
        'pourcentage': pourcentage,
        'couleur_appreciation': couleur_appreciation,
        'details_session': dernier_resultat_session
    }
    
    # Nettoyer la session après affichage
    if 'dernier_resultat' in request.session:
        del request.session['dernier_resultat']
        request.session.modified = True
    
    return render(request, 'celicaweb/resultat_test_termine.html', context)

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def resultat_detail_instructeur(request, resultat_id):
    """Affiche les détails complets d'un résultat pour les instructeurs et administrateurs"""
    resultat = get_object_or_404(Resultat, id=resultat_id)
    
    # Vérifier les permissions
    if request.user.role == 'admin':
        # Les administrateurs peuvent voir tous les résultats
        pass
    elif request.user.role == 'instructeur':
        # Les instructeurs ne peuvent voir que les résultats de leurs tests
        if resultat.test.instructeur != request.user:
            messages.error(request, "Vous n'avez pas accès à ce résultat.")
            return redirect('celica_web:resultats_apprenants')
    else:
        # Les apprenants ne peuvent voir que leurs propres résultats
        if resultat.apprenant != request.user:
            messages.error(request, "Accès refusé.")
            return redirect('celica_web:mes_resultats')
    
    # Calculer le pourcentage correct basé sur la note sur 20
    pourcentage = (resultat.note_sur_20 / 20) * 100 if resultat.note_sur_20 else 0
    
    # Déterminer la couleur de l'appréciation
    couleur_appreciation = {
        'excellent': 'success',
        'tres_bien': 'success', 
        'bien': 'info',
        'assez_bien': 'warning',
        'passable': 'warning',
        'insuffisant': 'danger'
    }.get(resultat.appreciation, 'secondary')
    
    # Récupérer les questions du test avec leurs réponses
    questions = resultat.test.questions.all().prefetch_related('reponses')
    
    # Utiliser les détails des réponses stockés dans le modèle
    if resultat.details_reponses:
        details_reponses = resultat.details_reponses
    else:
        # Fallback si pas de détails stockés
        details_reponses = []
        for question in questions:
            reponses_correctes = []
            if question.type_question == 'QCM':
                reponses_correctes = [r.texte for r in question.reponses.filter(est_correcte=True)]
            elif question.type_question == 'QRL':
                reponse_correcte = question.reponses.filter(est_correcte=True).first()
                reponses_correctes = [reponse_correcte.texte] if reponse_correcte else []
            
            details_reponses.append({
                'question_id': question.id,
                'question_enonce': question.enonce,
                'question_type': question.type_question,
                'reponse_donnee': 'Non disponible (réponse non stockée)',
                'reponse_donnee_affichage': 'Non disponible',
                'reponses_correctes': reponses_correctes,
                'score_obtenu': 0,
                'score_max': question.ponderation
            })
    
    context = {
        'resultat': resultat,
        'pourcentage': pourcentage,
        'couleur_appreciation': couleur_appreciation,
        'details_session': {
            'test_titre': resultat.test.titre,
            'note_sur_20': resultat.note_sur_20,
            'appreciation': resultat.appreciation,
            'score': resultat.score,
            'total_points': resultat.test.bareme,
            'details_reponses': details_reponses
        },
        'is_instructeur_view': True
    }
    
    return render(request, 'celicaweb/resultat_test_termine.html', context)

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def all_cours_list(request):
    cours = Cours.objects.all()
    modules = Module.objects.all()
    module_id = request.GET.get('module')
    status = request.GET.get('status')

    if module_id:
        cours = cours.filter(module__id=module_id)
    if status:
        cours = cours.filter(status=status)

    context = {
        'cours': cours,
        'modules': modules,
        'selected_module': module_id,
        'selected_status': status,
    }
    return render(request, 'celicaweb/cours_list.html', context)

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def importer_cours_depuis_pdf(request):
    modules = Module.objects.all()

    if request.method == 'POST':
        form = ImportCoursForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = form.cleaned_data['fichier']
            module = form.cleaned_data['module']
            try:
                # Utiliser request.user comme instructeur
                cours = Cours.importer_depuis_pdf(fichier, module, request.user)
                groupes = module.groupes.all()
                for groupe in groupes:
                    apprenants = groupe.apprenants.all()
                    for apprenant in apprenants:
                        # Utiliser la méthode creer_notification
                        Notification.creer_notification(
                            titre=f"Nouveau cours ajouté : {cours.titre}",
                            message=f"Un nouveau cours '{cours.titre}' a été ajouté au module '{module.intitule}'.",
                            type_notice="info",  # Correspond à type_notice dans le modèle
                            utilisateur=apprenant,  # Correspond à utilisateur dans le modèle
                            module=module,  # Associer le module pour la validation
                            instructeur=request.user  # Associer l'instructeur
                        )
                messages.success(request, "Cours importé avec succès.")
                return redirect('celica_web:cours_list')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation du cours : {str(e)}")
        else:
            # Afficher les erreurs détaillées
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"Champ {field} : {error}")
            messages.error(request, "Erreur dans le formulaire : " + " | ".join(error_messages))
    else:
        form = ImportCoursForm()

    return render(request, 'celicaweb/importer_cours.html', {'form': form, 'modules': modules})
@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def cours_detail(request, pk):
    cours = get_object_or_404(Cours, pk=pk)
    return render(request, 'celicaweb/cours_details.html', {'cours': cours})

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def supprimer_cours(request, pk):
    cours = get_object_or_404(Cours, pk=pk)
    if request.method == 'POST':
        try:
            cours.delete()
            messages.success(request, "Cours supprimé avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression : {str(e)}")
        return redirect('celica_web:cours_list')
    return render(request, 'celicaweb/confirm_delete.html', {'cours': cours})

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def exporter_cours(request, pk, format_fichier):
    cours = get_object_or_404(Cours, pk=pk)
    try:
        if format_fichier == 'original':
            # Exportation du fichier original
            if not cours.fichier:
                messages.error(request, "Aucun fichier original à exporter")
                return redirect('celica_web:mes_cours')
            
            # Retourner le fichier original directement
            response = HttpResponse(cours.fichier, content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{cours.fichier.name.split("/")[-1]}"'
            return response
        else:
            # Exportation dans d'autres formats
            data = cours.exporter(format_fichier)
            response = HttpResponse(data, content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="cours_{pk}.{format_fichier}"'
            return response
    except ValueError as e:
        messages.error(request, f"Erreur lors de l'exportation : {str(e)}")
        return redirect('celica_web:mes_cours')

@login_required
@permission_required('celica_web.consulter_cours', raise_exception=True)
def mes_cours(request):
    # Récupérer les cours des modules des groupes où l'utilisateur est apprenant
    cours = Cours.objects.filter(module__groupes__groupes_apprenant=request.user).distinct()
    
    # Ajouter des filtres
    module_id = request.GET.get('module')
    if module_id:
        cours = cours.filter(module__id=module_id)
    
    modules = Module.objects.filter(groupes__groupes_apprenant=request.user).distinct()
    return render(request, 'celicaweb/mes_cours.html', {
        'cours': cours,
        'modules': modules,
        'selected_module': module_id
    })

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def test_resultats(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    
    # Vérifier que l'utilisateur a le droit de voir les résultats de ce test
    if request.user.role != 'admin' and test.instructeur != request.user:
        messages.error(request, "Vous n'avez pas accès aux résultats de ce test.")
        return redirect('celica_web:test_list')
    
    resultats = Resultat.objects.filter(test=test)
    return render(request, 'celicaweb/test_resultats.html', {'test': test, 'resultats': resultats})

# Fonction utilitaire pour vérifier le rôle apprenant
def check_apprenant_role(user):
    return hasattr(user, 'role') and user.role == 'apprenant'

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def mes_resultats(request):
    if not check_apprenant_role(request.user):
        messages.error(request, "Accès refusé : vous n'êtes pas un apprenant.")
        return redirect('celica_web:visitor_index')
    
    # Récupérer tous les résultats de l'apprenant
    resultats = Resultat.objects.filter(apprenant=request.user).select_related('test', 'test__module').order_by('-date_passation')
    
    # Appliquer les filtres si fournis
    module_id = request.GET.get('module')
    test_id = request.GET.get('test')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if module_id:
        try:
            resultats = resultats.filter(test__module__id=int(module_id))
        except (ValueError, TypeError):
            pass
    
    if test_id:
        try:
            resultats = resultats.filter(test__id=int(test_id))
        except (ValueError, TypeError):
            pass
    
    if date_debut:
        try:
            from datetime import datetime
            date_debut_parsed = datetime.strptime(date_debut, '%Y-%m-%d').date()
            resultats = resultats.filter(date_passation__date__gte=date_debut_parsed)
        except ValueError:
            pass
    
    if date_fin:
        try:
            from datetime import datetime
            date_fin_parsed = datetime.strptime(date_fin, '%Y-%m-%d').date()
            resultats = resultats.filter(date_passation__date__lte=date_fin_parsed)
        except ValueError:
            pass
    
    # Récupérer les options pour les filtres
    modules = Module.objects.filter(tests__resultats__apprenant=request.user).distinct()
    tests = Test.objects.filter(resultats__apprenant=request.user).distinct()
    
    # Calculer des statistiques pour l'apprenant et préparer les données pour le template
    resultats_avec_pourcentage = []
    scores_pourcentages = []
    
    for resultat in resultats:
        # Calculer le pourcentage basé sur la note sur 20
        # Pourcentage = (note_sur_20 / 20) * 100
        pourcentage = (resultat.note_sur_20 / 20) * 100 if resultat.note_sur_20 else 0
        
        resultats_avec_pourcentage.append((resultat, round(pourcentage, 1)))
        scores_pourcentages.append(pourcentage)
    
    # Pagination
    paginator = Paginator(resultats_avec_pourcentage, 10)  # 10 résultats par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    if resultats.exists():
        stats = {
            'total_tests': resultats.count(),
            'score_moyen': sum(scores_pourcentages) / len(scores_pourcentages) if scores_pourcentages else 0,
            'meilleur_score': max(scores_pourcentages) if scores_pourcentages else 0,
            'dernier_test': resultats.first().date_passation if resultats.exists() else None
        }
    else:
        stats = {
            'total_tests': 0,
            'score_moyen': 0,
            'meilleur_score': 0,
            'dernier_test': None
        }
    
    return render(request, 'celicaweb/mes_resultats.html', {
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'modules': modules,
        'tests': tests,
        'stats': stats,
        'selected_module': module_id,
        'selected_test': test_id,
        'date_debut': date_debut,
        'date_fin': date_fin
    })

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def resultats_apprenants(request):
    utilisateur = request.user
    
    # Filtrage des résultats selon le rôle
    if utilisateur.role == 'admin':
        # Les administrateurs voient tous les résultats
        resultats = Resultat.objects.all()
        tests = Test.objects.all()
        modules = Module.objects.all()
        apprenants = Utilisateur.objects.filter(role='apprenant')
    else:
        # Les instructeurs voient uniquement les résultats de leurs tests
        resultats = Resultat.objects.filter(test__instructeur=utilisateur)
        tests = Test.objects.filter(instructeur=utilisateur)
        modules = Module.objects.filter(tests__instructeur=utilisateur).distinct()
        apprenants = Utilisateur.objects.filter(resultats__test__instructeur=utilisateur, role='apprenant').distinct()
    
    # Ajouter des filtres
    test_id = request.GET.get('test')
    module_id = request.GET.get('module')
    apprenant_id = request.GET.get('apprenant')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if test_id:
        resultats = resultats.filter(test__id=test_id)
    if module_id:
        resultats = resultats.filter(test__module__id=module_id)
    if apprenant_id:
        resultats = resultats.filter(apprenant__id=apprenant_id)
    if date_debut:
        resultats = resultats.filter(date_passation__gte=date_debut)
    if date_fin:
        resultats = resultats.filter(date_passation__lte=date_fin)
    
    # Exportation
    if request.GET.get('export'):
        format_fichier = request.GET.get('export')
        if format_fichier in ['csv', 'excel']:
            data = []
            # Regrouper et trier les résultats par test, puis par apprenant
            resultats_ordonnes = sorted(resultats, key=lambda r: (r.test.titre, r.apprenant.last_name, r.apprenant.first_name))
            for resultat in resultats_ordonnes:
                data.append({
                    'Test': resultat.test.titre,
                    'Apprenant': f"{resultat.apprenant.first_name} {resultat.apprenant.last_name}",
                    'Note': round(resultat.note_sur_20, 2) if resultat.note_sur_20 is not None else '',
                    'Date': resultat.date_passation,
                    'Appréciation': resultat.get_appreciation_display() if hasattr(resultat, 'get_appreciation_display') else resultat.appreciation
                })
            df = pd.DataFrame(data)
            # Correction pour Excel : convertir les dates avec timezone en format compatible
            if format_fichier == 'excel' and 'Date' in df.columns:
                df['Date'] = df['Date'].astype(str)
            if format_fichier == 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="resultats.csv"'
                df.to_csv(response, index=False)
            else:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="resultats.xlsx"'
                df.to_excel(response, index=False)
            return response
    
    return render(request, 'celicaweb/resultats_apprenants.html', {
        'resultats': resultats,
        'tests': tests,
        'modules': modules,
        'apprenants': apprenants,
        'selected_test': test_id,
        'selected_module': module_id,
        'selected_apprenant': apprenant_id,
        'date_debut': date_debut,
        'date_fin': date_fin
    })
@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def resultat_form(request):
    if request.method == 'POST':
        form = ResultatForm(request.POST)
        if form.is_valid():
            resultat = form.save(commit=False)
            resultat.appreciation = "Réussi" if resultat.score >= 50 else "Échoué"
            resultat.calculer_note(resultat.score)
            resultat.save()
            messages.success(request, "Résultat créé avec succès.")
            return redirect('celica_web:mes_resultats')
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = ResultatForm()
    return render(request, 'celicaweb/resultat_form.html', {'form': form})

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def exporter_resultat(request, pk, format_fichier):
    resultat = get_object_or_404(Resultat, pk=pk)
    try:
        data = resultat.exporter(format_fichier)
        response = HttpResponse(data, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="resultat_{pk}.{format_fichier}"'
        return response
    except ValueError as e:
        messages.error(request, f"Erreur lors de l'exportation : {str(e)}")
        return redirect('celica_web:mes_resultats')

@login_required
@permission_required('celica_web.gerer_plannings', raise_exception=True)
def planning_form(request, planning_id=None):
    planning = get_object_or_404(Planning, id=planning_id) if planning_id else None
    
    if request.method == 'POST':
        form = PlanningForm(request.POST, instance=planning)
        if form.is_valid():
            planning = form.save()
            
            # Collecter tous les apprenants des groupes concernés (via le champ FK uniquement)
            apprenants_notifier = set()
            
            # Via le champ `groupe` (ForeignKey)
            if planning.groupe:
                apprenants_notifier.update(planning.groupe.groupes_apprenant.filter(role='apprenant'))
            
            # Envoyer notifications aux apprenants
            for apprenant in apprenants_notifier:
                Notification.objects.create(
                    titre=f"Planning {'modifié' if planning_id else 'créé'}",
                    utilisateur=apprenant,
                    message=f"{'Modification du' if planning_id else 'Nouveau'} planning pour {planning.test.titre if planning.test else 'Session'} du {planning.date_debut} au {planning.date_fin}.",
                    type_notice='planning'
                )
            
            action = "modifié" if planning_id else "créé"
            messages.success(request, f"Planning {action} avec succès et notifications envoyées.")
            return redirect('celica_web:planning_list')
    else:
        form = PlanningForm(instance=planning)
    
    context = {
        'form': form,
        'planning': planning,
        'is_edit': planning_id is not None
    }
    return render(request, 'celicaweb/planning_form.html', context)

@login_required
@permission_required('celica_web.gerer_plannings', raise_exception=True)
def supprimer_planning(request, planning_id):
    planning = get_object_or_404(Planning, id=planning_id)
    if request.method == 'POST':
        # Supprimer le planning (plus de relation M2M à nettoyer)
        planning.delete()
        messages.success(request, "Planning supprimé avec succès.")
        return redirect('celica_web:planning_list')
    return render(request, 'celicaweb/confirm_delete_planning.html', {'planning': planning})

@login_required
@permission_required('celica_web.gerer_plannings', raise_exception=True)
def planning_list(request):
    utilisateur = request.user

    # Filtrage des plannings selon le rôle
    if utilisateur.role == 'admin':
        plannings = Planning.objects.select_related('test', 'groupe').all()
    else:
        plannings = Planning.objects.filter(
            models.Q(instructeur_responsable=utilisateur) |
            models.Q(test__instructeur=utilisateur)
        ).distinct()

    module_id = request.GET.get('module')
    if module_id:
        plannings = plannings.filter(test__module__id=module_id)

    context = {
        'plannings': plannings,
        'modules': Module.objects.all(),
        'selected_module': module_id,
        'nb_planifies': plannings.filter(statut='planifie').count(),
        'nb_en_cours': plannings.filter(statut='en_cours').count(),
        'nb_termines': plannings.filter(statut='termine').count(),
    }
    return render(request, 'celicaweb/planning_list.html', context)

@login_required
@permission_required('celica_web.gerer_plannings', raise_exception=True)
def exporter_planning(request, pk, format_fichier):
    planning = get_object_or_404(Planning, pk=pk)
    try:
        data = planning.exporter(format_fichier)
        response = HttpResponse(data, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="planning_{pk}.{format_fichier}"'
        return response
    except ValueError as e:
        messages.error(request, f"Erreur lors de l'exportation : {str(e)}")
        return redirect('celica_web:planning_list')

@login_required
@permission_required('celica_web.gerer_groupes', raise_exception=True)
def groupe_form(request, groupe_id=None):
    groupe = get_object_or_404(Groupe, id=groupe_id) if groupe_id else None
    if request.method == 'POST':
        form = GroupeForm(request.POST, instance=groupe)
        if form.is_valid():
            try:
                groupe = form.save()
                # Créer des notifications pour les apprenants ajoutés au groupe
                for apprenant in groupe.apprenants.all():
                    Notification.objects.create(
                        titre="Ajout au groupe",
                        utilisateur=apprenant,
                        message=f"Vous avez été ajouté au groupe {groupe.nom}.",
                        type_notice='info'
                    )
                
                action = "créé" if not groupe_id else "mis à jour"
                messages.success(request, f"Groupe {action} avec succès et notifications envoyées.")
                return redirect('celica_web:groupe_list')
            except Exception as e:
                messages.error(request, f"Erreur lors de la sauvegarde du groupe : {str(e)}")
        else:
            # Afficher les erreurs de validation
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans le champ {field}: {error}")
    else:
        form = GroupeForm(instance=groupe)
    
    context = {
        'form': form,
        'groupe': groupe,
        'is_edit': groupe_id is not None
    }
    return render(request, 'celicaweb/groupe_form.html', context)

@login_required
@permission_required('celica_web.gerer_groupes', raise_exception=True)
def supprimer_groupe(request, groupe_id):
    groupe = get_object_or_404(Groupe, id=groupe_id)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # --- Sécuriser la suppression ------------------------------------
                # 1. Vider toutes les relations Many-To-Many afin d'éviter les
                #    contraintes d'intégrité.
                groupe.groupes_apprenant.clear()   # Apprenants
                groupe.instructeurs.clear()        # Instructeurs
                groupe.modules.clear()             # Modules liés

                # 2. Plus de relation M2M à détacher (champ groupes supprimé)
                # Cette étape n'est plus nécessaire

                # 3. Mettre à NULL la FK "groupe" des plannings qui pointent dessus.
                Planning.objects.filter(groupe=groupe).update(groupe=None)

                # 4. Supprimer enfin le groupe
                groupe.delete()
        except IntegrityError:
            # Il peut rester d'anciennes relations M2M héritées (ex. table
            # `celica_web_groupe_apprenants`) qui ne sont plus gérées par l'ORM.
            # On les purge manuellement puis on retente la suppression.
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM celica_web_groupe_apprenants WHERE groupe_id = %s",
                    [groupe.id],
                )
            # Nouvelle tentative
            try:
                groupe.delete()
            except IntegrityError as e:
                messages.error(
                    request,
                    "Impossible de supprimer le groupe : des dépendances subsistent ({}).".format(e),
                )
                return redirect('celica_web:groupe_list')

        messages.success(request, "Groupe supprimé avec succès.")
        return redirect('celica_web:groupe_list')
    return render(request, 'celicaweb/confirm_delete_groupe.html', {'groupe': groupe})

@login_required
@permission_required('celica_web.gerer_groupes', raise_exception=True)
def groupe_list(request):
    groupes = Groupe.objects.all()
    return render(request, 'celicaweb/groupe_list.html', {'groupes': groupes})

@login_required
@permission_required('celica_web.gerer_groupes', raise_exception=True)
def groupe_gerer_membres(request, groupe_id):
    groupe = get_object_or_404(Groupe, id=groupe_id)
    # Utiliser directement le champ ManyToMany pour éviter l'erreur QuerySet
    apprenants_disponibles = Utilisateur.objects.filter(role='apprenant').exclude(id__in=groupe.groupes_apprenant.all())
    instructeurs_disponibles = Utilisateur.objects.filter(role='instructeur').exclude(id__in=groupe.instructeurs.all())
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'ajouter_apprenant':
            apprenant_id = request.POST.get('apprenant_id')
            if apprenant_id:
                apprenant = get_object_or_404(Utilisateur, id=apprenant_id, role='apprenant')
                # Utiliser le champ ManyToMany direct au lieu de la propriété
                groupe.groupes_apprenant.add(apprenant)
                messages.success(request, f"Apprenant {apprenant.email} ajouté au groupe.")
            else:
                messages.error(request, "Veuillez sélectionner un apprenant.")
            return redirect('celica_web:groupe_gerer_membres', groupe_id=groupe.id)
        elif action == 'supprimer_apprenant':
            apprenant_id = request.POST.get('apprenant_id')
            apprenant = get_object_or_404(Utilisateur, id=apprenant_id, role='apprenant')
            # Utiliser le champ ManyToMany direct au lieu de la propriété
            groupe.groupes_apprenant.remove(apprenant)
            messages.success(request, f"Apprenant {apprenant.email} retiré du groupe.")
            return redirect('celica_web:groupe_gerer_membres', groupe_id=groupe.id)
        elif action == 'ajouter_instructeur':
            instructeur_id = request.POST.get('instructeur_id')
            if instructeur_id:
                instructeur = get_object_or_404(Utilisateur, id=instructeur_id, role='instructeur')
                groupe.instructeurs.add(instructeur)
                messages.success(request, f"Instructeur {instructeur.email} ajouté au groupe.")
            else:
                messages.error(request, "Veuillez sélectionner un instructeur.")
            return redirect('celica_web:groupe_gerer_membres', groupe_id=groupe.id)
        elif action == 'supprimer_instructeur':
            instructeur_id = request.POST.get('instructeur_id')
            instructeur = get_object_or_404(Utilisateur, id=instructeur_id, role='instructeur')
            groupe.instructeurs.remove(instructeur)
            messages.success(request, f"Instructeur {instructeur.email} retiré du groupe.")
            return redirect('celica_web:groupe_gerer_membres', groupe_id=groupe.id)
    return render(request, 'celicaweb/groupe_gerer_membres.html', {
        'groupe': groupe, 'apprenants_disponibles': apprenants_disponibles, 'instructeurs_disponibles': instructeurs_disponibles
    })

@login_required
@permission_required('celica_web.gerer_modules', raise_exception=True)
def module_form(request, module_id=None):
    module = get_object_or_404(Module, id=module_id) if module_id else None
    # Vérifier si le module est lié à des tests ou plannings
    if module:
        linked_tests = Test.objects.filter(module=module)
        linked_plannings = Planning.objects.filter(test__module=module)
        if linked_tests.exists() or linked_plannings.exists():
            messages.warning(request, "Ce module est lié à des tests ou plannings. Soyez prudent lors de la modification.")
    
    if request.method == 'POST':
        form = ModuleForm(request.POST, instance=module)
        if form.is_valid():
            module = form.save()
            # Envoyer une notification aux utilisateurs concernés (apprenants des groupes + instructeur principal)
            utilisateurs_concernes = []
            
            # Récupérer l'instructeur principal s'il existe
            if module.instructeur_principal:
                utilisateurs_concernes.append(module.instructeur_principal)
            
            # Récupérer tous les apprenants des groupes associés
            for groupe in module.groupes.all():
                apprenants_groupe = groupe.groupes_apprenant.filter(role='apprenant')
                utilisateurs_concernes.extend(apprenants_groupe)
            
            # Envoyer les notifications
            for user in utilisateurs_concernes:
                Notification.objects.create(
                    titre=f"Module {'modifié' if module_id else 'créé'}",
                    utilisateur=user,
                    message=f"Le module {module.intitule} a été {'modifié' if module_id else 'créé'}.",
                    module=module
                )
            
            messages.success(request, f"Module {'modifié' if module_id else 'créé'} avec succès. Les utilisateurs concernés ont été notifiés.")
            return redirect('celica_web:module_list')
    else:
        form = ModuleForm(instance=module)
    return render(request, 'celicaweb/module_form.html', {'form': form})

@login_required
@permission_required('celica_web.gerer_modules', raise_exception=True)
def supprimer_module(request, module_id):
    module = get_object_or_404(Module, id=module_id)
    if request.method == 'POST':
        try:
            # Vérifier s'il y a des tests liés avant suppression
            tests_lies = module.tests.all()
            if tests_lies.exists():
                messages.error(request, f"Impossible de supprimer le module '{module.intitule}' car il contient {tests_lies.count()} test(s). Supprimez d'abord les tests associés.")
                return redirect('celica_web:module_list')
            
            module_nom = module.intitule
            module.delete()
            messages.success(request, f"Module '{module_nom}' supprimé avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression : {str(e)}")
        return redirect('celica_web:module_list')
    return render(request, 'celicaweb/confirm_delete_module.html', {'module': module})

@login_required
@permission_required('celica_web.gerer_modules', raise_exception=True)
def module_list(request):
    utilisateur = request.user

    # Filtrage des modules selon le rôle
    if utilisateur.role == 'admin':
        modules = Module.objects.all()
    else:
        modules = Module.objects.filter(
            models.Q(instructeur_principal=utilisateur) |
            models.Q(groupes__instructeurs=utilisateur)
        ).distinct()

    context = {
        'modules': modules,
    }
    return render(request, 'celicaweb/module_list.html', context)

@login_required
@permission_required('celica_web.gerer_notifications', raise_exception=True)
def notification_form(request, notification_id=None):
    notification = get_object_or_404(Notification, id=notification_id) if notification_id else None
    if request.method == 'POST':
        form = NotificationForm(request.POST, instance=notification)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.instructeur = request.user
            notification.save()
            # Envoyer une notification aux utilisateurs concernés (apprenants des groupes liés au module)
            if notification.module:
                for groupe in notification.module.groupes.all():
                    apprenants = groupe.get_apprenants()
                    for apprenant in apprenants:
                        Notification.objects.create(
                            titre=f"Notification: {notification.titre}",
                            utilisateur=apprenant,
                            message=f"Nouvelle notification: {notification.message}",
                            module=notification.module
                        )
            messages.success(request, f"Notification {'modifiée' if notification_id else 'créée'} avec succès et envoyée aux utilisateurs concernés.")
            return redirect('celica_web:notification_list')
    else:
        form = NotificationForm(instance=notification)
    return render(request, 'celicaweb/notification_form.html', {'form': form})

@login_required
@permission_required('celica_web.gerer_notifications', raise_exception=True)
def notification_form_edit(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    if request.method == 'POST':
        form = NotificationForm(request.POST, instance=notification)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.instructeur = request.user
            notification.save()
            # Envoyer une notification mise à jour aux utilisateurs concernés
            if notification.module:
                for groupe in notification.module.groupes.all():
                    apprenants = groupe.get_apprenants()
                    for apprenant in apprenants:
                        Notification.objects.create(
                            titre=f"Mise à jour: {notification.titre}",
                            utilisateur=apprenant,
                            message=f"Mise à jour de notification: {notification.message}",
                            module=notification.module
                        )
            messages.success(request, "Notification modifiée avec succès et mise à jour envoyée aux utilisateurs concernés.")
            return redirect('celica_web:notification_list')
    else:
        form = NotificationForm(instance=notification)
    return render(request, 'celicaweb/notification_form.html', {'form': form})

@login_required
def notification_list(request):
    # Récupérer toutes les notifications de l'utilisateur
    notifications = Notification.objects.filter(utilisateur=request.user).order_by('-date_envoi')
    
    # Calculer les statistiques
    total_notifications = notifications.count()
    notifications_non_lues = notifications.filter(est_lue=False).count()
    notifications_lues = notifications.filter(est_lue=True).count()
    
    # Notifications récentes (moins de 24h)
    from django.utils import timezone
    from datetime import timedelta
    date_limite = timezone.now() - timedelta(hours=24)
    notifications_recentes = notifications.filter(date_envoi__gte=date_limite).count()
    
    # Limiter l'affichage à 5 notifications pour la liste
    notifications_affichage = notifications[:5]
    
    context = {
        'notifications': notifications_affichage,
        'total_notifications': total_notifications,
        'notifications_non_lues': notifications_non_lues,
        'notifications_lues': notifications_lues,
        'notifications_recentes': notifications_recentes,
    }
    return render(request, 'celicaweb/notification_list.html', context)

def marquer_lue(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    if request.user == notification.utilisateur:
        notification.marquer_comme_lue()
    return redirect('celica_web:notification_list')

@login_required
@permission_required('celica_web.gerer_notifications', raise_exception=True)
def supprimer_notification(request, id):
    notification = get_object_or_404(Notification, id=id)
    if request.method == 'POST':
        notification.delete()
        messages.success(request, "Notification supprimée avec succès.")
        return redirect('celica_web:notification_list')
    return render(request, 'celicaweb/confirm_delete_notification.html', {'notification': notification})

@login_required
@permission_required('celica_web.consulter_notifications', raise_exception=True)
def lister_notifications(request):
    notifications = Notification.objects.filter(destinataire=request.user).order_by('-date_creation')
    
    # Ajouter des filtres
    type_notification = request.GET.get('type')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if type_notification:
        notifications = notifications.filter(type_notification=type_notification)
    if date_debut:
        notifications = notifications.filter(date_creation__gte=timezone.datetime.strptime(date_debut, '%Y-%m-%d'))
    if date_fin:
        notifications = notifications.filter(date_creation__lte=timezone.datetime.strptime(date_fin, '%Y-%m-%d'))
    
    return render(request, 'celicaweb/lister_notifications.html', {
        'notifications': notifications,
        'type_notification': type_notification,
        'date_debut': date_debut,
        'date_fin': date_fin
    })

@login_required
@permission_required('celica_web.consulter_notifications', raise_exception=True)
def marquer_notification_lue(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, destinataire=request.user)
    if request.method == 'POST':
        notification.marquer_comme_lue()
        messages.success(request, "Notification marquée comme lue.")
        return redirect('celica_web:lister_notifications')
    return render(request, 'celicaweb/confirm_action.html', {'notification': notification})

@login_required
@permission_required('celica_web.acceder_aide', raise_exception=True)
def consulter_aide(request):
    aides = Aide.objects.all()
    
    if request.method == 'POST':
        sujet = request.POST.get('sujet')
        message = request.POST.get('message')
        if sujet and message:
            # Envoyer une notification à l'admin pour la demande de support
            admin_user = Utilisateur.objects.filter(role='admin').first()
            if admin_user:
                Notification.objects.create(
                    titre=f"Demande de support: {sujet}",
                    utilisateur=admin_user,
                    message=f"Nouvelle demande de support de {request.user.email}: {sujet}\n{message}",
                    type_notice="urgence"
                )
            messages.success(request, "Votre demande de support a été envoyée.")
            return redirect('celica_web:consulter_aide')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire.")
    
    return render(request, 'celicaweb/consulter_aide.html', {'aides': aides})

@login_required
@permission_required('celica_web.acceder_aide', raise_exception=True)
def rechercher_aide(request):
    mot_cle = request.GET.get('mot_cle', '')
    aides = Aide.rechercher_aide(mot_cle) if mot_cle else Aide.objects.all()
    return render(request, 'celicaweb/rechercher_aide.html', {'aides': aides, 'mot_cle': mot_cle})

def consulter_a_propos(request):
    try:
        # Essayer de récupérer l'objet APropos
        a_propos = APropos.objects.first()
        if not a_propos:
            # Si aucun objet n'existe, en créer un en mémoire
            a_propos = APropos(
                version="1.0.0",
                nom_application="CelicaWeb",
                description="Application de gestion des tests QCM pour la CELICA Maintenance de l'ASECNA.",
                organisme="ASECNA",
                contact_email="admin@celicaweb.com",
                date_mise_a_jour=timezone.now()
            )
    except OperationalError as e:
        # Si la table a un schéma incorrect (ex: colonne manquante), créer un objet en mémoire
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Erreur de base de données lors de la consultation de APropos: {e}. Utilisation des valeurs par défaut.")
        a_propos = APropos(
            version="1.0.0",
            nom_application="CelicaWeb", 
            description="Application de gestion des tests QCM pour la CELICA Maintenance de l'ASECNA.",
            organisme="ASECNA",
            contact_email="admin@celicaweb.com",
            date_mise_a_jour=timezone.now()
        )
    
    return render(request, 'celicaweb/consulter_a_propos.html', {'a_propos': a_propos})

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def gerer_utilisateurs(request):
    # Récupérer le paramètre de recherche
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '')
    statut_filter = request.GET.get('statut', '')
    
    # Base queryset
    utilisateurs = Utilisateur.objects.all()
    
    # Appliquer les filtres
    if search_query:
        utilisateurs = utilisateurs.filter(
            Q(last_name__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(matricule__icontains=search_query)
        )
    
    if role_filter:
        utilisateurs = utilisateurs.filter(role=role_filter)
    
    if statut_filter:
        utilisateurs = utilisateurs.filter(statut=statut_filter)
    
    # Grouper les utilisateurs par statut
    utilisateurs_par_statut = {
        'actif': utilisateurs.filter(statut='actif').order_by('last_name', 'first_name'),
        'inactif': utilisateurs.filter(statut='inactif').order_by('last_name', 'first_name'),
        'suspendu': utilisateurs.filter(statut='suspendu').order_by('last_name', 'first_name'),
    }
    
    # Statistiques
    total_utilisateurs = utilisateurs.count()
    total_actifs = utilisateurs.filter(statut='actif').count()
    total_inactifs = utilisateurs.filter(statut='inactif').count()
    total_suspendus = utilisateurs.filter(statut='suspendu').count()
    
    context = {
        'utilisateurs_par_statut': utilisateurs_par_statut,
        'search_query': search_query,
        'role_filter': role_filter,
        'statut_filter': statut_filter,
        'stats': {
            'total': total_utilisateurs,
            'actifs': total_actifs,
            'inactifs': total_inactifs,
            'suspendus': total_suspendus,
        },
        'roles': Utilisateur.ROLES,
        'statuts': Utilisateur.STATUTS,
    }
    
    return render(request, 'celicaweb/utilisateur_list.html', context)

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def utilisateur_add_view(request):
    if request.method == 'POST':
        form = UtilisateurForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Utilisateur ajouté avec succès.")
            return redirect('celica_web:gerer_utilisateurs')
    else:
        form = UtilisateurForm()
    return render(request, 'celicaweb/utilisateur_add.html', {'form': form})



@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def rechercher_utilisateur(request):
    mot_cle = request.GET.get('mot_cle', '')
    utilisateurs = Utilisateur.objects.rechercher(mot_cle)
    return render(request, 'celicaweb/rechercher_utilisateur.html', {'utilisateurs': utilisateurs, 'mot_cle': mot_cle})

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def modifier_profil(request, user_id):
    user = get_object_or_404(Utilisateur, id=user_id)
    if request.method == 'POST':
        form = UtilisateurForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profil modifié avec succès.")
            return redirect('celica_web:gerer_utilisateurs')
    else:
        form = UtilisateurForm(instance=user)
    return render(request, 'celicaweb/modifier_profil.html', {'form': form, 'user': user})

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def changer_mot_de_passe(request, user_id):
    utilisateur = get_object_or_404(Utilisateur, pk=user_id)
    if request.method == 'POST':
        form = ChangerMotDePasseForm(request.POST)
        if form.is_valid():
            nouveau_mot_de_passe = form.cleaned_data['nouveau_mot_de_passe']
            utilisateur.set_password(nouveau_mot_de_passe)
            utilisateur.doit_changer_mot_de_passe = False
            utilisateur.save()
            messages.success(request, "Le mot de passe a été changé avec succès.")
            return redirect('celica_web:login')
    else:
        form = ChangerMotDePasseForm()
    return render(request, 'celicaweb/changer_mot_de_passe.html', {'form': form, 'utilisateur': utilisateur})

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def supprimer_utilisateur(request, user_id):
    user = get_object_or_404(Utilisateur, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, "Utilisateur supprimé avec succès.")
        return redirect('celica_web:gerer_utilisateurs')
    return render(request, 'celicaweb/confirm_delete.html', {'user': user})

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def consulter_statistiques(request):
    # Génération automatique si la table est vide ou incomplète
    tests = Test.objects.all()
    for test in tests:
        generer_ou_maj_statistiques(test)
    statistiques = Statistiques.objects.all()
    
    # Debug: afficher les taux individuels
    if statistiques.exists():
        taux_list = [s.taux_reussite for s in statistiques if s.taux_reussite is not None]
        moyenne_generale = sum(taux_list) / len(taux_list) if taux_list else 0
        
        # Debug: afficher les détails
        print(f"DEBUG - Nombre de statistiques: {len(statistiques)}")
        for i, stat in enumerate(statistiques):
            print(f"DEBUG - Stat {i+1}: Test='{stat.test.titre}', Taux={stat.taux_reussite}%, Participants={stat.nombre_participants}")
        print(f"DEBUG - Taux list: {taux_list}")
        print(f"DEBUG - Moyenne calculée: {moyenne_generale}")
    else:
        moyenne_generale = 0
        print("DEBUG - Aucune statistique trouvée")
    
    return render(request, 'celicaweb/consulter_statistiques.html', {
        'statistiques': statistiques,
        'moyenne_generale': moyenne_generale,
    })

@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def exporter_statistiques(request, pk, format_fichier):
    statistique = get_object_or_404(Statistiques, pk=pk)
    test = statistique.test
    if format_fichier == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="statistiques_test_{test.id}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Test', 'Module', 'Taux de réussite', 'Score moyen', 'Score médian', 'Participants', 'Période'])
        writer.writerow([
            test.titre,
            test.module.intitule,
            f"{statistique.taux_reussite:.1f}%",
            f"{statistique.score_moyen:.2f}",
            f"{statistique.score_median:.2f}",
            statistique.nombre_participants,
            f"{statistique.periode_debut:%d/%m/%Y} - {statistique.periode_fin:%d/%m/%Y}"
        ])
        return response
    elif format_fichier == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="statistiques_test_{test.id}.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques"
        headers = ['Test', 'Module', 'Taux de réussite', 'Score moyen', 'Score médian', 'Participants', 'Période']
        ws.append(headers)
        ws.append([
            test.titre,
            test.module.intitule,
            f"{statistique.taux_reussite:.1f}%",
            f"{statistique.score_moyen:.2f}",
            f"{statistique.score_median:.2f}",
            statistique.nombre_participants,
            f"{statistique.periode_debut:%d/%m/%Y} - {statistique.periode_fin:%d/%m/%Y}"
        ])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        wb.save(response)
        return response
    else:
        return HttpResponse("Format non supporté", status=400)

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def remplacer_question(request, test_id, question_id):
    test = get_object_or_404(Test, id=test_id)
    ancienne_question = get_object_or_404(Question, id=question_id, test=test)
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES)
        reponse_formset = ReponseFormSet(request.POST)
        if form.is_valid() and reponse_formset.is_valid():
            enonce = form.cleaned_data['enonce']
            niveau_difficulte = form.cleaned_data['niveau_difficulte']
            type_question = form.cleaned_data['type_question']
            reponses = [{'texte': rf.cleaned_data['texte'], 'est_correcte': rf.cleaned_data['est_correcte']} for rf in reponse_formset if rf.cleaned_data]
            try:
                test.remplacer_question(ancienne_question_id=question_id, enonce=enonce, niveau_difficulte=niveau_difficulte, type_question=type_question, reponses=reponses)
                messages.success(request, "La question a été remplacée avec succès.")
                return redirect('celica_web:question_list')
            except Exception as e:
                messages.error(request, f"Erreur lors du remplacement de la question : {str(e)}")
    else:
        form = QuestionForm(instance=ancienne_question)
        reponse_formset = ReponseFormSet(instance=ancienne_question)
    return render(request, 'celicaweb/remplacer_question.html', {'form': form, 'reponse_formset': reponse_formset, 'test': test, 'question': ancienne_question})

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def ajouter_questions_existantes(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    # Vérifier si le test est lié à une session active
    now = timezone.now()
    active_plannings = Planning.objects.filter(test=test, date_debut__lte=now, date_fin__gte=now, is_published=True)
    if active_plannings.exists():
        messages.error(request, "Ce test est lié à une session active et ne peut pas être modifié.")
        return redirect('celica_web:test_list')

    # Récupérer les questions disponibles : banque + questions d'autres tests du même module
    existing_questions = Question.objects.filter(
        Q(test__isnull=True, module=test.module) |  # Questions de la banque du même module
        Q(test__isnull=False, module=test.module)   # Questions d'autres tests du même module
    ).exclude(test=test)  # Exclure les questions déjà dans ce test
    messages.info(request, f"Nombre de questions disponibles : {existing_questions.count()}")

    # Créer un formset pour chaque question existante
    existing_questions_formsets = {}
    for question in existing_questions:
        formset_prefix = f'reponse_{question.id}'
        if request.method == 'POST':
            existing_questions_formsets[question.id] = ReponseFormSet(request.POST, prefix=formset_prefix, instance=question)
        else:
            existing_questions_formsets[question.id] = ReponseFormSet(prefix=formset_prefix, instance=question)

    if request.method == 'POST':
        form = TestForm(request.POST, request.FILES, instance=test)
        import_form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            test = form.save()

            # Gestion des questions en fonction de la source
            question_source = request.POST.get('question_source')

            if question_source == 'select':
                # Sélection de questions existantes
                question_ids = request.POST.getlist('existing_questions')
                if question_ids:
                    questions = Question.objects.filter(id__in=question_ids)
                    for question in questions:
                        # Mettre à jour les réponses existantes
                        for reponse in question.reponses.all():
                            texte_key = f'existing_reponse_{question.id}_{reponse.id}_texte'
                            est_correcte_key = f'existing_reponse_{question.id}_{reponse.id}_est_correcte'
                            if texte_key in request.POST:
                                reponse.texte = request.POST[texte_key]
                                reponse.est_correcte = request.POST.get(est_correcte_key) == 'True'
                                reponse.save()
                        # Ajouter des réponses supplémentaires
                        reponse_formset = existing_questions_formsets.get(question.id)
                        if reponse_formset and reponse_formset.is_valid():
                            reponse_formset.save()
                    test.questions.add(*questions)
                    messages.info(request, f"{len(questions)} question(s) existante(s) ajoutée(s) au test.")

            elif question_source == 'manual':
                # Ajout de nouvelles questions si saisie manuelle
                new_question_form = QuestionForm(request.POST, prefix='new_question')
                new_reponse_formset = ReponseFormSet(request.POST, prefix='new_reponse')
                if new_question_form.is_valid() and new_reponse_formset.is_valid():
                    question = new_question_form.save(commit=False)
                    type_question = new_question_form.cleaned_data['type_question']
                    response_count = sum(1 for form in new_reponse_formset if form.cleaned_data.get('texte'))
                    correct_count = sum(1 for form in new_reponse_formset if form.cleaned_data.get('est_correcte'))

                    # === AJOUT DE LA VALIDATION STRICTE ===
                    from .views import valider_ponderation_stricte
                    autorise_ajout, message = valider_ponderation_stricte(test, question.ponderation)
                    if not autorise_ajout:
                        messages.error(request, message)
                        return render(request, 'celicaweb/test_form.html', {
                            'form': form,
                            'existing_questions': existing_questions,
                            'existing_questions_formsets': existing_questions_formsets,
                            'new_question_form': new_question_form,
                            'new_reponse_formset': new_reponse_formset,
                            'import_form': import_form,
                            'test': test
                        })
                    # === FIN AJOUT ===

                    if type_question == 'QCM' and response_count < 2:
                        messages.error(request, "Un QCM doit avoir au moins 2 réponses.")
                        return render(request, 'celicaweb/test_form.html', {
                            'form': form,
                            'existing_questions': existing_questions,
                            'existing_questions_formsets': existing_questions_formsets,
                            'new_question_form': new_question_form,
                            'new_reponse_formset': new_reponse_formset,
                            'import_form': import_form,
                            'test': test
                        })
                    if type_question == 'QCM' and correct_count == 0:
                        messages.error(request, "Un QCM doit avoir au moins une bonne réponse.")
                        return render(request, 'celicaweb/test_form.html', {
                            'form': form,
                            'existing_questions': existing_questions,
                            'existing_questions_formsets': existing_questions_formsets,
                            'new_question_form': new_question_form,
                            'new_reponse_formset': new_reponse_formset,
                            'import_form': import_form,
                            'test': test
                        })
                    if type_question == 'QRL' and response_count != 1:
                        messages.error(request, "Un QRL doit avoir exactement une réponse.")
                        return render(request, 'celicaweb/test_form.html', {
                            'form': form,
                            'existing_questions': existing_questions,
                            'existing_questions_formsets': existing_questions_formsets,
                            'new_question_form': new_question_form,
                            'new_reponse_formset': new_reponse_formset,
                            'import_form': import_form,
                            'test': test
                        })
                    if type_question == 'QRL' and correct_count != 1:
                        messages.error(request, "Un QRL doit avoir exactement une bonne réponse.")
                        return render(request, 'celicaweb/test_form.html', {
                            'form': form,
                            'existing_questions': existing_questions,
                            'existing_questions_formsets': existing_questions_formsets,
                            'new_question_form': new_question_form,
                            'new_reponse_formset': new_reponse_formset,
                            'import_form': import_form,
                            'test': test
                        })
                    question.module = test.module
                    question.instructeur = test.instructeur
                    question.save()
                    new_reponse_formset.instance = question
                    new_reponse_formset.save()
                    test.questions.add(question)
                    messages.success(request, "Nouvelle question ajoutée et associée au test.")
                    return redirect('celica_web:test_form_edit', test_id=test.id)

            elif question_source == 'import' and import_form.is_valid():
                # Gestion de l'importation
                import_file = import_form.cleaned_data['file']
                import_format = import_form.cleaned_data['import_format']
                if import_format == 'csv':
                    df = pd.read_csv(import_file)
                else:
                    df = pd.read_excel(import_file)
                for _, row in df.iterrows():
                    type_question = row['type_question']
                    if type_question not in ['QCM', 'QRL']:
                        messages.error(request, f"Type de question invalide : {type_question}")
                        continue
                    question = Question(
                        enonce=row['enonce'],
                        type_question=type_question,
                        niveau_difficulte=row.get('niveau_difficulte', 'facile'),
                        ponderation=row.get('ponderation', 1),  # Valeur par défaut si non spécifiée
                        module=test.module,
                        test=test,
                        instructeur=test.instructeur
                    )
                    if request.POST.get('save_imported_to_db'):
                        question.save()
                    if type_question == 'QRL':
                        if pd.notna(row.get('reponse_1')):
                            Reponse.objects.create(
                                question=question,
                                texte=row['reponse_1'],
                                est_correcte=True
                            )
                        else:
                            messages.error(request, f"QRL sans réponse : {row['enonce']}")
                            continue
                    else:
                        response_count = 0
                        correct_count = 0
                        for i in range(1, 6):
                            response_key = f'reponse_{i}'
                            correct_key = f'est_correcte_{i}'
                            if response_key in row and pd.notna(row[response_key]):
                                response_count += 1
                                est_correcte = row.get(correct_key, False)
                                Reponse.objects.create(
                                    question=question,
                                    texte=row[response_key],
                                    est_correcte=est_correcte
                                )
                                if est_correcte:
                                    correct_count += 1
                        if response_count < 2:
                            messages.error(request, f"QCM avec moins de 2 réponses : {row['enonce']}")
                            continue
                        if correct_count == 0:
                            messages.error(request, f"QCM sans bonne réponse : {row['enonce']}")
                            continue
                if import_form.cleaned_data.get('import_images'):
                    with zipfile.ZipFile(import_form.cleaned_data['import_images'], 'r') as zip_ref:
                        zip_ref.extractall('media/questions/images/')

            messages.success(request, "Test mis à jour avec succès.")
            return redirect('celica_web:test_list')
    else:
        form = TestForm(instance=test)
        new_question_form = QuestionForm(prefix='new_question')
        new_reponse_formset = ReponseFormSet(prefix='new_reponse')
        import_form = ImportForm()

    return render(request, 'celicaweb/test_form.html', {
        'form': form,
        'existing_questions': existing_questions,
        'existing_questions_formsets': existing_questions_formsets,
        'new_question_form': new_question_form,
        'new_reponse_formset': new_reponse_formset,
        'import_form': import_form,
        'test': test
    })

@login_required
def consulter_planning_apprenant(request):
    from django.db.models import Q
    groupes = request.user.groupes.all()
    
    if not groupes.exists():
        messages.warning(request, "Vous n'êtes assigné à aucun groupe. Contactez votre instructeur.")
        plannings = Planning.objects.none()
        modules = Module.objects.none()
    else:
        # Filtrer les plannings via la relation FK 'groupe' uniquement
        plannings = Planning.objects.filter(
            Q(groupe__in=groupes),
            Q(statut__in=['planifie', 'en_cours', 'termine'])  # Inclure les terminés pour l'historique
        ).distinct()
        
        modules = Module.objects.filter(
            tests__plannings__groupe__in=groupes
        ).distinct()
    
    # Filtres
    module_id = request.GET.get('module')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if module_id:
        plannings = plannings.filter(test__module__id=module_id)
    if date_debut:
        plannings = plannings.filter(date_debut__gte=date_debut)
    if date_fin:
        plannings = plannings.filter(date_fin__lte=date_fin)
    
    return render(request, 'celicaweb/consulter_planning_apprenant.html', {
        'plannings': plannings,
        'modules': modules,
        'selected_module': module_id,
        'selected_date_debut': date_debut,
        'selected_date_fin': date_fin
    })

@login_required
@permission_required('celica_web.passer_tests', raise_exception=True)
def apprenant_tests(request):
    if not check_apprenant_role(request.user):
        messages.error(request, "Accès refusé : vous n'êtes pas un apprenant.")
        return redirect('celica_web:visitor_index')
    
    from django.db.models import Q
    now = timezone.now()
    
    # Récupérer les groupes de l'apprenant
    groupes_apprenant = request.user.groupes.all()
    
    if groupes_apprenant.exists():
        # Récupérer les tests avec plannings pour l'apprenant
        # Seulement ceux qui sont accessibles maintenant OU dans les prochaines 24h
        from datetime import timedelta
        tomorrow = now + timedelta(days=1)
        
        # Debug: Afficher les groupes de l'apprenant
        print(f"🔍 Groupes de l'apprenant {request.user.email}: {[g.nom for g in groupes_apprenant]}")
        
        tests_with_plannings = Test.objects.filter(
            actif=True,
            plannings__groupe__in=groupes_apprenant,
            plannings__statut__in=['planifie', 'en_cours']
        ).exclude(
            resultats__apprenant=request.user
        ).select_related('module').prefetch_related(
            'plannings__groupe',
            'questions'
        ).distinct()
        
        # Debug: Afficher les tests trouvés
        print(f"🔍 Tests trouvés pour l'apprenant: {tests_with_plannings.count()}")
        for test in tests_with_plannings:
            plannings = test.plannings.filter(groupe__in=groupes_apprenant, statut__in=['planifie', 'en_cours'])
            print(f"  - Test: {test.titre} (ID: {test.id})")
            for planning in plannings:
                print(f"    Planning: {planning.titre} (ID: {planning.id}, Statut: {planning.statut}, Groupe: {planning.groupe.nom if planning.groupe else 'Aucun'})")
        
        # Traiter chaque test pour déterminer s'il est accessible
        tests = []
        for test in tests_with_plannings:
            # Récupérer le planning le plus proche pour ce test
            planning = test.plannings.filter(
                groupe__in=groupes_apprenant,
                statut__in=['planifie', 'en_cours']
            ).order_by('date_debut').first()
            
            if planning:
                # Vérifier si le test est accessible maintenant
                is_accessible = planning.date_debut <= now <= planning.date_fin
                
                # Ne montrer que les tests accessibles maintenant OU dans les prochaines 24h
                if is_accessible or (planning.date_debut <= tomorrow):
                    test.is_accessible = is_accessible
                    test.planning = planning
                    tests.append(test)
        
    else:
        # Si l'apprenant n'appartient à aucun groupe, aucun test disponible
        tests = Test.objects.none()
        messages.info(request, "Vous n'êtes assigné à aucun groupe. Contactez votre instructeur.")
    
    return render(request, 'celicaweb/apprenant_tests.html', {'tests': tests})

def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def admin_dashboard(request):
    utilisateur = request.user
    
    # Vérification des imports de modèles
    from django.db import models
    from .models import Utilisateur, Cours, Test, Question, Module, Groupe, Planning, Notification, Resultat, TestEventLog, SecurityViolation
    try:
        from .models import ResultatTest
    except ImportError:
        ResultatTest = None

    # === STATISTIQUES DÉTAILLÉES - CONTRÔLE TOTAL ADMINISTRATEUR ===
    try:
        
        # Comptes basiques - TOUS les éléments de la plateforme
        utilisateurs_count = Utilisateur.objects.count()
        cours_count = Cours.objects.count()
        tests_count = Test.objects.count()
        questions_count = Question.objects.count()
        modules_count = Module.objects.count()
        groupes_count = Groupe.objects.count()
        plannings_count = Planning.objects.count()
        notifications_count = Notification.objects.count()
        
        # Correction pour ResultatTest vs Resultat
        try:
            resultats_count = ResultatTest.objects.count()
        except:
            resultats_count = Resultat.objects.count()
        
        # Statistiques calculées silencieusement
        
        # Statistiques par rôle
        apprenants_count = Utilisateur.objects.filter(role='apprenant').count()
        instructeurs_count = Utilisateur.objects.filter(role='instructeur').count()
        admins_count = Utilisateur.objects.filter(is_staff=True).count()
        
        # Statistiques d'activité récente (30 derniers jours)
        from django.utils import timezone
        from datetime import timedelta
        last_30_days = timezone.now() - timedelta(days=30)
        
        nouveaux_utilisateurs = Utilisateur.objects.filter(date_joined__gte=last_30_days).count()
        nouveaux_tests = Test.objects.filter(date_creation__gte=last_30_days).count()
        nouveaux_cours = Cours.objects.filter(created_at__gte=last_30_days).count()
        
        # === STATISTIQUES DE JOURNALISATION ===
        # Événements de test
        total_events = TestEventLog.objects.count()
        events_30_days = TestEventLog.objects.filter(timestamp__gte=last_30_days).count()
        
        # Violations de sécurité
        total_violations = SecurityViolation.objects.count()
        violations_30_days = SecurityViolation.objects.filter(timestamp__gte=last_30_days).count()
        
        # Types d'événements les plus fréquents
        event_types_stats = TestEventLog.objects.values('event_type').annotate(
            count=models.Count('id')
        ).order_by('-count')[:5]
        
        # Types de violations les plus fréquentes
        violation_types_stats = SecurityViolation.objects.values('violation_type').annotate(
            count=models.Count('id')
        ).order_by('-count')[:5]
        
        # Tests les plus surveillés (avec le plus d'événements)
        tests_most_monitored = TestEventLog.objects.values('test__titre').annotate(
            event_count=models.Count('id')
        ).order_by('-event_count')[:5]
        
        # Utilisateurs avec le plus de violations
        users_most_violations = SecurityViolation.objects.values('utilisateur__email').annotate(
            violation_count=models.Count('id')
        ).order_by('-violation_count')[:5]
        
        # Statistiques des résultats
        if resultats_count > 0:
            try:
                # Essayer d'abord avec ResultatTest
                moyenne_generale = ResultatTest.objects.aggregate(
                    avg_score=models.Avg('score')
                )['avg_score'] or 0
                
                # Résultats récents avec ResultatTest
                resultats_recents = ResultatTest.objects.select_related(
                    'utilisateur', 'test'
                ).order_by('-date_passage')[:10]
            except:
                # Sinon utiliser Resultat
                moyenne_generale = Resultat.objects.aggregate(
                    avg_score=models.Avg('score')
                )['avg_score'] or 0
                
                # Résultats récents avec Resultat
                resultats_recents = Resultat.objects.select_related(
                    'apprenant', 'test'
                ).order_by('-date_passation')[:10]
        else:
            moyenne_generale = 0
            resultats_recents = []
            
        stats = {
            'total_utilisateurs': utilisateurs_count,
            'total_cours': cours_count,
            'total_tests': tests_count,
            'total_questions': questions_count,
            'total_modules': modules_count,
            'total_groupes': groupes_count,
            'total_plannings': plannings_count,
            'total_resultats': resultats_count,
            'total_notifications': notifications_count,
            'apprenants_count': apprenants_count,
            'instructeurs_count': instructeurs_count,
            'admins_count': admins_count,
            'nouveaux_utilisateurs': nouveaux_utilisateurs,
            'nouveaux_tests': nouveaux_tests,
            'nouveaux_cours': nouveaux_cours,
            'moyenne_generale': round(moyenne_generale, 2) if moyenne_generale else 0,
            # Statistiques de journalisation
            'total_events': total_events,
            'events_30_days': events_30_days,
            'total_violations': total_violations,
            'violations_30_days': violations_30_days,
            'event_types_stats': event_types_stats,
            'violation_types_stats': violation_types_stats,
            'tests_most_monitored': tests_most_monitored,
            'users_most_violations': users_most_violations,
        }
        
    except Exception as e:
        
        # En cas d'erreur, essayer de récupérer au moins les comptes basiques
        try:
            stats = {
                'total_utilisateurs': Utilisateur.objects.count(),
                'total_cours': Cours.objects.count(),
                'total_tests': Test.objects.count(),
                'total_questions': Question.objects.count(),
                'total_modules': Module.objects.count(),
                'total_groupes': Groupe.objects.count(),
                'total_plannings': Planning.objects.count(),
                'total_resultats': Resultat.objects.count(),
                'total_notifications': Notification.objects.count(),
                'apprenants_count': Utilisateur.objects.filter(role='apprenant').count(),
                'instructeurs_count': Utilisateur.objects.filter(role='instructeur').count(),
                'admins_count': Utilisateur.objects.filter(is_staff=True).count(),
                'nouveaux_utilisateurs': 0,
                'nouveaux_tests': 0,
                'nouveaux_cours': 0,
                'moyenne_generale': 0,
                # Statistiques de journalisation par défaut
                'total_events': 0,
                'events_30_days': 0,
                'total_violations': 0,
                'violations_30_days': 0,
                'event_types_stats': [],
                'violation_types_stats': [],
                'tests_most_monitored': [],
                'users_most_violations': [],
            }
        except Exception as e2:
            stats = {
                'total_utilisateurs': 0, 'total_cours': 0, 'total_tests': 0,
                'total_questions': 0, 'total_modules': 0, 'total_groupes': 0,
                'total_plannings': 0, 'total_resultats': 0, 'total_notifications': 0,
                'apprenants_count': 0, 'instructeurs_count': 0, 'admins_count': 0,
                'nouveaux_utilisateurs': 0, 'nouveaux_tests': 0, 'nouveaux_cours': 0,
                'moyenne_generale': 0,
                # Statistiques de journalisation par défaut
                'total_events': 0, 'events_30_days': 0, 'total_violations': 0,
                'violations_30_days': 0, 'event_types_stats': [], 'violation_types_stats': [],
                'tests_most_monitored': [], 'users_most_violations': [],
            }
        resultats_recents = []

    # === DONNÉES RÉCENTES POUR AFFICHAGE ===
    try:
        # Utilisateurs récents
        utilisateurs_recents = Utilisateur.objects.order_by('-date_joined')[:5]
        
        # Tests récents
        tests_recents = Test.objects.select_related('module').order_by('-date_creation')[:5]
        
        # Cours récents
        cours_recents = Cours.objects.order_by('-created_at')[:5]
        
        # Modules récents
        modules_recents = Module.objects.order_by('-id')[:5]
        
        # Notifications non lues
        notifications_non_lues = Notification.objects.filter(
            est_lue=False
        ).order_by('-date_envoi')[:5]
        
        # Événements récents de journalisation
        events_recents = TestEventLog.objects.select_related('test', 'utilisateur').order_by('-timestamp')[:5]
        
        # Violations récentes
        violations_recents = SecurityViolation.objects.select_related('utilisateur').order_by('-timestamp')[:5]
        
    except Exception as e:
        utilisateurs_recents = tests_recents = cours_recents = modules_recents = notifications_non_lues = events_recents = violations_recents = []

    # === GESTION DES FORMULAIRES ===
    utilisateur_form = UtilisateurForm()
    notification_form = NotificationForm()

    if request.method == 'POST':
        if request.POST.get('action') == 'ajouter_utilisateur':
            utilisateur_form = UtilisateurForm(request.POST)
            if utilisateur_form.is_valid():
                utilisateur_form.save()
                messages.success(request, "Utilisateur ajouté avec succès.")
                return redirect('celica_web:admin_dashboard')
            else:
                messages.error(request, "Erreur lors de l'ajout de l'utilisateur.")

        elif request.POST.get('action') == 'envoyer_notification':
            notification_form = NotificationForm(request.POST)
            if notification_form.is_valid():
                notification = notification_form.save(commit=False)
                notification.emetteur = utilisateur
                notification.save()
                messages.success(request, "Notification envoyée avec succès.")
                return redirect('celica_web:admin_dashboard')
            else:
                messages.error(request, "Erreur lors de l'envoi de la notification.")

    context = {
        'utilisateur': utilisateur,
        'stats': stats,
        'utilisateurs_recents': utilisateurs_recents,
        'tests_recents': tests_recents,
        'cours_recents': cours_recents,
        'modules_recents': modules_recents,
        'resultats_recents': resultats_recents,
        'notifications_non_lues': notifications_non_lues,
        'events_recents': events_recents,
        'violations_recents': violations_recents,
        'utilisateur_form': utilisateur_form,
        'notification_form': notification_form,
        # Garder l'ancienne interface pour la compatibilité
        'utilisateurs_count': utilisateurs_count,
        'cours_count': cours_count,
        'tests_count': tests_count,
    }
    return render(request, 'celicaweb/admin_dashboard.html', context)

def test_preview(request, test_id):
    """Vue pour afficher l'aperçu d'un test"""
    test = get_object_or_404(Test, id=test_id)
    questions = test.questions.all().prefetch_related('reponses')
    
    context = {
        'test': test,
        'questions': questions,
        'is_preview': True,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'celicaweb/test_preview_partial.html', context)
    
    return render(request, 'celicaweb/test_preview.html', context)

from django.http import JsonResponse

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)

@login_required
@require_http_methods(["GET"])
def ajax_questions_existantes(request):
    """
    Vue AJAX pour récupérer les questions existantes filtrées
    """
    try:
        module_id = request.GET.get('module_id', 'tous')
        test_id = request.GET.get('test_id')
        
        # Base query pour les questions disponibles
        questions_query = Question.objects.filter(actif=True)
        
        # Filtrer par module si spécifié
        if module_id != 'tous' and module_id:
            questions_query = questions_query.filter(module_id=module_id)
        
        # Si un test est spécifié, exclure les questions déjà dans ce test
        if test_id:
            try:
                test = Test.objects.get(id=test_id)
                
                # Exclure les questions déjà dans le test courant
                questions_deja_dans_test = test.questions.values_list('id', flat=True)
                questions_query = questions_query.exclude(id__in=questions_deja_dans_test)
                
                # Si aucun module spécifique n'est demandé, filtrer par le module du test
                if module_id == 'tous':
                    questions_query = questions_query.filter(module=test.module)
                    
            except Test.DoesNotExist:
                logger.warning(f"Test avec ID {test_id} non trouvé")
                return JsonResponse({'success': False, 'error': 'Test non trouvé'}, status=404)
        
        # Préparer les données à retourner
        questions_data = []
        
        for question in questions_query.select_related('module').prefetch_related('reponses'):
            question_data = {
                'id': question.id,
                'enonce': question.enonce,
                'type_question': question.type_question,
                'niveau_difficulte': question.niveau_difficulte,
                'ponderation': float(question.ponderation),
                'module_nom': question.module.intitule if question.module else 'Sans module',
                'source': 'Banque' if not question.test else 'Test',
                'reponses': []
            }
            
            # Ajouter les réponses pour les QCM
            if question.type_question == 'QCM':
                for reponse in question.reponses.all():
                    question_data['reponses'].append({
                        'texte': reponse.texte,
                        'est_correcte': reponse.est_correcte
                    })
            
            questions_data.append(question_data)
        
        return JsonResponse({
            'success': True,
            'questions': questions_data,
            'total': len(questions_data)
        })
        
    except Exception as e:
        logger.error(f"Erreur dans ajax_questions_existantes: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des questions: {str(e)}'
        }, status=500)


@login_required
@user_passes_test(lambda u: u.is_staff or u.role in ['admin', 'instructeur'])
def diagnostic_plannings(request):
    """Vue de diagnostic pour identifier les problèmes d'affichage des plannings"""
    
    # Informations de base
    diagnostic = {
        'user': request.user,
        'user_groups': request.user.groupes.all(),
        'total_plannings': Planning.objects.count(),
        'total_groups': Groupe.objects.count(),
        'total_apprenants': Utilisateur.objects.filter(role='apprenant').count(),
    }
    
    # Analyse des plannings
    plannings_analysis = []
    for planning in Planning.objects.all()[:10]:  # Limiter à 10 pour performance
        analysis = {
            'planning': planning,
            'groupe_fk': planning.groupe,
            'groupes_m2m': [],  # Field removed
            'apprenants_via_fk': [],
            'apprenants_via_m2m': [],
            'total_apprenants': 0,
        }
        
        # Apprenants via ForeignKey
        if planning.groupe:
            analysis['apprenants_via_fk'] = list(planning.groupe.groupes_apprenant.filter(role='apprenant'))
        
        # Apprenants via ManyToMany (field removed)
        # No longer applicable
        
        # Total unique
        all_apprenants = set(analysis['apprenants_via_fk'] + analysis['apprenants_via_m2m'])
        analysis['total_apprenants'] = len(all_apprenants)
        
        plannings_analysis.append(analysis)
    
    # Analyse des groupes
    groupes_analysis = []
    for groupe in Groupe.objects.all()[:5]:  # Limiter à 5
        analysis = {
            'groupe': groupe,
            'apprenants_count': groupe.groupes_apprenant.filter(role='apprenant').count(),
            'apprenants_list': list(groupe.groupes_apprenant.filter(role='apprenant')),
            'plannings_via_fk': Planning.objects.filter(groupe=groupe).count(),
            'plannings_via_m2m': 0,  # Field removed
        }
        groupes_analysis.append(analysis)
    
    # Test avec un utilisateur apprenant
    if request.user.role == 'apprenant' or request.GET.get('test_user'):
        from django.db.models import Q
        user_groups = request.user.groupes.all()
        
        # Test des requêtes des vues apprenants
        plannings_old_query = Planning.objects.filter(groupe__in=user_groups).distinct()
        plannings_new_query = Planning.objects.filter(
            Q(groupe__in=user_groups),
            Q(statut__in=['planifie', 'en_cours']),
            Q(date_fin__gte=timezone.now())
        ).distinct()
        
        diagnostic.update({
            'plannings_old_count': plannings_old_query.count(),
            'plannings_new_count': plannings_new_query.count(),
            'plannings_old_list': list(plannings_old_query[:5]),
            'plannings_new_list': list(plannings_new_query[:5]),
        })
    
    context = {
        'diagnostic': diagnostic,
        'plannings_analysis': plannings_analysis,
        'groupes_analysis': groupes_analysis,
    }
    
    return render(request, 'celicaweb/diagnostic_plannings.html', context)

# Import des nouveaux services
import os
import tempfile
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

# Import des services Excel
from .services.excel_import_service import ExcelImportService, ExcelImportValidator
from .utils.excel_generator import ExcelTemplateGenerator

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def download_excel_template(request):
    """Télécharge un template Excel pour l'import de tests"""
    try:
        generator = ExcelTemplateGenerator()
        
        # Générer le template avec exemples
        if request.GET.get('example') == 'true':
            buffer = generator.generate_example_template()
            filename = "template_test_avec_exemples.xlsx"
        else:
            buffer = generator.generate_test_template()
            filename = "template_test_vide.xlsx"
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du template: {str(e)}")
        return redirect('celica_web:test_list')

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def import_test_excel(request):
    """Importe un test depuis un fichier Excel"""
    if request.method == 'POST':
        try:
            # Vérifier qu'un fichier a été uploadé
            if 'excel_file' not in request.FILES:
                messages.error(request, "Aucun fichier sélectionné")
                return redirect('celica_web:test_list')
            
            excel_file = request.FILES['excel_file']
            module_id = request.POST.get('module_id')
            
            if not module_id:
                messages.error(request, "Veuillez sélectionner un module")
                return redirect('celica_web:test_list')
            
            module = get_object_or_404(Module, id=module_id)
            
            # Validation du fichier
            if excel_file.size > 10 * 1024 * 1024:  # 10 MB
                messages.error(request, "Le fichier est trop volumineux (maximum 10 MB)")
                return redirect('celica_web:test_list')
            
            # Vérifier l'extension
            if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                messages.error(request, "Format de fichier non supporté. Utilisez Excel (.xlsx, .xls)")
                return redirect('celica_web:test_list')
            
            # Sauvegarder temporairement le fichier
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in excel_file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
            
            try:
                # Valider la structure du fichier
                is_valid, validation_errors = ExcelImportValidator.validate_file_structure(tmp_file_path)
                
                if not is_valid:
                    messages.error(request, f"Structure de fichier invalide: {'; '.join(validation_errors)}")
                    return redirect('celica_web:test_list')
                
                # Importer le test
                import_service = ExcelImportService(request.user)
                success, test, messages_list = import_service.import_test_from_excel(tmp_file_path, module)
                
                if success:
                    # Afficher les messages de succès et d'avertissement
                    for msg in messages_list:
                        if "erreur" in msg.lower():
                            messages.error(request, msg)
                        elif "avertissement" in msg.lower() or "warning" in msg.lower():
                            messages.warning(request, msg)
                        else:
                            messages.info(request, msg)
                    
                    # Message de succès principal
                    summary = import_service.get_import_summary()
                    messages.success(
                        request, 
                        f"Test '{test.titre}' importé avec succès! "
                        f"{summary['imported_count']} question(s) créée(s)."
                    )
                    
                    return redirect('celica_web:test_form_edit', test_id=test.id)
                else:
                    # Afficher les erreurs
                    for error in messages_list:
                        messages.error(request, error)
                    return redirect('celica_web:test_list')
                    
            finally:
                # Nettoyer le fichier temporaire
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                    
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import: {str(e)}")
            logger.error(f"Erreur dans import_test_excel: {str(e)}")
            return redirect('celica_web:test_list')
    
    # GET request - afficher le formulaire d'import
    modules = Module.objects.filter(status='actif').order_by('intitule')
    return render(request, 'celicaweb/import_test_excel.html', {
        'modules': modules
    })

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def preview_excel_import(request):
    """Prévisualise l'import Excel sans sauvegarder"""
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                return JsonResponse({'error': 'Aucun fichier sélectionné'}, status=400)
            
            excel_file = request.FILES['excel_file']
            
            # Sauvegarder temporairement
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in excel_file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
            
            try:
                # Valider et lire les données
                is_valid, validation_errors = ExcelImportValidator.validate_file_structure(tmp_file_path)
                
                if not is_valid:
                    return JsonResponse({'error': f"Structure invalide: {'; '.join(validation_errors)}"}, status=400)
                
                # Lire les données pour prévisualisation
                import pandas as pd
                
                preview_data = {
                    'test_info': {},
                    'questions': []
                }
                
                # Lire l'onglet Test
                try:
                    test_df = pd.read_excel(tmp_file_path, sheet_name='Test', header=None)
                    for _, row in test_df.iterrows():
                        if pd.notna(row[0]) and pd.notna(row[1]):
                            key = str(row[0]).strip()
                            value = str(row[1]).strip()
                            preview_data['test_info'][key] = value
                except:
                    preview_data['test_info'] = {'Titre du test': 'Non spécifié'}
                
                # Lire l'onglet Questions
                try:
                    questions_df = pd.read_excel(tmp_file_path, sheet_name='Questions')
                    for index, row in questions_df.iterrows():
                        if pd.notna(row.get('Type', '')) and str(row['Type']).strip():
                            question_preview = {
                                'type': str(row.get('Type', '')).strip(),
                                'enonce': str(row.get('Énoncé', '')).strip(),
                                'niveau': str(row.get('Niveau', '')).strip(),
                                'points': float(row.get('Points', 0)),
                                'reponses': []
                            }
                            
                            # Prévisualiser les réponses
                            for i in range(1, 6):
                                reponse_key = f'Réponse {i}'
                                if reponse_key in row and pd.notna(row[reponse_key]):
                                    reponse_text = str(row[reponse_key]).strip()
                                    if reponse_text:
                                        est_correcte = '✓' in reponse_text
                                        question_preview['reponses'].append({
                                            'texte': reponse_text.replace('✓', '').strip(),
                                            'correcte': est_correcte
                                        })
                            
                            preview_data['questions'].append(question_preview)
                except Exception as e:
                    return JsonResponse({'error': f"Erreur lecture questions: {str(e)}"}, status=400)
                
                return JsonResponse(preview_data)
                
            finally:
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                    
        except Exception as e:
            return JsonResponse({'error': f"Erreur prévisualisation: {str(e)}"}, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def cours_list(request):
    utilisateur = request.user

    # Filtrage des cours selon le rôle
    if utilisateur.role == 'admin':
        cours = Cours.objects.select_related('module', 'instructeur').all()
    else:
        cours = Cours.objects.filter(instructeur=utilisateur)

    modules = Module.objects.all()
    module_id = request.GET.get('module')
    status = request.GET.get('status')

    if module_id:
        cours = cours.filter(module__id=module_id)
    if status:
        cours = cours.filter(status=status)

    context = {
        'cours': cours,
        'modules': modules,
        'selected_module': module_id,
        'selected_status': status,
    }
    return render(request, 'celicaweb/cours_list.html', context)

@csrf_exempt
@login_required
def security_violation(request):
    """Gère les violations de sécurité pendant les tests"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            violation = data.get('violation', 'Violation non spécifiée')
            violation_type = data.get('violation_type', 'unknown')
            url = data.get('url', '')
            
            print(f"🔒 VIOLATION REÇUE: {violation} ({violation_type}) par {request.user.email}")
            
            # Enregistrer la violation dans la base de données
            security_violation = SecurityViolation.objects.create(
                utilisateur=request.user,
                violation=violation,
                violation_type=violation_type,
                url=url,
                ip_address=request.META.get('REMOTE_ADDR', ''),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            print(f"✅ Violation enregistrée avec l'ID: {security_violation.id}")
            
            # METTRE À JOUR LE COMPTEUR DE VIOLATIONS DANS LA SESSION
            violation_count = data.get('violation_count', 0)
            
            # Identifier le test depuis l'URL
            test_id = None
            test = None
            if '/test/passer/' in url:
                url_parts = url.split('/test/passer/')
                if len(url_parts) > 1:
                    test_id_part = url_parts[1].split('/')[0]
                    if test_id_part.isdigit():
                        test_id = int(test_id_part)
                        test = Test.objects.get(id=test_id)
            
            # Mettre à jour le compteur dans la session
            if test_id:
                violation_count_key = f'violation_count_{test_id}_{request.user.id}'
                current_count = request.session.get(violation_count_key, 0)
                new_count = current_count + 1
                request.session[violation_count_key] = new_count
                request.session.modified = True
                print(f"📊 Compteur de violations mis à jour: {current_count} → {new_count}")
                
                # Si 3 violations atteintes, rediriger vers test_terminated
                if new_count >= 3:
                    print(f"🚨 LIMITE DE VIOLATIONS ATTEINTE ({new_count}/3)")
                    # Nettoyer les données de session
                    test_keys_to_clean = [key for key in request.session.keys() if key.startswith('test_') or key.startswith('violation_')]
                    for key in test_keys_to_clean:
                        del request.session[key]
                    request.session.modified = True
                    return JsonResponse({'status': 'terminated', 'redirect': '/test/terminated/'})
            
            # Envoyer des notifications pour TOUTES les violations
            try:
                # Récupérer l'administrateur
                admin = Utilisateur.objects.filter(role='admin', statut='actif').first()
                
                # Identifier le test depuis l'URL
                test_id = None
                test = None
                instructeur_test = None
                
                if '/test/passer/' in url:
                    url_parts = url.split('/test/passer/')
                    if len(url_parts) > 1:
                        test_id_part = url_parts[1].split('/')[0]
                        if test_id_part.isdigit():
                            test_id = int(test_id_part)
                            test = Test.objects.get(id=test_id)
                            instructeur_test = test.instructeur
                
                # Déterminer le type de notification selon le type de violation
                if violation_type == 'test_terminated':
                    print("🚨 VIOLATION DE TERMINAISON DE TEST DÉTECTÉE")
                    titre = "🚨 Test terminé automatiquement - Violations répétées"
                    priorite = 'critique'
                    
                    # Récupérer toutes les violations récentes pour ce test et cet utilisateur
                    violations_recentes = SecurityViolation.objects.filter(
                        utilisateur=request.user,
                        url__contains=f'/test/passer/{test_id}/',
                        timestamp__gte=timezone.now() - timezone.timedelta(hours=1)  # Dernière heure
                    ).order_by('timestamp')
                    
                    # Construire la liste des violations
                    violations_list = []
                    for i, v in enumerate(violations_recentes, 1):
                        violations_list.append(f"{i}. {v.violation_type.replace('_', ' ').title()} - {v.timestamp.strftime('%H:%M:%S')}")
                    
                    violations_text = "\n".join(violations_list) if violations_list else "Aucune violation détaillée disponible"
                    
                    message = f"""
🚨 TEST TERMINÉ AUTOMATIQUEMENT - VIOLATIONS RÉPÉTÉES

Apprenant: {request.user.get_full_name()} ({request.user.email})
Test: {test.titre if test else 'Non identifié'} (ID: {test_id if test_id else 'N/A'})
Nombre total de violations: {data.get('violation_count', 'N/A')}
Date/Heure de terminaison: {security_violation.timestamp.strftime('%d/%m/%Y %H:%M:%S')}

VIOLATIONS DÉTECTÉES:
{violations_text}

Le test a été automatiquement interrompu après 3 violations de sécurité.
L'apprenant a été redirigé vers la page de test terminé.

Actions recommandées:
• Contacter l'apprenant pour comprendre la situation
• Reprogrammer le test si nécessaire
• Consulter les détails des violations dans l'interface d'administration
                    """.strip()
                else:
                    print(f"🔔 VIOLATION DE SÉCURITÉ DÉTECTÉE: {violation_type}")
                    titre = f"⚠️ Violation de sécurité - {violation_type.replace('_', ' ').title()}"
                    priorite = 'haute'
                    
                    message = f"""
⚠️ VIOLATION DE SÉCURITÉ DÉTECTÉE

Apprenant: {request.user.get_full_name()} ({request.user.email})
Test: {test.titre if test else 'Non identifié'} (ID: {test_id if test_id else 'N/A'})
Type de violation: {violation_type}
Action détectée: {violation}
Date/Heure: {security_violation.timestamp.strftime('%d/%m/%Y %H:%M:%S')}
URL: {url}

Cette violation a été détectée pendant la passation du test.
L'apprenant a été averti et le compteur de violations a été incrémenté.

Actions recommandées:
• Surveiller l'activité de l'apprenant
• Contacter l'apprenant si les violations se répètent
• Consulter les détails de la violation
                    """.strip()
                
                # Envoyer à l'instructeur du test
                if instructeur_test:
                    Notification.objects.create(
                        titre=titre,
                        message=message,
                        type_notice='urgence',
                        priorite=priorite,
                        utilisateur=instructeur_test,
                        test=test,
                        date_expiration=timezone.now() + timezone.timedelta(days=7)
                    )
                    print(f"📧 Notification envoyée à l'instructeur: {instructeur_test.email}")
                
                # Envoyer à l'administrateur
                if admin:
                    Notification.objects.create(
                        titre=titre,
                        message=message,
                        type_notice='urgence',
                        priorite=priorite,
                        utilisateur=admin,
                        test=test,
                        date_expiration=timezone.now() + timezone.timedelta(days=7)
                    )
                    print(f"📧 Notification envoyée à l'administrateur: {admin.email}")
                
                print(f"✅ Notifications envoyées (Instructeur: {instructeur_test is not None}, Admin: {admin is not None})")
                
            except Exception as e:
                print(f"❌ Erreur lors de l'envoi des notifications: {str(e)}")
            
            # Retourner le nouveau compteur dans la réponse
            return JsonResponse({
                'status': 'success', 
                'id': security_violation.id,
                'violation_count': new_count
            })
        except Exception as e:
            print(f"❌ Erreur lors de l'enregistrement de la violation: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    # Gestion des requêtes GET (AJAX pour récupérer le compteur)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Identifier le test depuis l'URL de référence
        referer = request.META.get('HTTP_REFERER', '')
        test_id = None
        if '/test/passer/' in referer:
            url_parts = referer.split('/test/passer/')
            if len(url_parts) > 1:
                test_id_part = url_parts[1].split('/')[0]
                if test_id_part.isdigit():
                    test_id = int(test_id_part)
        
        if test_id:
            violation_count_key = f'violation_count_{test_id}_{request.user.id}'
            violation_count = request.session.get(violation_count_key, 0)
            return JsonResponse({'violation_count': violation_count})
        else:
            return JsonResponse({'violation_count': 0})
    
    # Page d'affichage des violations
    violations = SecurityViolation.objects.filter(utilisateur=request.user).order_by('-timestamp')[:10]
    print(f"📊 Affichage de {violations.count()} violations pour {request.user.email}")
    return render(request, 'celicaweb/security_violation.html', {
        'violations': violations,
        'debug': settings.DEBUG  # Activer le mode debug pour les tests
    })

@login_required
def test_terminated(request):
    """Vue affichée quand un test est terminé automatiquement"""
    # Récupérer les violations récentes de l'utilisateur
    violations_recentes = SecurityViolation.objects.filter(
        utilisateur=request.user,
        timestamp__gte=timezone.now() - timezone.timedelta(hours=1)  # Dernière heure
    ).order_by('-timestamp')[:10]
    
    # Récupérer l'ID du test depuis la session
    test_id = request.session.get('current_test_id')
    
    if test_id:
        try:
            # Récupérer le test
            test = Test.objects.get(id=test_id)
            
            # Vérifier si un résultat existe déjà pour ce test et cet apprenant
            try:
                resultat = Resultat.objects.get(test=test, apprenant=request.user)
                # Si le résultat existe déjà, le marquer comme terminé par violation
                resultat.score = 0
                resultat.commentaires = 'Test terminé automatiquement en raison de violations de sécurité'
                resultat.save()
                print(f"✅ Résultat existant mis à jour pour le test {test_id}")
            except Resultat.DoesNotExist:
                # Créer un nouveau résultat "terminé par violation"
                resultat = Resultat.objects.create(
                    test=test,
                    apprenant=request.user,
                    score=0,
                    commentaires='Test terminé automatiquement en raison de violations de sécurité',
                    date_passation=timezone.now()
                )
                print(f"✅ Nouveau résultat créé pour le test {test_id}")
            
        except Test.DoesNotExist:
            print(f"❌ Test {test_id} non trouvé")
        except Exception as e:
            print(f"❌ Erreur lors de la création du résultat: {str(e)}")
    
    # Nettoyer les données de session liées aux tests
    test_keys_to_clean = [key for key in request.session.keys() if key.startswith('test_') or key.startswith('violation_')]
    for key in test_keys_to_clean:
        del request.session[key]
    request.session.modified = True
    
    return render(request, 'celicaweb/test_terminated.html', {
        'message': 'Test terminé automatiquement en raison de violations de sécurité.',
        'violations': violations_recentes
    })





@login_required
@permission_required('celica_web.gerer_utilisateurs', raise_exception=True)
def utilisateur_list(request):
    # Récupérer le paramètre de recherche
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '')
    statut_filter = request.GET.get('statut', '')
    
    # Base queryset
    utilisateurs = Utilisateur.objects.all()
    
    # Appliquer les filtres
    if search_query:
        utilisateurs = utilisateurs.filter(
            Q(last_name__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(matricule__icontains=search_query)
        )
    
    if role_filter:
        utilisateurs = utilisateurs.filter(role=role_filter)
    
    if statut_filter:
        utilisateurs = utilisateurs.filter(statut=statut_filter)
    
    # Grouper les utilisateurs par statut
    utilisateurs_par_statut = {
        'actif': utilisateurs.filter(statut='actif').order_by('last_name', 'first_name'),
        'inactif': utilisateurs.filter(statut='inactif').order_by('last_name', 'first_name'),
        'suspendu': utilisateurs.filter(statut='suspendu').order_by('last_name', 'first_name'),
    }
    
    # Statistiques
    total_utilisateurs = utilisateurs.count()
    total_actifs = utilisateurs.filter(statut='actif').count()
    total_inactifs = utilisateurs.filter(statut='inactif').count()
    total_suspendus = utilisateurs.filter(statut='suspendu').count()
    
    context = {
        'utilisateurs_par_statut': utilisateurs_par_statut,
        'search_query': search_query,
        'role_filter': role_filter,
        'statut_filter': statut_filter,
        'stats': {
            'total': total_utilisateurs,
            'actifs': total_actifs,
            'inactifs': total_inactifs,
            'suspendus': total_suspendus,
        },
        'roles': Utilisateur.ROLES,
        'statuts': Utilisateur.STATUTS,
    }
    
    return render(request, 'celicaweb/utilisateur_list.html', context)

def get_client_ip(request):
    """Récupère l'adresse IP du client"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@csrf_exempt
@login_required
def log_test_event(request):
    """Vue pour enregistrer les événements de test"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            event_type = data.get('event_type')
            test_id = data.get('test_id')
            question_number = data.get('question_number')
            event_data = data.get('event_data', {})
            session_id = data.get('session_id')
            duration = data.get('duration')
            
            # Récupérer le test
            try:
                test = Test.objects.get(id=test_id)
            except Test.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Test non trouvé'}, status=404)
            
            # Enregistrer l'événement
            event = TestEventLog.log_event(
                test=test,
                utilisateur=request.user,
                event_type=event_type,
                event_data=event_data,
                question_number=question_number,
                session_id=session_id,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                duration=duration
            )
            
            print(f"📝 Événement enregistré: {event_type} - Test {test_id} - Question {question_number}")
            
            return JsonResponse({'status': 'success', 'message': 'Événement enregistré'})
            
        except Exception as e:
            print(f"❌ Erreur lors de l'enregistrement de l'événement: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'}, status=405)

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def test_event_logs(request, test_id):
    """Vue pour consulter les journaux d'événements d'un test"""
    try:
        test = Test.objects.get(id=test_id)
        
        # Vérifier que l'utilisateur a accès à ce test
        if request.user.role == 'instructeur' and test.instructeur != request.user:
            messages.error(request, "Vous n'avez pas accès à ce test.")
            return redirect('celica_web:test_list')
        
        # Récupérer les événements
        events = TestEventLog.objects.filter(test=test).select_related('utilisateur').order_by('-timestamp')
        
        # Filtres
        event_type_filter = request.GET.get('event_type')
        user_filter = request.GET.get('user')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if event_type_filter:
            events = events.filter(event_type=event_type_filter)
        
        if user_filter:
            events = events.filter(utilisateur__email__icontains=user_filter)
        
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                events = events.filter(timestamp__date__gte=date_from_obj.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                events = events.filter(timestamp__date__lte=date_to_obj.date())
            except ValueError:
                pass
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(events, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'test': test,
            'events': page_obj,
            'event_types': TestEventLog.EVENT_TYPES,
            'filters': {
                'event_type': event_type_filter,
                'user': user_filter,
                'date_from': date_from,
                'date_to': date_to,
            }
        }
        
        return render(request, 'celicaweb/test_event_logs.html', context)
        
    except Test.DoesNotExist:
        messages.error(request, "Test non trouvé.")
        return redirect('celica_web:test_list')

@login_required
@permission_required('celica_web.gerer_tests', raise_exception=True)
def user_test_events(request, user_id, test_id):
    """Vue pour consulter les événements d'un utilisateur spécifique pour un test"""
    try:
        user = Utilisateur.objects.get(id=user_id)
        test = Test.objects.get(id=test_id)
        
        # Vérifier les permissions
        if request.user.role == 'instructeur' and test.instructeur != request.user:
            messages.error(request, "Vous n'avez pas accès à ce test.")
            return redirect('celica_web:test_list')
        
        # Récupérer les événements de l'utilisateur pour ce test
        events = TestEventLog.objects.filter(
            test=test,
            utilisateur=user
        ).select_related('utilisateur').order_by('-timestamp')
        
        # Statistiques
        total_events = events.count()
        violation_events = events.filter(event_type__in=['violation_detected', 'copy_attempt', 'paste_attempt', 'right_click', 'keyboard_shortcut', 'dev_tools', 'screenshot_attempt']).count()
        focus_events = events.filter(event_type__in=['page_focus', 'page_blur']).count()
        question_events = events.filter(event_type__in=['question_view', 'question_answer', 'question_change']).count()
        
        # Durée totale du test (si disponible)
        test_start = events.filter(event_type='test_start').first()
        test_end = events.filter(event_type='test_end').first()
        
        context = {
            'user': user,
            'test': test,
            'events': events,
            'stats': {
                'total_events': total_events,
                'violation_events': violation_events,
                'focus_events': focus_events,
                'question_events': question_events,
                'test_start': test_start,
                'test_end': test_end,
            }
        }
        
        return render(request, 'celicaweb/user_test_events.html', context)
        
    except (Utilisateur.DoesNotExist, Test.DoesNotExist):
        messages.error(request, "Utilisateur ou test non trouvé.")
        return redirect('celica_web:test_list')

@login_required
@user_passes_test(is_superuser)
def journalisation_surveillance(request):
    """Vue pour afficher les détails de journalisation et surveillance"""
    from django.core.paginator import Paginator
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # Filtres
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    event_type_filter = request.GET.get('event_type')
    violation_type_filter = request.GET.get('violation_type')
    user_filter = request.GET.get('user')
    test_filter = request.GET.get('test')
    
    # Récupérer les événements de test
    events = TestEventLog.objects.select_related('test', 'utilisateur').all()
    
    # Récupérer les violations de sécurité
    violations = SecurityViolation.objects.select_related('utilisateur').all()
    
    # Appliquer les filtres pour les événements
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            events = events.filter(timestamp__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            events = events.filter(timestamp__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    
    if user_filter:
        events = events.filter(utilisateur__email__icontains=user_filter)
    
    if test_filter:
        events = events.filter(test__titre__icontains=test_filter)
    
    # Appliquer les filtres pour les violations (pas de filtre par test car pas de relation directe)
    violations_filtered = violations
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            violations_filtered = violations_filtered.filter(timestamp__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            violations_filtered = violations_filtered.filter(timestamp__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if violation_type_filter:
        violations_filtered = violations_filtered.filter(violation_type=violation_type_filter)
    
    if user_filter:
        violations_filtered = violations_filtered.filter(utilisateur__email__icontains=user_filter)
    
    # Statistiques générales
    now = timezone.now()
    last_30_days = now - timedelta(days=30)
    last_7_days = now - timedelta(days=7)
    
    stats = {
        'total_events': events.count(),
        'total_violations': violations_filtered.count(),
        'events_30_days': events.filter(timestamp__gte=last_30_days).count(),
        'violations_30_days': violations_filtered.filter(timestamp__gte=last_30_days).count(),
        'events_7_days': events.filter(timestamp__gte=last_7_days).count(),
        'violations_7_days': violations_filtered.filter(timestamp__gte=last_7_days).count(),
    }
    
    # Statistiques par type
    event_types_stats = events.values('event_type').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    violation_types_stats = violations_filtered.values('violation_type').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Tests les plus surveillés
    tests_most_monitored = events.values('test__titre').annotate(
        event_count=Count('id')
    ).order_by('-event_count')[:10]
    
    # Utilisateurs avec le plus de violations
    users_most_violations = violations_filtered.values('utilisateur__email').annotate(
        violation_count=Count('id')
    ).order_by('-violation_count')[:10]
    
    # Pagination pour les événements
    events = events.order_by('-timestamp')
    paginator_events = Paginator(events, 50)
    page_number_events = request.GET.get('page_events')
    events_page = paginator_events.get_page(page_number_events)
    
    # Pagination pour les violations
    violations_filtered = violations_filtered.order_by('-timestamp')
    paginator_violations = Paginator(violations_filtered, 50)
    page_number_violations = request.GET.get('page_violations')
    violations_page = paginator_violations.get_page(page_number_violations)
    
    # Options pour les filtres
    event_types = TestEventLog.EVENT_TYPES
    violation_types = SecurityViolation.VIOLATION_TYPES
    tests = Test.objects.all().order_by('titre')
    users = Utilisateur.objects.all().order_by('email')
    
    context = {
        'events_page': events_page,
        'violations_page': violations_page,
        'stats': stats,
        'event_types_stats': event_types_stats,
        'violation_types_stats': violation_types_stats,
        'tests_most_monitored': tests_most_monitored,
        'users_most_violations': users_most_violations,
        'event_types': event_types,
        'violation_types': violation_types,
        'tests': tests,
        'users': users,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'event_type': event_type_filter,
            'violation_type': violation_type_filter,
            'user': user_filter,
            'test': test_filter,
        }
    }
    
    return render(request, 'celicaweb/journalisation_surveillance.html', context)

@login_required
@user_passes_test(is_superuser)
def supprimer_violation(request, violation_id):
    """Vue pour supprimer une violation de sécurité"""
    try:
        violation = SecurityViolation.objects.get(id=violation_id)
        
        if request.method == 'POST':
            # Supprimer la violation
            violation.delete()
            messages.success(request, "Violation de sécurité supprimée avec succès.")
            return redirect('celica_web:journalisation_surveillance')
        
        # Afficher la page de confirmation
        context = {
            'violation': violation,
        }
        return render(request, 'celicaweb/confirm_delete_violation.html', context)
        
    except SecurityViolation.DoesNotExist:
        messages.error(request, "Violation de sécurité non trouvée.")
        return redirect('celica_web:journalisation_surveillance')

@login_required
@user_passes_test(is_superuser)
def supprimer_toutes_violations(request):
    """Vue pour supprimer toutes les violations de sécurité"""
    if request.method == 'POST':
        # Récupérer les paramètres de filtrage
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        violation_type = request.POST.get('violation_type')
        user_filter = request.POST.get('user')
        
        # Construire le queryset de base
        violations = SecurityViolation.objects.all()
        
        # Appliquer les mêmes filtres que dans la vue principale
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                violations = violations.filter(timestamp__date__gte=date_from_obj.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                violations = violations.filter(timestamp__date__lte=date_to_obj.date())
            except ValueError:
                pass
        
        if violation_type:
            violations = violations.filter(violation_type=violation_type)
        
        if user_filter:
            violations = violations.filter(utilisateur__email__icontains=user_filter)
        
        # Compter les violations avant suppression
        count = violations.count()
        
        # Supprimer les violations
        violations.delete()
        
        messages.success(request, f"{count} violation(s) de sécurité supprimée(s) avec succès.")
        return redirect('celica_web:journalisation_surveillance')
    
    # Gestion des requêtes GET avec confirmation
    elif request.method == 'GET' and request.GET.get('confirm') == 'yes':
        # Supprimer TOUTES les violations sans filtres
        count = SecurityViolation.objects.count()
        SecurityViolation.objects.all().delete()
        
        messages.success(request, f"{count} violation(s) de sécurité supprimée(s) avec succès.")
        return redirect('celica_web:journalisation_surveillance')
    
    # Si ce n'est pas une requête POST ou GET avec confirmation, rediriger
    return redirect('celica_web:journalisation_surveillance')

@login_required
@user_passes_test(is_superuser)
def supprimer_tous_evenements(request):
    """Vue pour supprimer tous les événements de test"""
    if request.method == 'POST':
        # Récupérer les paramètres de filtrage
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        event_type = request.POST.get('event_type')
        user_filter = request.POST.get('user')
        test_filter = request.POST.get('test')
        
        # Construire le queryset de base
        events = TestEventLog.objects.all()
        
        # Appliquer les mêmes filtres que dans la vue principale
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                events = events.filter(timestamp__date__gte=date_from_obj.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                events = events.filter(timestamp__date__lte=date_to_obj.date())
            except ValueError:
                pass
        
        if event_type:
            events = events.filter(event_type=event_type)
        
        if user_filter:
            events = events.filter(utilisateur__email__icontains=user_filter)
        
        if test_filter:
            events = events.filter(test__titre__icontains=test_filter)
        
        # Compter les événements avant suppression
        count = events.count()
        
        # Supprimer les événements
        events.delete()
        
        messages.success(request, f"{count} événement(s) de test supprimé(s) avec succès.")
        return redirect('celica_web:journalisation_surveillance')
    
    # Gestion des requêtes GET avec confirmation
    elif request.method == 'GET' and request.GET.get('confirm') == 'yes':
        # Supprimer TOUS les événements sans filtres
        count = TestEventLog.objects.count()
        TestEventLog.objects.all().delete()
        
        messages.success(request, f"{count} événement(s) de test supprimé(s) avec succès.")
        return redirect('celica_web:journalisation_surveillance')
    
    # Si ce n'est pas une requête POST ou GET avec confirmation, rediriger
    return redirect('celica_web:journalisation_surveillance')

# Fonction utilitaire pour vérifier le rôle apprenant
def check_apprenant_role(user):
    return hasattr(user, 'role') and user.role == 'apprenant'

@login_required
@permission_required('celica_web.consulter_resultats', raise_exception=True)
def apprenant_dashboard(request):
    if not check_apprenant_role(request.user):
        messages.error(request, "Accès refusé : vous n'êtes pas un apprenant.")
        return redirect('celica_web:visitor_index')
    
    try:
        from django.db.models import Avg, Count, Q
        from datetime import timedelta
        
        now = timezone.now()
        today = now.date()
        
        # === STATISTIQUES PERSONNELLES ===
        
        # Mes groupes
        mes_groupes = request.user.groupes.all()
        mes_groupes_list = list(mes_groupes)
        
        # Mes cours - tous les cours des modules de mes groupes
        mes_cours = Cours.objects.filter(
            module__groupes__groupes_apprenant=request.user,
            status='actif'
        ).distinct()
        
        # Tests disponibles - non encore passés avec vérification des plannings
        tests_disponibles = Test.objects.filter(
            actif=True,
            plannings__groupe__in=mes_groupes_list,
            plannings__date_debut__lte=now,
            plannings__date_fin__gte=now,
            plannings__statut__in=['planifie', 'en_cours']
        ).exclude(resultats__apprenant=request.user).distinct()
        
        # Mes résultats avec calculs
        mes_resultats = Resultat.objects.filter(apprenant=request.user).order_by('-date_passation')
        
        # Calcul de la moyenne générale (en pourcentage)
        if mes_resultats.exists():
            # Calculer la moyenne des pourcentages comme dans mes_resultats
            scores_pourcentages = []
            for resultat in mes_resultats:
                pourcentage = (resultat.note_sur_20 / 20) * 100 if resultat.note_sur_20 else 0
                scores_pourcentages.append(pourcentage)
            moyenne_generale = sum(scores_pourcentages) / len(scores_pourcentages) if scores_pourcentages else 0
            moyenne_generale = round(moyenne_generale, 1)
        else:
            moyenne_generale = 0
        
        # Planning d'aujourd'hui et de la semaine
        plannings_aujourdhui = Planning.objects.filter(
            groupe__in=mes_groupes_list,
            date_debut__date=today
        ).distinct().order_by('date_debut')
        
        # Plannings cette semaine
        semaine_debut = today - timedelta(days=today.weekday())
        semaine_fin = semaine_debut + timedelta(days=6)
        plannings_semaine = Planning.objects.filter(
            groupe__in=mes_groupes_list,
            date_debut__date__range=[semaine_debut, semaine_fin]
        ).distinct().order_by('date_debut')
        
        # Notifications non lues
        notifications_non_lues = Notification.objects.filter(
            utilisateur=request.user,
            est_lue=False
        ).order_by('-date_envoi')[:5]
        
        # Toutes les notifications récentes
        toutes_notifications = Notification.objects.filter(
            utilisateur=request.user
        ).order_by('-date_envoi')[:10]
        
        # === STATISTIQUES POUR LES CARTES ===
        stats = {
            'total_cours': mes_cours.count(),
            'tests_passes': mes_resultats.count(),
            'moyenne_generale': moyenne_generale,
            'plannings_semaine': plannings_semaine.count(),
            'notifications_non_lues': notifications_non_lues.count(),
            'tests_disponibles': tests_disponibles.count(),
        }
        
        # === PROGRESSION ET COURS RÉCENTS ===
        # Cours avec progression fictive (à adapter selon votre logique)
        cours_avec_progression = []
        for cours in mes_cours[:6]:  # Limiter à 6 cours
            # Logique de progression - vous pouvez l'adapter selon vos besoins
            progression = min(100, (cours.id * 23) % 101)  # Progression fictive
            cours_avec_progression.append({
                'cours': cours,
                'progression': progression
            })
        
        # === RÉSULTATS RÉCENTS AVEC BADGES ===
        resultats_recents = []
        for resultat in mes_resultats[:5]:
            # Déterminer le badge selon le score
            if resultat.score >= 16:
                badge_class = 'success'
                badge_text = 'Excellent'
            elif resultat.score >= 14:
                badge_class = 'primary'
                badge_text = 'Très bien'
            elif resultat.score >= 12:
                badge_class = 'info'
                badge_text = 'Bien'
            elif resultat.score >= 10:
                badge_class = 'warning'
                badge_text = 'Passable'
            else:
                badge_class = 'danger'
                badge_text = 'Insuffisant'
            
            resultats_recents.append({
                'resultat': resultat,
                'badge_class': badge_class,
                'badge_text': badge_text
            })
        
        # === ACCÈS RAPIDE - ACTIONS PRIORITAIRES ===
        actions_prioritaires = []
        
        # Tests urgents à passer
        if tests_disponibles.exists():
            actions_prioritaires.append({
                'titre': f"{tests_disponibles.count()} test(s) disponible(s)",
                'description': "Des tests sont disponibles pour vous",
                'icon': 'fas fa-clipboard-list',
                'color': 'primary',
                'url': '/test/list/',
                'urgent': tests_disponibles.count() > 2
            })
        
        # Notifications non lues
        if notifications_non_lues.exists():
            actions_prioritaires.append({
                'titre': f"{notifications_non_lues.count()} notification(s)",
                'description': "Nouvelles notifications à consulter",
                'icon': 'fas fa-bell',
                'color': 'warning',
                'url': '/notifications/',
                'urgent': notifications_non_lues.count() > 3
            })
        
        # Plannings du jour
        if plannings_aujourdhui.exists():
            actions_prioritaires.append({
                'titre': f"{plannings_aujourdhui.count()} session(s) aujourd'hui",
                'description': "Vous avez des sessions planifiées aujourd'hui",
                'icon': 'fas fa-calendar-day',
                'color': 'info',
                'url': '/planning/',
                'urgent': True
            })
        
        # === MODULES AVEC CATÉGORIES AVIATION ===
        modules_aviation = Module.objects.filter(
            groupes__in=mes_groupes_list
        ).distinct()
        
        modules_par_categorie = {}
        for module in modules_aviation:
            categorie = module.categorie
            if categorie not in modules_par_categorie:
                modules_par_categorie[categorie] = []
            modules_par_categorie[categorie].append(module)
        
        context = {
            'user': request.user,
            'stats': stats,
            'mes_cours': cours_avec_progression,
            'tests_disponibles': tests_disponibles,  # Afficher tous les tests disponibles
            'resultats_recents': resultats_recents,
            'plannings_aujourdhui': plannings_aujourdhui,
            'plannings_semaine': plannings_semaine[:5],  # Limiter l'affichage
            'notifications': toutes_notifications,
            'notifications_non_lues': notifications_non_lues,
            'actions_prioritaires': actions_prioritaires,
            'modules_par_categorie': modules_par_categorie,
            'mes_groupes': mes_groupes,
            'now': now,
            'today': today,
        }
        
        return render(request, 'celicaweb/apprenant_dashboard.html', context)
        
    except Exception as e:
        messages.error(request, f"Erreur lors du chargement du tableau de bord: {str(e)}")
        return render(request, 'celicaweb/apprenant_dashboard.html', {
            'user': request.user,
            'stats': {
                'total_cours': 0, 'tests_passes': 0, 'moyenne_generale': 0,
                'plannings_semaine': 0, 'notifications_non_lues': 0, 'tests_disponibles': 0
            },
            'mes_cours': [], 'tests_disponibles': [], 'resultats_recents': [],
            'plannings_aujourdhui': [], 'plannings_semaine': [], 'notifications': [],
            'notifications_non_lues': [], 'actions_prioritaires': [],
            'modules_par_categorie': {}, 'mes_groupes': [],
            'now': timezone.now(), 'today': timezone.now().date()
        })

@login_required
@permission_required('celica_web.gerer_cours', raise_exception=True)
def instructeur_dashboard(request):
    utilisateur = request.user
    logger = logging.getLogger(__name__)
    logger.debug(f"\n=== DÉBUT DIAGNOSTIC DASHBOARD pour {utilisateur.email} ===")

    # Vérifier le rôle de l'instructeur
    if utilisateur.role != 'instructeur':
        messages.error(request, "Accès refusé : vous n'êtes pas un instructeur.")
        return redirect('celica_web:visitor_index')

    # Récupérer le module sélectionné
    module_id = request.GET.get('module')
    module_filter = {'module__id': module_id} if module_id else {}

    # === RÉCUPÉRATION DES DONNÉES D'AFFICHAGE ===
    # Modules (où il est instructeur principal ou associé)
    modules_qs = Module.objects.filter(
        models.Q(instructeur_principal=utilisateur) | models.Q(groupes__instructeurs=utilisateur)
    ).distinct()
    # Cours
    cours_qs = Cours.objects.filter(instructeur=utilisateur, **module_filter)
    # Tests
    tests_qs = Test.objects.filter(instructeur=utilisateur, **module_filter)
    # Questions
    questions_qs = Question.objects.filter(instructeur=utilisateur, **module_filter)
    # Résultats
    resultats_qs = Resultat.objects.filter(test__instructeur=utilisateur, **module_filter)
    # Plannings (où il est responsable ou liés à ses modules/tests)
    plannings_qs = Planning.objects.filter(
        models.Q(instructeur_responsable=utilisateur) |
        models.Q(test__instructeur=utilisateur) |
        models.Q(test__module__in=modules_qs)
    ).distinct()

    # Groupes (où il est instructeur associé)
    groupes_qs = Groupe.objects.filter(instructeurs=utilisateur).distinct()

    # === COMPTAGE SYNCHRONISÉ ===
    stats_cleaned = {
        'total_cours': cours_qs.count(),
        'total_tests': tests_qs.count(),
        'total_questions': questions_qs.count(),
        'total_resultats': resultats_qs.count(),
        'total_plannings': plannings_qs.count(),
        'total_modules': modules_qs.count(),
        'total_groupes': groupes_qs.count(),
    }

    # === STATISTIQUES POUR LES GRAPHIQUES ===
    from django.db.models import Avg, Count
    import json
    from django.core.serializers.json import DjangoJSONEncoder

    stats_labels = []
    stats_data = []
    stats_count = []

    tests_stats = (
        Test.objects.filter(instructeur=utilisateur)
        .annotate(
            avg_score=Avg('resultats__score'),
            count_results=Count('resultats')
        )
        .filter(count_results__gt=0)
        .order_by('-avg_score')[:5]
    )

    for test in tests_stats:
        titre_court = test.titre[:15] + "..." if len(test.titre) > 15 else test.titre
        stats_labels.append(titre_court)
        stats_data.append(round(test.avg_score or 0, 1))
        stats_count.append(test.count_results)

    context = {
        'user': utilisateur,
        'stats': stats_cleaned,
        'cours': list(cours_qs.select_related('module').order_by('-created_at')[:10]),
        'tests': list(tests_qs.select_related('module').order_by('-date_creation')[:10]),
        'modules': list(modules_qs),
        'questions': list(questions_qs.select_related('module').order_by('-date_creation')[:10]),
        'resultats': list(resultats_qs.select_related('test', 'apprenant').order_by('-date_passation')[:10]),
        'plannings': list(plannings_qs.select_related('groupe', 'test').order_by('-date_debut')[:10]),
        'notifications': list(Notification.objects.filter(utilisateur=utilisateur, est_lue=False).order_by('-date_envoi')[:5]),
        'stats_labels': json.dumps(stats_labels, cls=DjangoJSONEncoder),
        'stats_data': json.dumps(stats_data, cls=DjangoJSONEncoder),
        'stats_count': json.dumps(stats_count, cls=DjangoJSONEncoder),
        'selected_module': module_id,
        'selected_status': request.GET.get('status'),
    }

    return render(request, 'celicaweb/instructeur_dashboard.html', context)

@login_required
@permission_required('celica_web.consulter_plannings', raise_exception=True)
def apprenant_plannings(request):
    if not check_apprenant_role(request.user):
        messages.error(request, "Accès refusé : vous n'êtes pas un apprenant.")
        return redirect('celica_web:visitor_index')
    
    from django.db.models import Q
    now = timezone.now()
    groupes = request.user.groupes.all()
    
    if not groupes.exists():
        messages.warning(request, "Vous n'êtes assigné à aucun groupe. Contactez votre instructeur pour voir les plannings.")
        plannings = Planning.objects.none()
    else:
        # Filtrer les plannings via le champ FK 'groupe' uniquement
        plannings = Planning.objects.filter(
            Q(groupe__in=groupes),
            Q(statut__in=['planifie', 'en_cours', 'termine'])
        ).distinct()
    
    # Ajouter des filtres
    module_id = request.GET.get('module')
    if module_id:
        plannings = plannings.filter(test__module__id=module_id)
    
    modules = Module.objects.filter(
        tests__plannings__groupe__in=groupes
    ).distinct()
    return render(request, 'celicaweb/apprenant_plannings.html', {
        'plannings': plannings,
        'modules': modules,
        'selected_module': module_id,
        'now': now
    })

@login_required
@permission_required('celica_web.gerer_plannings', raise_exception=True)
def consulter_plannings_instructeur(request):
    utilisateur = request.user
    
    # Filtrage des plannings selon le rôle
    if utilisateur.role == 'admin':
        plannings = Planning.objects.all()
        modules = Module.objects.all()
        groupes = Groupe.objects.all()
    else:
        # Les instructeurs voient uniquement les plannings de leurs tests ou dont ils sont responsables
        plannings = Planning.objects.filter(
            models.Q(test__instructeur=utilisateur) |
            models.Q(instructeur_responsable=utilisateur)
        ).distinct()
        modules = Module.objects.filter(tests__instructeur=utilisateur).distinct()
        groupes = Groupe.objects.filter(plannings__test__instructeur=utilisateur).distinct()
    
    # Ajouter des filtres
    module_id = request.GET.get('module')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    groupe_id = request.GET.get('groupe')
    type_test = request.GET.get('type_test')
    
    if module_id:
        plannings = plannings.filter(test__module__id=module_id)
    if date_debut:
        plannings = plannings.filter(date_debut__gte=date_debut)
    if date_fin:
        plannings = plannings.filter(date_fin__lte=date_fin)
    if groupe_id:
        plannings = plannings.filter(groupe_id=groupe_id)
    if type_test:
        plannings = plannings.filter(test__type_test=type_test)
    return render(request, 'celicaweb/consulter_plannings_instructeurs.html', {
        'plannings': plannings,
        'modules': modules,
        'groupes': groupes,
        'type_tests': type_test,
        'selected_module': module_id,
        'selected_date_debut': date_debut,
        'selected_date_fin': date_fin,
        'selected_groupe': groupe_id,
        'selected_type_test': type_test
    })

@login_required
@permission_required('celica_web.gerer_questions', raise_exception=True)
def supprimer_question(request, question_id):
    question = get_object_or_404(Question, id=question_id, instructeur=request.user)
    # Vérifier si la question est utilisée dans un test actif
    now = timezone.now()
    active_tests = Test.objects.filter(questions=question)
    for test in active_tests:
        active_plannings = Planning.objects.filter(test=test, date_debut__lte=now, date_fin__gte=now, statut__in=['planifie', 'en_cours'])
        if active_plannings.exists():
            messages.error(request, "Cette question est utilisée dans un test actif et ne peut pas être supprimée.")
            return redirect('celica_web:question_list')
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, "Question supprimée avec succès.")
        return redirect('celica_web:question_list')
    
    return render(request, 'celicaweb/confirm_delete.html', {'question': question})